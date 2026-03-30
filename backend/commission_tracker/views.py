import csv
import io
from decimal import Decimal
from django.db.models import Q, Sum, Count
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from core.permissions import require_auth, require_permission
from core.status_history import record_status_change
from .models import CommissionStudent, CommissionEntry, StudentProvider, CommissionTerm
from core.pagination import StandardPagination
from core.field_permissions import filter_fields, filter_fields_list
from .services import calculate_entry, compute_master_from_entries, num, round2


def _intake_sort_key(intake_str):
    if not intake_str:
        return (0, 0)
    parts = intake_str.strip().upper().split()
    year = 0
    term = 0
    for p in parts:
        if p.isdigit() and len(p) == 4:
            year = int(p)
        elif p.startswith('T') and len(p) == 2 and p[1:].isdigit():
            term = int(p[1:])
    return (year, term)


TERMINAL_STATUSES = {'withdrawn', 'complete'}


def get_excluded_student_ids_for_year(target_year):
    prior_terms = list(
        CommissionTerm.objects.filter(year__lt=target_year).order_by('-year', '-sort_order')
    )
    if not prior_terms:
        return set()
    prior_term_names = [t.term_name for t in prior_terms]

    prior_entries = CommissionEntry.objects.filter(term_name__in=prior_term_names)
    prior_student_ids = set(prior_entries.values_list('commission_student_id', flat=True).distinct())
    if not prior_student_ids:
        return set()

    term_order = {t.term_name: (t.year, t.sort_order) for t in prior_terms}

    latest_status_per_provider = {}
    for e in prior_entries:
        sid = e.commission_student_id
        prov_key = (sid, e.student_provider_id)
        order = term_order.get(e.term_name, (0, 0))
        if prov_key not in latest_status_per_provider or order > latest_status_per_provider[prov_key][0]:
            latest_status_per_provider[prov_key] = (order, (e.student_status or 'Under Enquiry').lower())

    has_active_provider = set()
    all_terminal_students = set()
    for (sid, _prov_id), (order, status) in latest_status_per_provider.items():
        if status not in TERMINAL_STATUSES:
            has_active_provider.add(sid)

    target_year_str = str(target_year)
    students_with_target_year_provider = set(
        StudentProvider.objects.filter(
            commission_student_id__in=prior_student_ids,
            start_intake__icontains=target_year_str,
        ).values_list('commission_student_id', flat=True).distinct()
    )
    students_with_target_year_main = set(
        CommissionStudent.objects.filter(
            id__in=prior_student_ids,
            start_intake__icontains=target_year_str,
        ).values_list('id', flat=True).distinct()
    )
    has_active_provider |= students_with_target_year_provider
    has_active_provider |= students_with_target_year_main

    for sid in prior_student_ids:
        if sid not in has_active_provider:
            all_terminal_students.add(sid)

    return all_terminal_students


def _clean_id(val):
    if val and isinstance(val, str) and val.endswith('.0'):
        stripped = val[:-2]
        if stripped.isdigit():
            return stripped
    return val


def student_to_dict(s):
    return {
        'id': s.id, 'agentName': s.agent_name, 'studentId': _clean_id(s.student_id),
        'agentsicId': _clean_id(s.agentsic_id), 'studentName': s.student_name,
        'provider': s.provider, 'country': s.country, 'startIntake': s.start_intake,
        'courseLevel': s.course_level, 'courseName': s.course_name,
        'courseDurationYears': str(s.course_duration_years) if s.course_duration_years is not None else None,
        'commissionRatePct': str(s.commission_rate_pct) if s.commission_rate_pct is not None else None,
        'gstRatePct': str(s.gst_rate_pct) if s.gst_rate_pct is not None else None,
        'gstApplicable': s.gst_applicable,
        'scholarshipType': s.scholarship_type,
        'scholarshipValue': str(s.scholarship_value) if s.scholarship_value is not None else '0',
        'status': s.status, 'notes': s.notes,
        'totalReceived': str(s.total_received) if s.total_received is not None else '0',
        'createdAt': s.created_at.isoformat() if s.created_at else None,
        'updatedAt': s.updated_at.isoformat() if s.updated_at else None,
    }


def entry_to_dict(e):
    return {
        'id': e.id, 'commissionStudentId': e.commission_student_id,
        'studentProviderId': e.student_provider_id, 'termName': e.term_name,
        'academicYear': e.academic_year,
        'feeGross': str(e.fee_gross) if e.fee_gross is not None else '0',
        'commissionRateAuto': str(e.commission_rate_auto) if e.commission_rate_auto is not None else None,
        'commissionRateOverridePct': str(e.commission_rate_override_pct) if e.commission_rate_override_pct is not None else None,
        'commissionRateUsedPct': str(e.commission_rate_used_pct) if e.commission_rate_used_pct is not None else None,
        'commissionAmount': str(e.commission_amount) if e.commission_amount is not None else '0',
        'bonus': str(e.bonus) if e.bonus is not None else '0',
        'gstAmount': str(e.gst_amount) if e.gst_amount is not None else '0',
        'totalAmount': str(e.total_amount) if e.total_amount is not None else '0',
        'paymentStatus': e.payment_status, 'paidDate': str(e.paid_date) if e.paid_date else None,
        'invoiceNo': e.invoice_no, 'paymentRef': e.payment_ref, 'notes': e.notes,
        'studentStatus': e.student_status, 'rateChangeWarning': e.rate_change_warning,
        'scholarshipTypeAuto': e.scholarship_type_auto,
        'scholarshipValueAuto': str(e.scholarship_value_auto) if e.scholarship_value_auto is not None else None,
        'scholarshipTypeOverride': e.scholarship_type_override,
        'scholarshipValueOverride': str(e.scholarship_value_override) if e.scholarship_value_override is not None else None,
        'scholarshipTypeUsed': e.scholarship_type_used,
        'scholarshipValueUsed': str(e.scholarship_value_used) if e.scholarship_value_used is not None else None,
        'scholarshipChangeWarning': e.scholarship_change_warning,
        'scholarshipAmount': str(e.scholarship_amount) if e.scholarship_amount is not None else '0',
        'feeAfterScholarship': str(e.fee_after_scholarship) if e.fee_after_scholarship is not None else '0',
        'createdAt': e.created_at.isoformat() if e.created_at else None,
        'updatedAt': e.updated_at.isoformat() if e.updated_at else None,
    }


def provider_to_dict(sp):
    return {
        'id': sp.id, 'commissionStudentId': sp.commission_student_id,
        'provider': sp.provider, 'studentId': _clean_id(sp.student_id), 'country': sp.country,
        'courseLevel': sp.course_level, 'courseName': sp.course_name,
        'courseDurationYears': str(sp.course_duration_years) if sp.course_duration_years is not None else None,
        'startIntake': sp.start_intake,
        'commissionRatePct': str(sp.commission_rate_pct) if sp.commission_rate_pct is not None else None,
        'gstRatePct': str(sp.gst_rate_pct) if sp.gst_rate_pct is not None else None,
        'gstApplicable': sp.gst_applicable,
        'scholarshipType': sp.scholarship_type,
        'scholarshipValue': str(sp.scholarship_value) if sp.scholarship_value is not None else '0',
        'status': sp.status, 'notes': sp.notes,
        'createdAt': sp.created_at.isoformat() if sp.created_at else None,
    }


def get_term_order():
    terms = CommissionTerm.objects.filter(is_active=True).order_by('sort_order')
    return [t.term_name for t in terms]


def sync_per_provider_notes(student, entries, term_order):
    main_entries = [e for e in entries if not e.student_provider_id]
    main_master = compute_master_from_entries(main_entries, term_order)
    student.notes = main_master['notes']

    provider_entries = {}
    for e in entries:
        if e.student_provider_id:
            provider_entries.setdefault(e.student_provider_id, []).append(e)

    for prov_id, prov_ents in provider_entries.items():
        prov_master = compute_master_from_entries(prov_ents, term_order)
        try:
            sp = StudentProvider.objects.get(id=prov_id)
            sp.notes = prov_master['notes']
            sp.save(update_fields=['notes'])
        except StudentProvider.DoesNotExist:
            pass


def recalculate_student(student, user_id=None):
    entries = list(CommissionEntry.objects.filter(commission_student_id=student.id))
    term_order = get_term_order()

    provider_ids = set(e.student_provider_id for e in entries if e.student_provider_id)
    providers_lookup = {sp.id: sp for sp in StudentProvider.objects.filter(id__in=provider_ids)} if provider_ids else {}

    for entry in entries:
        prov_config = providers_lookup.get(entry.student_provider_id) if entry.student_provider_id else None

        calc = calculate_entry(student, entry, prov_config)
        entry.commission_rate_auto = calc['commissionRateAuto']
        entry.commission_rate_used_pct = calc['commissionRateUsedPct']
        entry.commission_amount = calc['commissionAmount']
        entry.gst_amount = calc['gstAmount']
        entry.total_amount = calc['totalAmount']
        entry.rate_change_warning = calc['rateChangeWarning']
        entry.scholarship_type_auto = calc['scholarshipTypeAuto']
        entry.scholarship_value_auto = calc['scholarshipValueAuto']
        entry.scholarship_type_used = calc['scholarshipTypeUsed']
        entry.scholarship_value_used = calc['scholarshipValueUsed']
        entry.scholarship_change_warning = calc['scholarshipChangeWarning']
        entry.scholarship_amount = calc['scholarshipAmount']
        entry.fee_after_scholarship = calc['feeAfterScholarship']
        entry.save()

    old_status = student.status
    entries = list(CommissionEntry.objects.filter(commission_student_id=student.id))
    master = compute_master_from_entries(entries, term_order)
    student.status = master['status']
    student.total_received = master['totalReceived']
    sync_per_provider_notes(student, entries, term_order)
    student.save(update_fields=['status', 'notes', 'total_received'])
    record_status_change('commission_student', student.id, old_status, student.status, user_id, notes='Recalculated from entries')


class StudentsListView(APIView):
    pagination_class = StandardPagination

    @require_permission("commission_tracker.student.read")
    def get(self, request):
        qs = CommissionStudent.objects.all()
        agent = request.query_params.get('agent')
        prov = request.query_params.get('provider')
        country = request.query_params.get('country')
        status = request.query_params.get('status')
        search = request.query_params.get('search')
        year = request.query_params.get('year')
        exclude_year = request.query_params.get('excludeYear')

        if agent:
            agent_list = [a.strip() for a in agent.split(",") if a.strip()]
            if len(agent_list) == 1:
                qs = qs.filter(agent_name__icontains=agent_list[0])
            else:
                qs = qs.filter(agent_name__in=agent_list)
        if prov:
            prov_list = [p.strip() for p in prov.split(",") if p.strip()]
            if len(prov_list) == 1:
                qs = qs.filter(provider__icontains=prov_list[0])
            else:
                qs = qs.filter(provider__in=prov_list)
        if country:
            qs = qs.filter(country__iexact=country)
        if status:
            status_list = [s.strip() for s in status.split(",") if s.strip()]
            if len(status_list) == 1:
                qs = qs.filter(status=status_list[0])
            else:
                qs = qs.filter(status__in=status_list)
        if search:
            qs = qs.filter(Q(student_name__icontains=search) | Q(student_id__icontains=search) | Q(agentsic_id__icontains=search) | Q(agent_name__icontains=search))
        if year:
            qs = qs.filter(start_intake__icontains=year)

        if exclude_year:
            excluded_ids = get_excluded_student_ids_for_year(int(exclude_year))
            if excluded_ids:
                qs = qs.exclude(id__in=excluded_ids)

        students_list = sorted(qs, key=lambda s: _intake_sort_key(s.start_intake), reverse=True)
        user_perms = request.session.get('userPermissions', [])

        if request.query_params.get('pageSize') == 'all':
            data = filter_fields_list([student_to_dict(s) for s in students_list], 'commission_student', user_perms)
            return Response({'count': len(data), 'next': None, 'previous': None, 'results': data})

        from django.core.paginator import Paginator as DjPaginator
        page_num = int(request.query_params.get('page', 1))
        page_sz = int(request.query_params.get('pageSize', 50))
        djp = DjPaginator(students_list, page_sz)
        page_obj = djp.get_page(page_num)
        data = filter_fields_list([student_to_dict(s) for s in page_obj], 'commission_student', user_perms)
        return Response({'count': djp.count, 'next': None, 'previous': None, 'results': data})

    @require_permission("commission_tracker.student.add")
    def post(self, request):
        try:
            d = request.data
            user_id = request.session.get('userId')

            student_id_val = _clean_id(d.get('studentId', '').strip())
            provider_val = d.get('provider', '').strip()
            agentsic_id_val = _clean_id(d.get('agentsicId', '').strip())

            if student_id_val and provider_val:
                existing = CommissionStudent.objects.filter(
                    student_id=student_id_val, provider=provider_val
                ).first()
                if existing:
                    return Response({
                        'message': f'Student with ID "{student_id_val}" already exists for provider "{provider_val}"'
                    }, status=400)

            if agentsic_id_val:
                existing = CommissionStudent.objects.filter(agentsic_id=agentsic_id_val).first()
                if existing:
                    return Response({
                        'message': f'Student with Agentsic ID "{agentsic_id_val}" already exists'
                    }, status=400)

            s = CommissionStudent.objects.create(
                agent_name=d.get('agentName', ''),
                student_id=_clean_id(d.get('studentId')),
                agentsic_id=_clean_id(d.get('agentsicId')),
                student_name=d.get('studentName', ''),
                provider=d.get('provider', ''),
                country=d.get('country', 'AU'),
                start_intake=d.get('startIntake'),
                course_level=d.get('courseLevel'),
                course_name=d.get('courseName'),
                course_duration_years=d.get('courseDurationYears'),
                commission_rate_pct=d.get('commissionRatePct'),
                gst_rate_pct=d.get('gstRatePct', 10),
                gst_applicable=d.get('gstApplicable', 'Yes'),
                scholarship_type=d.get('scholarshipType', 'None'),
                scholarship_value=d.get('scholarshipValue', 0),
                status=d.get('status', 'Under Enquiry'),
                notes=d.get('notes'),
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            record_status_change('commission_student', s.id, None, s.status, user_id, notes='Student created')
            user_perms = request.session.get('userPermissions', [])
            return Response(filter_fields(student_to_dict(s), 'commission_student', user_perms))
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class StudentDetailView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request, student_id):
        try:
            s = CommissionStudent.objects.get(id=student_id)
        except CommissionStudent.DoesNotExist:
            return Response({'message': 'Student not found'}, status=404)
        user_perms = request.session.get('userPermissions', [])
        return Response(filter_fields(student_to_dict(s), 'commission_student', user_perms))

    @require_permission("commission_tracker.student.update")
    def patch(self, request, student_id):
        try:
            try:
                s = CommissionStudent.objects.get(id=student_id)
            except CommissionStudent.DoesNotExist:
                return Response({'message': 'Student not found'}, status=404)

            old_status = s.status
            field_map = {
                'agentName': 'agent_name', 'studentId': 'student_id', 'agentsicId': 'agentsic_id',
                'studentName': 'student_name', 'provider': 'provider', 'country': 'country',
                'startIntake': 'start_intake', 'courseLevel': 'course_level', 'courseName': 'course_name',
                'courseDurationYears': 'course_duration_years', 'commissionRatePct': 'commission_rate_pct',
                'gstRatePct': 'gst_rate_pct', 'gstApplicable': 'gst_applicable',
                'scholarshipType': 'scholarship_type', 'scholarshipValue': 'scholarship_value',
                'status': 'status', 'notes': 'notes',
            }
            decimal_fields = {'commission_rate_pct', 'gst_rate_pct', 'scholarship_value', 'course_duration_years'}
            id_fields = {'student_id', 'agentsic_id'}
            for js_field, db_field in field_map.items():
                if js_field in request.data:
                    val = request.data[js_field]
                    if db_field in decimal_fields:
                        if val == '' or val is None:
                            val = None
                        else:
                            try:
                                val = Decimal(str(val))
                            except Exception:
                                val = None
                    if db_field in id_fields:
                        val = _clean_id(val)
                    setattr(s, db_field, val)
            if 'status' in request.data:
                record_status_change('commission_student', s.id, old_status, s.status, request.session.get('userId'), notes='Manual status update')
            s.updated_by_user_id = request.session.get('userId')
            s.save()

            recalculate_student(s, user_id=request.session.get('userId'))
            s.refresh_from_db()
            user_perms = request.session.get('userPermissions', [])
            return Response(filter_fields(student_to_dict(s), 'commission_student', user_perms))
        except Exception as e:
            return Response({'message': str(e)}, status=500)

    @require_permission("commission_tracker.student.delete")
    def delete(self, request, student_id):
        try:
            try:
                s = CommissionStudent.objects.get(id=student_id)
            except CommissionStudent.DoesNotExist:
                return Response({'message': 'Student not found'}, status=404)
            s.delete()
            return Response({'message': 'Deleted'})
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class StudentRecalculateView(APIView):
    @require_permission("commission_tracker.student.update")
    def post(self, request, student_id):
        try:
            try:
                s = CommissionStudent.objects.get(id=student_id)
            except CommissionStudent.DoesNotExist:
                return Response({'message': 'Student not found'}, status=404)
            recalculate_student(s, user_id=request.session.get('userId'))
            s.refresh_from_db()
            entries = CommissionEntry.objects.filter(commission_student_id=student_id)
            user_perms = request.session.get('userPermissions', [])
            return Response({
                'student': filter_fields(student_to_dict(s), 'commission_student', user_perms),
                'entries': filter_fields_list([entry_to_dict(e) for e in entries], 'commission_entry', user_perms),
            })
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class RecalculateAllView(APIView):
    @require_permission("commission_tracker.student.update")
    def post(self, request):
        try:
            term_name = request.data.get('termName')
            students = CommissionStudent.objects.all()
            recalculated = 0
            for student in students:
                entries = list(CommissionEntry.objects.filter(commission_student_id=student.id))
                if term_name:
                    entries = [e for e in entries if e.term_name == term_name]
                if not entries:
                    continue

                provider_ids = set(e.student_provider_id for e in entries if e.student_provider_id)
                providers_lookup = {sp.id: sp for sp in StudentProvider.objects.filter(id__in=provider_ids)} if provider_ids else {}

                for entry in entries:
                    prov_config = providers_lookup.get(entry.student_provider_id) if entry.student_provider_id else None
                    calc = calculate_entry(student, entry, prov_config)
                    entry.commission_rate_auto = calc['commissionRateAuto']
                    entry.commission_rate_used_pct = calc['commissionRateUsedPct']
                    entry.commission_amount = calc['commissionAmount']
                    entry.gst_amount = calc['gstAmount']
                    entry.total_amount = calc['totalAmount']
                    entry.rate_change_warning = calc['rateChangeWarning']
                    entry.scholarship_type_auto = calc['scholarshipTypeAuto']
                    entry.scholarship_value_auto = calc['scholarshipValueAuto']
                    entry.scholarship_type_used = calc['scholarshipTypeUsed']
                    entry.scholarship_value_used = calc['scholarshipValueUsed']
                    entry.scholarship_change_warning = calc['scholarshipChangeWarning']
                    entry.scholarship_amount = calc['scholarshipAmount']
                    entry.fee_after_scholarship = calc['feeAfterScholarship']
                    entry.save()
                    recalculated += 1

                all_entries = list(CommissionEntry.objects.filter(commission_student_id=student.id))
                term_order = get_term_order()
                master = compute_master_from_entries(all_entries, term_order)
                student.status = master['status']
                student.total_received = master['totalReceived']
                sync_per_provider_notes(student, all_entries, term_order)
                student.save(update_fields=['status', 'notes', 'total_received'])

            return Response({'message': f'Recalculated {recalculated} entries', 'recalculated': recalculated})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'message': str(e)}, status=500)


class StudentProvidersView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request, student_id):
        sps = StudentProvider.objects.filter(commission_student_id=student_id).order_by('id')
        return Response([provider_to_dict(sp) for sp in sps])

    @require_permission("commission_tracker.student.add")
    def post(self, request, student_id):
        try:
            d = request.data
            user_id = request.session.get('userId')
            sp = StudentProvider.objects.create(
                commission_student_id=student_id,
                provider=d.get('provider', ''),
                student_id=_clean_id(d.get('studentId')),
                country=d.get('country', 'Australia'),
                course_level=d.get('courseLevel'),
                course_name=d.get('courseName'),
                course_duration_years=d.get('courseDurationYears'),
                start_intake=d.get('startIntake'),
                commission_rate_pct=d.get('commissionRatePct'),
                gst_rate_pct=d.get('gstRatePct', 10),
                gst_applicable=d.get('gstApplicable', 'Yes'),
                scholarship_type=d.get('scholarshipType', 'None'),
                scholarship_value=d.get('scholarshipValue', 0),
                status=d.get('status', 'Under Enquiry'),
                notes=d.get('notes'),
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            return Response(provider_to_dict(sp))
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class StudentProviderDeleteView(APIView):
    @require_permission("commission_tracker.student.delete")
    def delete(self, request, student_id, provider_id):
        try:
            try:
                sp = StudentProvider.objects.get(id=provider_id, commission_student_id=student_id)
            except StudentProvider.DoesNotExist:
                return Response({'message': 'Provider not found'}, status=404)
            sp.delete()
            return Response({'message': 'Deleted'})
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class AllStudentProvidersView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request):
        sps = StudentProvider.objects.all().order_by('id')
        return Response([provider_to_dict(sp) for sp in sps])


class StudentProviderUpdateView(APIView):
    @require_permission("commission_tracker.student.update")
    def patch(self, request, provider_id):
        try:
            try:
                sp = StudentProvider.objects.get(id=provider_id)
            except StudentProvider.DoesNotExist:
                return Response({'message': 'Provider not found'}, status=404)
            field_map = {
                'provider': 'provider', 'studentId': 'student_id', 'country': 'country',
                'courseLevel': 'course_level', 'courseName': 'course_name',
                'courseDurationYears': 'course_duration_years', 'startIntake': 'start_intake',
                'commissionRatePct': 'commission_rate_pct', 'gstRatePct': 'gst_rate_pct',
                'gstApplicable': 'gst_applicable', 'scholarshipType': 'scholarship_type',
                'scholarshipValue': 'scholarship_value', 'status': 'status', 'notes': 'notes',
            }
            for js_field, db_field in field_map.items():
                if js_field in request.data:
                    val = request.data[js_field]
                    if db_field == 'student_id':
                        val = _clean_id(val)
                    setattr(sp, db_field, val)
            sp.updated_by_user_id = request.session.get('userId')
            sp.save()

            entries = CommissionEntry.objects.filter(student_provider_id=sp.id)
            parent_student = CommissionStudent.objects.filter(id=sp.commission_student_id).first()
            for entry in entries:
                calc = calculate_entry(
                    parent_student,
                    entry, sp
                )
                entry.commission_rate_auto = calc['commissionRateAuto']
                entry.commission_rate_used_pct = calc['commissionRateUsedPct']
                entry.commission_amount = calc['commissionAmount']
                entry.gst_amount = calc['gstAmount']
                entry.total_amount = calc['totalAmount']
                entry.rate_change_warning = calc['rateChangeWarning']
                entry.scholarship_type_auto = calc['scholarshipTypeAuto']
                entry.scholarship_value_auto = calc['scholarshipValueAuto']
                entry.scholarship_type_used = calc['scholarshipTypeUsed']
                entry.scholarship_value_used = calc['scholarshipValueUsed']
                entry.scholarship_change_warning = calc['scholarshipChangeWarning']
                entry.scholarship_amount = calc['scholarshipAmount']
                entry.fee_after_scholarship = calc['feeAfterScholarship']
                entry.save()

            return Response(provider_to_dict(sp))
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class StudentEntriesView(APIView):
    @require_permission("commission_tracker.entry.read")
    def get(self, request, student_id):
        entries = CommissionEntry.objects.filter(commission_student_id=student_id).order_by('term_name')
        user_perms = request.session.get('userPermissions', [])
        return Response(filter_fields_list([entry_to_dict(e) for e in entries], 'commission_entry', user_perms))

    @require_permission("commission_tracker.entry.add")
    def post(self, request, student_id):
        try:
            d = request.data
            try:
                student = CommissionStudent.objects.get(id=student_id)
            except CommissionStudent.DoesNotExist:
                return Response({'message': 'Student not found'}, status=404)

            provider_config = None
            sp_id = d.get('studentProviderId')
            if sp_id:
                try:
                    provider_config = StudentProvider.objects.get(id=sp_id)
                except StudentProvider.DoesNotExist:
                    pass

            calc = calculate_entry(student, d, provider_config)

            user_id = request.session.get('userId')
            entry = CommissionEntry.objects.create(
                commission_student_id=student_id,
                student_provider_id=sp_id,
                term_name=d.get('termName', ''),
                academic_year=d.get('academicYear'),
                fee_gross=d.get('feeGross', 0),
                commission_rate_auto=calc['commissionRateAuto'],
                commission_rate_override_pct=d.get('commissionRateOverridePct'),
                commission_rate_used_pct=calc['commissionRateUsedPct'],
                commission_amount=calc['commissionAmount'],
                bonus=d.get('bonus', 0),
                gst_amount=calc['gstAmount'],
                total_amount=calc['totalAmount'],
                payment_status=d.get('paymentStatus', 'Pending'),
                paid_date=d.get('paidDate') or None,
                invoice_no=d.get('invoiceNo'),
                payment_ref=d.get('paymentRef'),
                notes=d.get('notes'),
                student_status=d.get('studentStatus', 'Under Enquiry'),
                rate_change_warning=calc['rateChangeWarning'],
                scholarship_type_auto=calc['scholarshipTypeAuto'],
                scholarship_value_auto=calc['scholarshipValueAuto'],
                scholarship_type_override=d.get('scholarshipTypeOverride'),
                scholarship_value_override=d.get('scholarshipValueOverride'),
                scholarship_type_used=calc['scholarshipTypeUsed'],
                scholarship_value_used=calc['scholarshipValueUsed'],
                scholarship_change_warning=calc['scholarshipChangeWarning'],
                scholarship_amount=calc['scholarshipAmount'],
                fee_after_scholarship=calc['feeAfterScholarship'],
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )

            old_student_status = student.status
            entries = list(CommissionEntry.objects.filter(commission_student_id=student_id))
            term_order = get_term_order()
            master = compute_master_from_entries(entries, term_order)
            student.status = master['status']
            student.total_received = master['totalReceived']
            sync_per_provider_notes(student, entries, term_order)
            student.save(update_fields=['status', 'notes', 'total_received'])
            record_status_change('commission_student', student.id, old_student_status, student.status, user_id, notes='Entry added')

            user_perms = request.session.get('userPermissions', [])
            return Response(filter_fields(entry_to_dict(entry), 'commission_entry', user_perms))
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class EntryDetailView(APIView):
    @require_permission("commission_tracker.entry.update")
    def patch(self, request, entry_id):
        try:
            try:
                entry = CommissionEntry.objects.get(id=entry_id)
            except CommissionEntry.DoesNotExist:
                return Response({'message': 'Entry not found'}, status=404)

            d = request.data
            field_map = {
                'termName': 'term_name', 'academicYear': 'academic_year', 'feeGross': 'fee_gross',
                'commissionRateOverridePct': 'commission_rate_override_pct', 'bonus': 'bonus',
                'paymentStatus': 'payment_status', 'paidDate': 'paid_date', 'invoiceNo': 'invoice_no',
                'paymentRef': 'payment_ref', 'notes': 'notes', 'studentStatus': 'student_status',
                'scholarshipTypeOverride': 'scholarship_type_override',
                'scholarshipValueOverride': 'scholarship_value_override',
                'studentProviderId': 'student_provider_id',
            }
            decimal_fields = {
                'fee_gross', 'commission_rate_override_pct', 'bonus',
                'scholarship_value_override',
            }
            nullable_fields = {
                'paid_date', 'invoice_no', 'payment_ref', 'notes',
                'commission_rate_override_pct', 'scholarship_type_override',
                'scholarship_value_override',
            }
            for js_field, db_field in field_map.items():
                if js_field in d:
                    val = d[js_field]
                    if db_field in nullable_fields and (val == '' or val is None):
                        val = None
                    elif db_field in decimal_fields and val is not None:
                        try:
                            val = float(val) if val != '' else 0
                        except (ValueError, TypeError):
                            val = 0
                    setattr(entry, db_field, val)

            entry.updated_by_user_id = request.session.get('userId')
            student = CommissionStudent.objects.get(id=entry.commission_student_id)
            provider_config = None
            if entry.student_provider_id:
                try:
                    provider_config = StudentProvider.objects.get(id=entry.student_provider_id)
                except StudentProvider.DoesNotExist:
                    pass

            calc = calculate_entry(student, entry, provider_config)
            entry.commission_rate_auto = calc['commissionRateAuto']
            entry.commission_rate_used_pct = calc['commissionRateUsedPct']
            entry.commission_amount = calc['commissionAmount']
            entry.gst_amount = calc['gstAmount']
            entry.total_amount = calc['totalAmount']
            entry.rate_change_warning = calc['rateChangeWarning']
            entry.scholarship_type_auto = calc['scholarshipTypeAuto']
            entry.scholarship_value_auto = calc['scholarshipValueAuto']
            entry.scholarship_type_used = calc['scholarshipTypeUsed']
            entry.scholarship_value_used = calc['scholarshipValueUsed']
            entry.scholarship_change_warning = calc['scholarshipChangeWarning']
            entry.scholarship_amount = calc['scholarshipAmount']
            entry.fee_after_scholarship = calc['feeAfterScholarship']
            entry.save()

            old_student_status = student.status
            entries = list(CommissionEntry.objects.filter(commission_student_id=entry.commission_student_id))
            term_order = get_term_order()
            master = compute_master_from_entries(entries, term_order)
            student.status = master['status']
            student.total_received = master['totalReceived']
            sync_per_provider_notes(student, entries, term_order)
            student.save(update_fields=['status', 'notes', 'total_received'])
            record_status_change('commission_student', student.id, old_student_status, student.status, request.session.get('userId'), notes='Entry updated')

            user_perms = request.session.get('userPermissions', [])
            return Response(filter_fields(entry_to_dict(entry), 'commission_entry', user_perms))
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'message': str(e)}, status=500)

    @require_permission("commission_tracker.entry.delete")
    def delete(self, request, entry_id):
        try:
            try:
                entry = CommissionEntry.objects.get(id=entry_id)
            except CommissionEntry.DoesNotExist:
                return Response({'message': 'Entry not found'}, status=404)

            student_id = entry.commission_student_id
            entry.delete()

            try:
                student = CommissionStudent.objects.get(id=student_id)
                old_student_status = student.status
                entries = list(CommissionEntry.objects.filter(commission_student_id=student_id))
                term_order = get_term_order()
                master = compute_master_from_entries(entries, term_order)
                student.status = master['status']
                student.total_received = master['totalReceived']
                sync_per_provider_notes(student, entries, term_order)
                student.save(update_fields=['status', 'notes', 'total_received'])
                record_status_change('commission_student', student.id, old_student_status, student.status, request.session.get('userId'), notes='Entry deleted')
            except CommissionStudent.DoesNotExist:
                pass

            return Response({'message': 'Deleted'})
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class AllEntriesView(APIView):
    @require_permission("commission_tracker.entry.read")
    def get(self, request):
        year_param = request.query_params.get('year')
        entries = CommissionEntry.objects.all().order_by('commission_student_id', 'term_name')

        if year_param:
            target_year = int(year_param)
            excluded_ids = get_excluded_student_ids_for_year(target_year)
            if excluded_ids:
                entries = entries.exclude(commission_student_id__in=excluded_ids)

        user_perms = request.session.get('userPermissions', [])
        grouped = {}
        for e in entries:
            sid = e.commission_student_id
            if sid not in grouped:
                grouped[sid] = []
            grouped[sid].append(entry_to_dict(e))
        for sid in grouped:
            grouped[sid] = filter_fields_list(grouped[sid], 'commission_entry', user_perms)
        return Response(grouped)


class TermsView(APIView):
    @require_auth
    def get(self, request):
        terms = CommissionTerm.objects.all().order_by('sort_order')
        return Response([{
            'id': t.id, 'termName': t.term_name, 'termLabel': t.term_label,
            'year': t.year, 'termNumber': t.term_number, 'sortOrder': t.sort_order,
            'isActive': t.is_active,
            'createdAt': t.created_at.isoformat() if t.created_at else None,
        } for t in terms])

    @require_permission("commission_tracker.student.update")
    def post(self, request):
        try:
            d = request.data
            t = CommissionTerm.objects.create(
                term_name=d.get('termName', ''),
                term_label=d.get('termLabel', ''),
                year=d.get('year', 0),
                term_number=d.get('termNumber', 0),
                sort_order=d.get('sortOrder', 0),
                is_active=d.get('isActive', True),
            )
            return Response({
                'id': t.id, 'termName': t.term_name, 'termLabel': t.term_label,
                'year': t.year, 'termNumber': t.term_number, 'sortOrder': t.sort_order,
                'isActive': t.is_active,
                'createdAt': t.created_at.isoformat() if t.created_at else None,
            })
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class TermDeleteView(APIView):
    @require_permission("commission_tracker.student.update")
    def delete(self, request, term_id):
        try:
            CommissionTerm.objects.filter(id=term_id).delete()
            return Response({'message': 'Deleted'})
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class FiltersView(APIView):
    @require_auth
    def get(self, request):
        agents = list(CommissionStudent.objects.values_list('agent_name', flat=True).distinct())
        providers = list(CommissionStudent.objects.values_list('provider', flat=True).distinct())
        countries = list(CommissionStudent.objects.values_list('country', flat=True).distinct())
        statuses = ["Under Enquiry", "Claim Next Semester", "On Break", "Withdrawn", "Complete", "Active", "Other"]
        return Response({
            'agents': sorted(set(a for a in agents if a)),
            'providers': sorted(set(p for p in providers if p)),
            'countries': sorted(set(c for c in countries if c)),
            'statuses': statuses,
        })


class YearsView(APIView):
    @require_auth
    def get(self, request):
        from datetime import datetime
        years = set(CommissionTerm.objects.values_list('year', flat=True).distinct())
        years.add(datetime.now().year)
        return Response(sorted(years, reverse=True))


class CommissionTrackerExportView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request):
        try:
            qs = CommissionStudent.objects.all()
            agent = request.query_params.get('agent')
            prov = request.query_params.get('provider')
            country = request.query_params.get('country')
            status = request.query_params.get('status')
            year = request.query_params.get('year')

            if agent:
                qs = qs.filter(agent_name__icontains=agent)
            if prov:
                qs = qs.filter(provider__icontains=prov)
            if country:
                qs = qs.filter(country__iexact=country)
            if status:
                qs = qs.filter(status=status)
            if year:
                qs = qs.filter(start_intake__icontains=year)

            qs = qs.order_by('-id')
            headers = [
                'ID', 'Agent Name', 'Student ID', 'AgentSIC ID', 'Student Name',
                'Provider', 'Country', 'Start Intake', 'Course Level', 'Course Name',
                'Course Duration (Years)', 'Commission Rate %', 'GST Rate %',
                'GST Applicable', 'Scholarship Type', 'Scholarship Value',
                'Status', 'Total Received', 'Created At',
            ]
            rows = []
            for s in qs:
                rows.append([
                    s.id, s.agent_name, _clean_id(s.student_id), _clean_id(s.agentsic_id), s.student_name,
                    s.provider, s.country, s.start_intake, s.course_level, s.course_name,
                    str(s.course_duration_years) if s.course_duration_years is not None else '',
                    str(s.commission_rate_pct) if s.commission_rate_pct is not None else '',
                    str(s.gst_rate_pct) if s.gst_rate_pct is not None else '',
                    s.gst_applicable, s.scholarship_type,
                    str(s.scholarship_value) if s.scholarship_value is not None else '0',
                    s.status, str(s.total_received) if s.total_received is not None else '0',
                    s.created_at.isoformat() if s.created_at else '',
                ])
            from core.exports import export_data
            return export_data(request, 'commission_tracker_export', headers, rows, 'Commission Tracker')
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class DashboardView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request):
        try:
            students = CommissionStudent.objects.all()
            total_students = students.count()
            total_received = float(students.aggregate(total=Sum('total_received'))['total'] or 0)

            status_counts = {}
            for s in students:
                status_counts[s.status] = status_counts.get(s.status, 0) + 1

            entries = CommissionEntry.objects.all()
            total_entries = entries.count()
            total_commission = float(entries.aggregate(total=Sum('commission_amount'))['total'] or 0)
            total_gst = float(entries.aggregate(total=Sum('gst_amount'))['total'] or 0)

            user_perms = request.session.get('userPermissions', [])
            result = {
                'totalStudents': total_students,
                'statusCounts': status_counts,
                'totalEntries': total_entries,
            }
            if 'commission_tracker.field.financials' in user_perms:
                result['totalReceived'] = round2(total_received)
                result['totalCommission'] = round2(total_commission)
                result['totalGst'] = round2(total_gst)
            return Response(result)
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class DashboardYearView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request, year):
        try:
            user_perms = request.session.get('userPermissions', [])
            has_financials = 'commission_tracker.field.financials' in user_perms
            target_year = int(year)
            terms = CommissionTerm.objects.filter(year=target_year).order_by('sort_order')
            term_names = [t.term_name for t in terms]

            intake_filter = request.query_params.get('intake', '')

            excluded_ids = get_excluded_student_ids_for_year(target_year)

            entries = CommissionEntry.objects.filter(term_name__in=term_names)
            if excluded_ids:
                entries = entries.exclude(commission_student_id__in=excluded_ids)

            entry_student_ids = set(entries.values_list('commission_student_id', flat=True).distinct())

            all_students_qs = CommissionStudent.objects.all()
            if excluded_ids:
                all_students_qs = all_students_qs.exclude(id__in=excluded_ids)

            if intake_filter and intake_filter.upper() != 'ALL':
                all_students_qs = all_students_qs.filter(start_intake__icontains=intake_filter)
                filtered_student_ids = set(all_students_qs.values_list('id', flat=True))
                entries = entries.filter(commission_student_id__in=filtered_student_ids)
                entry_student_ids = entry_student_ids & filtered_student_ids

            students = all_students_qs

            total_students = students.count()
            providers_set = set()
            agent_counts = {}
            provider_counts = {}
            status_counts = {}
            for s in students:
                providers_set.add(s.provider)
                agent_counts[s.agent_name] = agent_counts.get(s.agent_name, 0) + 1
                provider_counts[s.provider] = provider_counts.get(s.provider, 0) + 1
                status_counts[s.status] = status_counts.get(s.status, 0) + 1

            by_agent = sorted([{'agent': k, 'count': v} for k, v in agent_counts.items()], key=lambda x: -x['count'])
            by_provider = [{'provider': k, 'count': v} for k, v in provider_counts.items()]
            term_stats = []
            for t in terms:
                term_entries = [e for e in entries if e.term_name == t.term_name]
                total = sum(float(e.total_amount or 0) for e in term_entries)
                comm = sum(float(e.commission_amount or 0) for e in term_entries)
                gst = sum(float(e.gst_amount or 0) for e in term_entries)
                bonus = sum(float(e.bonus or 0) for e in term_entries)
                term_stat = {
                    'termName': t.term_name, 'termLabel': t.term_label,
                    'entryCount': len(term_entries),
                }
                if has_financials:
                    term_stat['totalAmount'] = round2(total)
                    term_stat['commissionAmount'] = round2(comm)
                    term_stat['gstAmount'] = round2(gst)
                    term_stat['bonusAmount'] = round2(bonus)
                term_stats.append(term_stat)

            if has_financials:
                provider_financials = {}
                student_provider_map = {s.id: s.provider for s in students}
                for e in entries:
                    prov = student_provider_map.get(e.commission_student_id, 'Unknown')
                    if prov not in provider_financials:
                        provider_financials[prov] = {'totalCommission': 0, 'totalBonus': 0, 'totalReceived': 0, 'pending': 0}
                    provider_financials[prov]['totalCommission'] += float(e.commission_amount or 0)
                    provider_financials[prov]['totalBonus'] += float(e.bonus or 0)
                    if (e.payment_status or '').lower() == 'paid':
                        provider_financials[prov]['totalReceived'] += float(e.total_amount or 0)

                for p in by_provider:
                    fin = provider_financials.get(p['provider'], {})
                    p['totalCommission'] = round2(fin.get('totalCommission', 0))
                    p['totalBonus'] = round2(fin.get('totalBonus', 0))
                    p['totalReceived'] = round2(fin.get('totalReceived', 0))
                    p['pending'] = round2(p['totalCommission'] - p['totalReceived'])

                by_provider.sort(key=lambda x: -x.get('totalCommission', 0))

            paid_entries = [e for e in entries if (e.payment_status or '').lower() == 'paid']
            total_received = sum(float(e.total_amount or 0) for e in paid_entries)

            result = {
                'year': int(year),
                'totalStudents': total_students,
                'totalProviders': len(providers_set),
                'terms': term_stats,
                'totalEntries': sum(ts['entryCount'] for ts in term_stats),
                'byStatus': status_counts,
                'byAgent': by_agent,
                'byProvider': by_provider,
            }
            if has_financials:
                total_comm = sum(ts.get('commissionAmount', 0) for ts in term_stats)
                total_bonus = sum(ts.get('bonusAmount', 0) for ts in term_stats)
                result['totalAmount'] = round2(sum(ts.get('totalAmount', 0) for ts in term_stats))
                result['totalCommission'] = round2(total_comm)
                result['totalBonus'] = round2(total_bonus)
                result['totalGst'] = round2(sum(ts.get('gstAmount', 0) for ts in term_stats))
                result['totalReceived'] = round2(total_received)
                result['totalPending'] = round2(total_comm - total_received)
            return Response(result)
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class CommissionInsightsView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request, year):
        try:
            from sub_agent.models import SubAgentEntry, SubAgentTermEntry
            user_perms = request.session.get('userPermissions', [])
            has_financials = 'commission_tracker.field.financials' in user_perms
            if not has_financials:
                return Response({'insights': None})

            target_year = int(year)
            terms = CommissionTerm.objects.filter(year=target_year).order_by('sort_order')
            term_names = [t.term_name for t in terms]
            excluded_ids = get_excluded_student_ids_for_year(target_year)

            entries = list(CommissionEntry.objects.filter(term_name__in=term_names))
            if excluded_ids:
                entries = [e for e in entries if e.commission_student_id not in excluded_ids]

            all_students_qs = CommissionStudent.objects.all()
            if excluded_ids:
                all_students_qs = all_students_qs.exclude(id__in=excluded_ids)
            students = {s.id: s for s in all_students_qs}

            sub_masters = {m.commission_student_id: m for m in SubAgentEntry.objects.all()}
            sub_term_entries = list(SubAgentTermEntry.objects.filter(term_name__in=term_names))
            if excluded_ids:
                sub_term_entries = [e for e in sub_term_entries if e.commission_student_id not in excluded_ids]

            sub_paid_by_student = {}
            for se in sub_term_entries:
                sub_paid_by_student[se.commission_student_id] = sub_paid_by_student.get(se.commission_student_id, 0) + float(se.total_paid or 0)

            provider_data = {}
            agent_data = {}
            agent_provider_data = {}
            student_term_data = {}
            for e in entries:
                s = students.get(e.commission_student_id)
                if not s:
                    continue
                prov = (s.provider or 'Unknown').strip()
                agent = (s.agent_name or 'Unknown').strip()
                comm = float(e.commission_amount or 0) + float(e.bonus or 0)
                total_with_gst = float(e.total_amount or 0)

                if prov not in provider_data:
                    provider_data[prov] = {'commission': 0, 'totalWithGst': 0, 'subPaid': 0, 'students': set(), 'hasSubAgent': set()}
                provider_data[prov]['commission'] += comm
                provider_data[prov]['totalWithGst'] += total_with_gst
                provider_data[prov]['students'].add(e.commission_student_id)

                if agent not in agent_data:
                    agent_data[agent] = {'commission': 0, 'subPaid': 0, 'students': set(), 'hasSubAgent': set()}
                agent_data[agent]['commission'] += comm
                agent_data[agent]['students'].add(e.commission_student_id)

                ap_key = (agent, prov)
                if ap_key not in agent_provider_data:
                    agent_provider_data[ap_key] = {'commission': 0, 'subPaid': 0, 'students': set()}
                agent_provider_data[ap_key]['commission'] += comm
                agent_provider_data[ap_key]['students'].add(e.commission_student_id)

                tn = e.term_name
                if tn not in student_term_data:
                    student_term_data[tn] = {'commission': 0, 'subPaid': 0, 'students': set()}
                student_term_data[tn]['commission'] += comm
                student_term_data[tn]['students'].add(e.commission_student_id)

            for sid, sub_paid in sub_paid_by_student.items():
                s = students.get(sid)
                if not s:
                    continue
                prov = (s.provider or 'Unknown').strip()
                agent = (s.agent_name or 'Unknown').strip()
                if prov in provider_data:
                    provider_data[prov]['subPaid'] += sub_paid
                    provider_data[prov]['hasSubAgent'].add(sid)
                if agent in agent_data:
                    agent_data[agent]['subPaid'] += sub_paid
                    agent_data[agent]['hasSubAgent'].add(sid)
                ap_key = (agent, prov)
                if ap_key in agent_provider_data:
                    agent_provider_data[ap_key]['subPaid'] += sub_paid

            for se in sub_term_entries:
                tn = se.term_name
                if tn in student_term_data:
                    student_term_data[tn]['subPaid'] += float(se.total_paid or 0)

            comm_by_student = {}
            for e in entries:
                sid = e.commission_student_id
                comm_by_student[sid] = comm_by_student.get(sid, 0) + float(e.commission_amount or 0) + float(e.bonus or 0)

            overpaid_students = []
            for sid, sub_paid in sub_paid_by_student.items():
                s = students.get(sid)
                if not s:
                    continue
                student_comm = comm_by_student.get(sid, 0)
                if sub_paid > student_comm and student_comm > 0:
                    overpaid_students.append({
                        'studentName': s.student_name,
                        'agentName': s.agent_name or 'Unknown',
                        'provider': (s.provider or 'Unknown').strip(),
                        'commission': round2(student_comm),
                        'subAgentPaid': round2(sub_paid),
                        'loss': round2(sub_paid - student_comm),
                    })
            overpaid_students.sort(key=lambda x: x['loss'], reverse=True)

            total_commission = sum(pd['commission'] for pd in provider_data.values())
            total_sub_paid = sum(pd['subPaid'] for pd in provider_data.values())
            total_margin = total_commission - total_sub_paid
            margin_pct = (total_margin / total_commission * 100) if total_commission > 0 else 0
            total_students_count = sum(len(pd['students']) for pd in provider_data.values())
            avg_comm_all = (total_commission / total_students_count) if total_students_count > 0 else 0
            avg_margin_all = (total_margin / total_students_count) if total_students_count > 0 else 0

            prev_year_terms = CommissionTerm.objects.filter(year=target_year - 1).order_by('sort_order')
            prev_total = 0
            prev_sub_paid = 0
            prev_students_count = 0
            prev_provider_data = {}
            prev_agent_data = {}
            if prev_year_terms.exists():
                prev_term_names = [t.term_name for t in prev_year_terms]
                prev_entries_qs = CommissionEntry.objects.filter(term_name__in=prev_term_names)
                prev_excluded = get_excluded_student_ids_for_year(target_year - 1)
                if prev_excluded:
                    prev_entries_qs = prev_entries_qs.exclude(commission_student_id__in=prev_excluded)
                prev_entries_list = list(prev_entries_qs)
                prev_total = sum(float(e.commission_amount or 0) + float(e.bonus or 0) for e in prev_entries_list)
                prev_student_ids = set(e.commission_student_id for e in prev_entries_list)
                prev_students_count = len(prev_student_ids)
                for e in prev_entries_list:
                    s = students.get(e.commission_student_id)
                    if not s:
                        continue
                    prov = (s.provider or 'Unknown').strip()
                    agent = (s.agent_name or 'Unknown').strip()
                    c = float(e.commission_amount or 0) + float(e.bonus or 0)
                    if prov not in prev_provider_data:
                        prev_provider_data[prov] = {'commission': 0, 'students': set()}
                    prev_provider_data[prov]['commission'] += c
                    prev_provider_data[prov]['students'].add(e.commission_student_id)
                    if agent not in prev_agent_data:
                        prev_agent_data[agent] = {'commission': 0, 'students': set()}
                    prev_agent_data[agent]['commission'] += c
                    prev_agent_data[agent]['students'].add(e.commission_student_id)
                prev_sub_entries = list(SubAgentTermEntry.objects.filter(term_name__in=prev_term_names))
                if prev_excluded:
                    prev_sub_entries = [e for e in prev_sub_entries if e.commission_student_id not in prev_excluded]
                prev_sub_paid = sum(float(e.total_paid or 0) for e in prev_sub_entries)

            avg_margin_per_student = (total_margin / total_students_count) if total_students_count > 0 else 0

            def calc_opportunity_score(margin, margin_pct_val, avg_per_student, student_count, growth_pct=None, sub_rate_pct=0):
                margin_contrib = min(margin / max(total_margin, 1) * 100, 100) if total_margin > 0 else 0
                pct_score = min(max(margin_pct_val, 0), 100)
                margin_per_student = (margin / student_count) if student_count > 0 else 0
                per_student_score = min(margin_per_student / max(avg_margin_per_student, 1) * 50, 100) if avg_margin_per_student > 0 else 0
                vol_score = min(student_count / max(total_students_count, 1) * 200, 100)
                growth_score = min(max(growth_pct or 0, 0), 100)
                payout_efficiency = max(0, 100 - sub_rate_pct) if sub_rate_pct > 0 else 100
                score = (margin_contrib * 0.30 + pct_score * 0.20 + per_student_score * 0.15 + vol_score * 0.10 + growth_score * 0.10 + payout_efficiency * 0.15)
                return round(min(score, 100))

            def get_ai_action(margin_pct_val, avg_per_student, student_count, growth_pct=None, sub_rate_pct=0):
                if margin_pct_val < 0:
                    return 'High Risk'
                if sub_rate_pct > 80 and student_count >= 2:
                    return 'Review Payout'
                if avg_per_student < avg_comm_all * 0.5 and student_count >= 3:
                    return 'Negotiate'
                if margin_pct_val > 60 and student_count <= 3:
                    return 'Scale'
                if margin_pct_val > 50 and student_count > 5:
                    return 'Scale'
                if margin_pct_val < 30 and student_count > 3:
                    return 'Negotiate'
                if growth_pct is not None and growth_pct < -20:
                    return 'Re-engage'
                if growth_pct is not None and growth_pct > 30:
                    return 'Scale'
                return 'Monitor'

            provider_insights = []
            for prov, pd in provider_data.items():
                margin = pd['commission'] - pd['subPaid']
                pct = (margin / pd['commission'] * 100) if pd['commission'] > 0 else 0
                avg_comm = pd['commission'] / len(pd['students']) if pd['students'] else 0
                sub_rate_pct = (pd['subPaid'] / pd['commission'] * 100) if pd['commission'] > 0 else 0
                prev_comm = prev_provider_data.get(prov, {}).get('commission', 0)
                growth = ((pd['commission'] - prev_comm) / prev_comm * 100) if prev_comm > 0 else None
                prev_sc = len(prev_provider_data.get(prov, {}).get('students', set()))
                sc = len(pd['students'])
                opp_score = calc_opportunity_score(margin, pct, avg_comm, sc, growth, sub_rate_pct)
                action = get_ai_action(pct, avg_comm, sc, growth, sub_rate_pct)
                pi = {
                    'provider': prov,
                    'commission': round2(pd['commission']),
                    'subAgentPaid': round2(pd['subPaid']),
                    'margin': round2(margin),
                    'marginPct': round(pct, 1),
                    'studentCount': sc,
                    'subAgentStudents': len(pd['hasSubAgent']),
                    'avgCommPerStudent': round2(avg_comm),
                    'subAgentRatePct': round(sub_rate_pct, 1),
                    'opportunityScore': opp_score,
                    'aiAction': action,
                    'trend': round(growth, 1) if growth is not None else None,
                    'prevCommission': round2(prev_comm) if prev_comm > 0 else None,
                    'prevStudents': prev_sc if prev_sc > 0 else None,
                }
                provider_insights.append(pi)
            provider_insights.sort(key=lambda x: x['margin'], reverse=True)

            agent_insights = []
            for agent, ad in agent_data.items():
                margin = ad['commission'] - ad['subPaid']
                pct = (margin / ad['commission'] * 100) if ad['commission'] > 0 else 0
                sub_rate = (ad['subPaid'] / ad['commission'] * 100) if ad['commission'] > 0 else 0
                avg_comm = ad['commission'] / len(ad['students']) if ad['students'] else 0
                prev_comm = prev_agent_data.get(agent, {}).get('commission', 0)
                growth = ((ad['commission'] - prev_comm) / prev_comm * 100) if prev_comm > 0 else None
                sc = len(ad['students'])
                opp_score = calc_opportunity_score(margin, pct, avg_comm, sc, growth, sub_rate)
                action = get_ai_action(pct, avg_comm, sc, growth, sub_rate)
                agent_insights.append({
                    'agent': agent,
                    'commission': round2(ad['commission']),
                    'subAgentPaid': round2(ad['subPaid']),
                    'margin': round2(margin),
                    'marginPct': round(pct, 1),
                    'studentCount': sc,
                    'subAgentStudents': len(ad['hasSubAgent']),
                    'subAgentRatePct': round(sub_rate, 1),
                    'avgCommPerStudent': round2(avg_comm),
                    'opportunityScore': opp_score,
                    'aiAction': action,
                    'trend': round(growth, 1) if growth is not None else None,
                    'prevCommission': round2(prev_comm) if prev_comm > 0 else None,
                })
            agent_insights.sort(key=lambda x: x['margin'], reverse=True)

            agent_provider_pairs = []
            for (agent, prov), apd in agent_provider_data.items():
                margin = apd['commission'] - apd['subPaid']
                pct = (margin / apd['commission'] * 100) if apd['commission'] > 0 else 0
                avg_comm = apd['commission'] / len(apd['students']) if apd['students'] else 0
                sub_rate_ap = (apd['subPaid'] / apd['commission'] * 100) if apd['commission'] > 0 else 0
                agent_provider_pairs.append({
                    'agent': agent, 'provider': prov,
                    'commission': round2(apd['commission']),
                    'subAgentPaid': round2(apd['subPaid']),
                    'margin': round2(margin),
                    'marginPct': round(pct, 1),
                    'studentCount': len(apd['students']),
                    'avgCommPerStudent': round2(avg_comm),
                    'subAgentRatePct': round(sub_rate_ap, 1),
                })
            agent_provider_pairs.sort(key=lambda x: x['margin'], reverse=True)

            def calc_recommended_payout(current_pct, margin_pct_val, student_count, is_high_margin_provider=False):
                if current_pct <= 0:
                    return current_pct, 'Optimized'
                target = current_pct
                if margin_pct_val < 0:
                    target = max(current_pct - 10, 30)
                elif margin_pct_val < 30 and student_count >= 3:
                    target = max(current_pct - 10, 40)
                elif margin_pct_val < 30:
                    target = max(current_pct - 5, 40)
                elif is_high_margin_provider and current_pct > 60:
                    target = max(current_pct - 5, 50)
                elif student_count <= 2 and current_pct < 70:
                    target = min(current_pct + 2, 70)
                reduction = current_pct - target
                if reduction > 10:
                    target = current_pct - 10
                    reduction = 10
                if current_pct < 30 and student_count >= 3:
                    return round(min(current_pct + 5, 50), 1), 'Underpaid'
                if reduction < 0:
                    reduction = 0
                    target = current_pct
                if reduction >= 5:
                    badge = 'High Payout Risk'
                elif reduction > 0:
                    badge = 'Optimize'
                else:
                    badge = 'Optimized'
                return round(target, 1), badge

            payout_optimizations = []
            for a in agent_insights:
                if a['subAgentPaid'] <= 0:
                    continue
                current_pct = a['subAgentRatePct']
                is_high_margin_prov = False
                rec_pct, badge = calc_recommended_payout(current_pct, a['marginPct'], a['studentCount'], is_high_margin_prov)
                reduction = current_pct - rec_pct
                expected_gain = a['commission'] * (reduction / 100) if reduction > 0 else 0
                avg_margin_per_app = (a['margin'] / a['studentCount']) if a['studentCount'] > 0 else 0
                payout_optimizations.append({
                    'agent': a['agent'],
                    'currentPayoutPct': round(current_pct, 1),
                    'recommendedPayoutPct': round(rec_pct, 1),
                    'reductionPct': round(reduction, 1),
                    'expectedGain': round2(expected_gain),
                    'commission': a['commission'],
                    'subAgentPaid': a['subAgentPaid'],
                    'marginPct': a['marginPct'],
                    'studentCount': a['studentCount'],
                    'avgMarginPerApp': round2(avg_margin_per_app),
                    'badge': badge,
                })
                a['payoutBadge'] = badge
                a['recommendedPayoutPct'] = round(rec_pct, 1)
            payout_optimizations.sort(key=lambda x: x['expectedGain'], reverse=True)
            total_payout_recovery = sum(p['expectedGain'] for p in payout_optimizations if p['expectedGain'] > 0)

            ap_payout_optimizations = []
            for ap in agent_provider_pairs:
                if ap['subAgentPaid'] <= 0:
                    continue
                current_pct = ap['subAgentRatePct']
                prov_margin_high = any(p['marginPct'] > 50 for p in provider_insights if p['provider'] == ap['provider'])
                rec_pct, badge = calc_recommended_payout(current_pct, ap['marginPct'], ap['studentCount'], prov_margin_high)
                reduction = current_pct - rec_pct
                expected_gain = ap['commission'] * (reduction / 100) if reduction > 0 else 0
                if reduction > 0:
                    ap_payout_optimizations.append({
                        'agent': ap['agent'], 'provider': ap['provider'],
                        'currentPayoutPct': round(current_pct, 1),
                        'recommendedPayoutPct': round(rec_pct, 1),
                        'reductionPct': round(reduction, 1),
                        'expectedGain': round2(expected_gain),
                        'marginPct': ap['marginPct'],
                        'studentCount': ap['studentCount'],
                        'badge': badge,
                    })
            ap_payout_optimizations.sort(key=lambda x: x['expectedGain'], reverse=True)

            leakage_alerts = []
            for a in agent_insights:
                if a['studentCount'] >= 2 and a['marginPct'] < 30 and a['subAgentPaid'] > 0:
                    missed = avg_margin_all * a['studentCount'] - a['margin'] if a['margin'] < avg_margin_all * a['studentCount'] else 0
                    if missed > 0:
                        root_cause_parts = []
                        if a['subAgentRatePct'] > 60:
                            root_cause_parts.append(f'High payout ({a["subAgentRatePct"]}%)')
                        if a['avgCommPerStudent'] < avg_comm_all * 0.7:
                            root_cause_parts.append('Low provider commission rates')
                        root_cause = ' + '.join(root_cause_parts) if root_cause_parts else 'Low overall margin'
                        action_parts = []
                        if a['subAgentRatePct'] > 60:
                            rec = a.get('recommendedPayoutPct', a['subAgentRatePct'] - 5)
                            action_parts.append(f'Reduce payout to {rec}%')
                        if a['avgCommPerStudent'] < avg_comm_all * 0.7:
                            action_parts.append('Renegotiate provider rates')
                        action = ' OR '.join(action_parts) if action_parts else 'Review margin structure'
                        leakage_alerts.append({
                            'entity': a['agent'], 'entityType': 'agent',
                            'issue': f'High application volume ({a["studentCount"]} applications) but low margin ({a["marginPct"]}%)',
                            'rootCause': root_cause,
                            'estimatedLoss': round2(missed),
                            'action': action,
                        })
            for p in provider_insights:
                if p['studentCount'] >= 2 and p['avgCommPerStudent'] < avg_comm_all * 0.5:
                    missed = (avg_comm_all - p['avgCommPerStudent']) * p['studentCount']
                    root_cause = f'Commission rate ${p["avgCommPerStudent"]:,.2f}/app is {round((1 - p["avgCommPerStudent"] / avg_comm_all) * 100)}% below portfolio avg'
                    action = f'Negotiate rate increase — closing the gap to ${avg_comm_all:,.2f}/app across {p["studentCount"]} applications'
                    leakage_alerts.append({
                        'entity': p['provider'], 'entityType': 'provider',
                        'issue': f'Below-average commission (${p["avgCommPerStudent"]:,.2f}/application vs ${avg_comm_all:,.2f} avg)',
                        'rootCause': root_cause,
                        'estimatedLoss': round2(missed),
                        'action': action,
                    })
            leakage_alerts.sort(key=lambda x: x['estimatedLoss'], reverse=True)

            negotiation_opps = []
            benchmark_avg = avg_comm_all
            for p in provider_insights:
                if p['studentCount'] >= 3 and p['avgCommPerStudent'] < benchmark_avg * 0.7 and p['commission'] > 0:
                    uplift_5pct = p['commission'] * 0.05
                    uplift_10pct = p['commission'] * 0.10
                    gap_pct = ((benchmark_avg - p['avgCommPerStudent']) / benchmark_avg * 100) if benchmark_avg > 0 else 0
                    if gap_pct > 50 and p['studentCount'] >= 5:
                        priority = 'High'
                    elif gap_pct > 30 or p['studentCount'] >= 5:
                        priority = 'Medium'
                    else:
                        priority = 'Low'
                    vol_factor = min(p['studentCount'] / max(total_students_count, 1) * 100, 100)
                    gap_factor = min(gap_pct, 100)
                    confidence = round(min((vol_factor * 0.4 + gap_factor * 0.6), 100))
                    negotiation_opps.append({
                        'provider': p['provider'],
                        'currentAvg': round2(p['avgCommPerStudent']),
                        'benchmarkAvg': round2(benchmark_avg),
                        'studentCount': p['studentCount'],
                        'gap': round2(benchmark_avg - p['avgCommPerStudent']),
                        'gapPct': round(gap_pct, 1),
                        'uplift5pct': round2(uplift_5pct),
                        'uplift10pct': round2(uplift_10pct),
                        'priority': priority,
                        'confidence': confidence,
                    })
            negotiation_opps.sort(key=lambda x: x['uplift10pct'], reverse=True)

            focus_opportunities = []
            agents_with_multi_provs = {}
            for (agent, prov), apd in agent_provider_data.items():
                if agent not in agents_with_multi_provs:
                    agents_with_multi_provs[agent] = []
                margin = apd['commission'] - apd['subPaid']
                pct = (margin / apd['commission'] * 100) if apd['commission'] > 0 else 0
                avg_margin_per = (margin / len(apd['students'])) if apd['students'] else 0
                agents_with_multi_provs[agent].append({
                    'provider': prov, 'margin': margin, 'marginPct': round(pct, 1),
                    'students': len(apd['students']),
                    'avgComm': round2(apd['commission'] / len(apd['students'])) if apd['students'] else 0,
                    'avgMarginPerStudent': round2(avg_margin_per),
                })
            for agent, provs in agents_with_multi_provs.items():
                if len(provs) < 2:
                    continue
                provs.sort(key=lambda x: x['marginPct'], reverse=True)
                best_p = provs[0]
                for wp in provs[1:]:
                    if wp['marginPct'] < best_p['marginPct'] * 0.5 and wp['students'] >= 2 and best_p['marginPct'] > 30:
                        margin_diff_per_student = best_p['avgMarginPerStudent'] - wp['avgMarginPerStudent']
                        if margin_diff_per_student > 0:
                            potential_gain = margin_diff_per_student * wp['students']
                            focus_opportunities.append({
                                'agent': agent,
                                'lowMarginProvider': wp['provider'],
                                'highMarginProvider': best_p['provider'],
                                'lowMarginPct': wp['marginPct'],
                                'highMarginPct': best_p['marginPct'],
                                'currentVolume': wp['students'],
                                'marginDiffPerApp': round2(margin_diff_per_student),
                                'potentialGain': round2(potential_gain),
                                'recommendation': f'Encourage {agent} to prioritize future applications with {best_p["provider"]} instead of {wp["provider"]}, based on significantly higher margin performance ({best_p["marginPct"]}% vs {wp["marginPct"]}%).',
                            })
            focus_opportunities.sort(key=lambda x: x['potentialGain'], reverse=True)

            top_loss_areas = []
            for ap in agent_provider_pairs:
                if ap['marginPct'] < 20 and ap['subAgentPaid'] > 0 and ap['studentCount'] >= 2:
                    loss_vs_avg = (avg_margin_per_student * ap['studentCount']) - ap['margin'] if ap['margin'] < avg_margin_per_student * ap['studentCount'] else 0
                    if loss_vs_avg > 0:
                        top_loss_areas.append({
                            'agent': ap['agent'], 'provider': ap['provider'],
                            'marginLoss': round2(loss_vs_avg),
                            'marginPct': ap['marginPct'],
                            'subAgentRatePct': ap['subAgentRatePct'],
                            'studentCount': ap['studentCount'],
                            'commission': ap['commission'],
                            'subAgentPaid': ap['subAgentPaid'],
                        })
            top_loss_areas.sort(key=lambda x: x['marginLoss'], reverse=True)

            intake_intelligence = []
            for tn in term_names:
                td = student_term_data.get(tn, {'commission': 0, 'subPaid': 0, 'students': set()})
                comm = td['commission']
                sub = td['subPaid']
                margin_t = comm - sub
                pct_t = (margin_t / comm * 100) if comm > 0 else 0
                sc = len(td['students'])
                intake_intelligence.append({
                    'term': tn, 'commission': round2(comm), 'subAgentPaid': round2(sub),
                    'margin': round2(margin_t), 'marginPct': round(pct_t, 1), 'studentCount': sc,
                    'avgPerStudent': round2(comm / sc) if sc > 0 else 0,
                })

            suggestions = []

            if total_payout_recovery > 0:
                top_po = payout_optimizations[0] if payout_optimizations else None
                risk_count = sum(1 for p in payout_optimizations if p['badge'] == 'High Payout Risk')
                msg = f'Optimizing sub-agent payout ratios could recover ~${total_payout_recovery:,.2f} in annual margin.'
                if top_po:
                    msg += f' Highest impact: {top_po["agent"]} (current {top_po["currentPayoutPct"]}% → recommended {top_po["recommendedPayoutPct"]}%, +${top_po["expectedGain"]:,.2f}).'
                if risk_count > 0:
                    msg += f' {risk_count} agent(s) flagged as High Payout Risk.'
                suggestions.append({'type': 'danger', 'title': 'Sub-Agent Payout Optimization', 'message': msg})

            if prev_total > 0:
                prev_margin_val = prev_total - prev_sub_paid
                margin_change = total_margin - prev_margin_val
                prev_margin_pct_val = (prev_margin_val / prev_total * 100) if prev_total > 0 else 0
                margin_pct_change = margin_pct - prev_margin_pct_val
                if margin_change > 0:
                    suggestions.append({'type': 'success', 'title': 'Net Margin Growth',
                        'message': f'Net margin improved by ${margin_change:,.2f} vs {target_year - 1} (${prev_margin_val:,.2f} → ${total_margin:,.2f}). Margin rate {"also improved" if margin_pct_change > 0 else "held steady"} at {round(margin_pct, 1)}% (was {round(prev_margin_pct_val, 1)}%).'})
                elif margin_change < 0:
                    suggestions.append({'type': 'warning', 'title': 'Net Margin Decline',
                        'message': f'Net margin dropped ${abs(margin_change):,.2f} vs {target_year - 1} (${prev_margin_val:,.2f} → ${total_margin:,.2f}, {round(prev_margin_pct_val, 1)}% → {round(margin_pct, 1)}%). Review sub-agent payout ratios and provider rates to recover margin.'})

            if top_loss_areas:
                tla = top_loss_areas[0]
                suggestions.append({'type': 'danger', 'title': 'Biggest Margin Loss Driver',
                    'message': f'{tla["agent"]} + {tla["provider"]} combination causing ${tla["marginLoss"]:,.2f} margin loss ({tla["studentCount"]} applications at {tla["marginPct"]}% margin, {tla["subAgentRatePct"]}% payout). Address payout structure and provider rates for this pair.'})

            negotiate_provs = [p for p in provider_insights if p['aiAction'] == 'Negotiate']
            if negotiate_provs:
                names = ', '.join(p['provider'] for p in negotiate_provs[:3])
                total_gap = sum((avg_comm_all - p['avgCommPerStudent']) * p['studentCount'] for p in negotiate_provs if p['avgCommPerStudent'] < avg_comm_all)
                suggestions.append({'type': 'info', 'title': 'Provider Negotiation Opportunity',
                    'message': f'{names} pay below-average commission per application. Closing the gap to the ${avg_comm_all:,.2f} benchmark through rate negotiation could recover ~${total_gap:,.2f} in margin.'})

            scale_provs = [p for p in provider_insights if p['aiAction'] == 'Scale' and p['marginPct'] > 50]
            if scale_provs:
                names = ', '.join(p['provider'] for p in scale_provs[:3])
                total_pot = sum(p['margin'] * 0.25 for p in scale_provs)
                suggestions.append({'type': 'success', 'title': 'High-Margin Growth Targets',
                    'message': f'{names} have >50% margin. Encouraging agents to increase future application volume by 25% with these providers could add ~${total_pot:,.2f} in net margin.'})

            if focus_opportunities:
                total_focus_gain = sum(r['potentialGain'] for r in focus_opportunities[:5])
                top_r = focus_opportunities[0]
                suggestions.append({'type': 'info', 'title': 'Application Focus Opportunity',
                    'message': f'{top_r["agent"]} earns {top_r["highMarginPct"]}% margin with {top_r["highMarginProvider"]} but only {top_r["lowMarginPct"]}% with {top_r["lowMarginProvider"]}. Guiding future application volume toward higher-margin providers could improve margin by ~${total_focus_gain:,.2f}.'})

            if leakage_alerts:
                total_leak = sum(l['estimatedLoss'] for l in leakage_alerts[:5])
                agent_leaks = [l for l in leakage_alerts if l['entityType'] == 'agent']
                prov_leaks = [l for l in leakage_alerts if l['entityType'] == 'provider']
                detail_parts = []
                if agent_leaks:
                    detail_parts.append(f'{len(agent_leaks)} agent(s) with high volume but low margin')
                if prov_leaks:
                    detail_parts.append(f'{len(prov_leaks)} provider(s) with below-benchmark rates')
                suggestions.append({'type': 'warning', 'title': 'Margin Leakage',
                    'message': f'~${total_leak:,.2f} in estimated margin leakage: {" and ".join(detail_parts)}. Review sub-agent payout arrangements and negotiate stronger provider rates to recover margin on future business.'})

            if overpaid_students:
                total_loss = sum(o['loss'] for o in overpaid_students)
                suggestions.append({'type': 'danger', 'title': 'Sub-Agent Overpayment',
                    'message': f'{len(overpaid_students)} application(s) where sub-agent payout exceeds commission earned, totalling ${total_loss:,.2f} in net loss. Immediately review payout records for these cases.'})

            filled_intakes = [t for t in intake_intelligence if t['studentCount'] > 0]
            empty_intakes = [t for t in intake_intelligence if t['studentCount'] == 0]
            if filled_intakes and len(filled_intakes) >= 2:
                best_intake = max(filled_intakes, key=lambda t: t['margin'])
                worst_intake = min(filled_intakes, key=lambda t: t['marginPct'])
                if best_intake['marginPct'] - worst_intake['marginPct'] > 15:
                    gap_value = (best_intake['margin'] / best_intake['studentCount'] - worst_intake['margin'] / worst_intake['studentCount']) * worst_intake['studentCount'] if worst_intake['studentCount'] > 0 and best_intake['studentCount'] > 0 else 0
                    if gap_value > 0:
                        suggestions.append({'type': 'info', 'title': 'Intake Margin Gap',
                            'message': f'{best_intake["term"]} yields {best_intake["marginPct"]}% margin vs {worst_intake["term"]} at {worst_intake["marginPct"]}%. Improving provider mix and payout structures for the weaker intake could add ~${gap_value:,.2f} in margin.'})

            if empty_intakes and filled_intakes:
                avg_margin_per_intake = sum(t['margin'] for t in filled_intakes) / len(filled_intakes)
                potential = avg_margin_per_intake * len(empty_intakes)
                if potential > 0:
                    names = ', '.join(t['term'] for t in empty_intakes[:3])
                    suggestions.append({'type': 'info', 'title': 'Unfilled Intakes',
                        'message': f'{names} — {len(empty_intakes)} intake(s) with no commission yet. Growing future application volume in these intakes could add ~${potential:,.2f} in net margin based on current averages.'})

            top_10_provs = sorted(provider_insights, key=lambda p: p['opportunityScore'], reverse=True)[:10]
            top_10_agents = sorted(agent_insights, key=lambda a: a['opportunityScore'], reverse=True)[:10]

            trend_data = None
            if prev_total > 0:
                prev_margin_val = prev_total - prev_sub_paid
                prev_margin_pct = (prev_margin_val / prev_total * 100) if prev_total > 0 else 0
                prev_avg = prev_total / prev_students_count if prev_students_count > 0 else 0
                trend_data = {
                    'prevCommission': round2(prev_total),
                    'prevSubPaid': round2(prev_sub_paid),
                    'prevMargin': round2(prev_margin_val),
                    'prevMarginPct': round(prev_margin_pct, 1),
                    'prevStudents': prev_students_count,
                    'prevAvgPerStudent': round2(prev_avg),
                    'commissionChange': round2(total_commission - prev_total),
                    'commissionChangePct': round(((total_commission - prev_total) / prev_total * 100) if prev_total > 0 else 0, 1),
                    'marginChange': round2(total_margin - prev_margin_val),
                    'marginChangePct': round(((total_margin - prev_margin_val) / prev_margin_val * 100) if prev_margin_val > 0 else 0, 1),
                    'studentChange': total_students_count - prev_students_count,
                }

            return Response({
                'insights': {
                    'totalCommission': round2(total_commission),
                    'totalSubAgentPaid': round2(total_sub_paid),
                    'totalMargin': round2(total_margin),
                    'marginPct': round(margin_pct, 1),
                    'totalStudents': total_students_count,
                    'avgPerStudent': round2(avg_comm_all),
                    'byProvider': provider_insights,
                    'byAgent': agent_insights,
                    'overpaidStudents': overpaid_students,
                    'suggestions': suggestions,
                    'topProviders': top_10_provs,
                    'topAgents': top_10_agents,
                    'agentProviderPairs': agent_provider_pairs[:20],
                    'leakageAlerts': leakage_alerts[:10],
                    'negotiationOpps': negotiation_opps[:10],
                    'focusOpportunities': focus_opportunities[:10],
                    'payoutOptimizations': payout_optimizations[:20],
                    'apPayoutOptimizations': ap_payout_optimizations[:15],
                    'totalPayoutRecovery': round2(total_payout_recovery),
                    'topLossAreas': top_loss_areas[:5],
                    'intakeIntelligence': intake_intelligence,
                    'trendData': trend_data,
                }
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'message': str(e)}, status=500)


class PredictionView(APIView):
    @require_permission("commission_tracker.student.read")
    def get(self, request, year):
        try:
            user_perms = request.session.get('userPermissions', [])
            has_financials = 'commission_tracker.field.financials' in user_perms
            if not has_financials:
                return Response({'prediction': None, 'message': 'No financial access'})

            target_year = int(year)
            all_terms_raw = list(CommissionTerm.objects.all().order_by('year', 'sort_order'))
            target_terms = [t for t in all_terms_raw if t.year == target_year]
            all_past_years = sorted(set(t.year for t in all_terms_raw if t.year < target_year))
            past_years = all_past_years[-5:] if len(all_past_years) > 5 else all_past_years
            all_terms = [t for t in all_terms_raw if t.year in past_years or t.year == target_year]

            if not past_years:
                return Response({'prediction': None, 'message': 'No historical data available'})

            past_term_names = [t.term_name for t in all_terms if t.year < target_year]
            target_term_names = [t.term_name for t in target_terms]
            term_number_map = {t.term_name: t.term_number for t in all_terms}
            term_year_map = {t.term_name: t.year for t in all_terms}

            excluded_ids = get_excluded_student_ids_for_year(target_year)

            past_entries = list(CommissionEntry.objects.filter(term_name__in=past_term_names))
            target_entries = list(CommissionEntry.objects.filter(term_name__in=target_term_names))
            if excluded_ids:
                target_entries = [e for e in target_entries if e.commission_student_id not in excluded_ids]

            past_student_ids = set(e.commission_student_id for e in past_entries)
            target_student_ids = set(e.commission_student_id for e in target_entries)

            all_students = {s.id: s for s in CommissionStudent.objects.filter(
                id__in=past_student_ids | target_student_ids
            )}

            all_master_qs = CommissionStudent.objects.all()
            if excluded_ids:
                all_master_qs = all_master_qs.exclude(id__in=excluded_ids)
            current_year_students = {s.id: s for s in all_master_qs}
            all_students.update(current_year_students)

            provider_comm_rules = {}
            try:
                from provider_commission.models import ProviderCommissionEntry
                for pce in ProviderCommissionEntry.objects.filter(is_active=True):
                    prov_key = (pce.provider_name or '').strip().lower()
                    level_key = (pce.degree_level or 'any').strip().lower()
                    if prov_key not in provider_comm_rules:
                        provider_comm_rules[prov_key] = {}
                    provider_comm_rules[prov_key][level_key] = {
                        'basis': pce.commission_basis or 'full_course',
                        'followup_year_rates': pce.followup_year_rates,
                    }
            except Exception:
                provider_comm_rules = {}

            def normalize_course_level(course_level):
                cl = (course_level or '').strip().lower()
                if not cl:
                    return 'any'
                undergrad_terms = ['bachelor', 'undergraduate', 'bachelors', "bachelor's", 'bsc', 'ba', 'beng', 'bcom', 'undergrad']
                postgrad_terms = ['master', 'postgraduate', 'masters', "master's", 'msc', 'ma', 'mba', 'meng', 'postgrad', 'graduate diploma', 'graduate certificate', 'grad dip', 'grad cert']
                phd_terms = ['phd', 'doctorate', 'doctoral', 'ph.d']
                vet_terms = ['vet', 'vocational', 'certificate iii', 'certificate iv', 'cert iii', 'cert iv', 'advanced diploma']
                diploma_terms = ['diploma']
                foundation_terms = ['foundation', 'pathway']
                english_terms = ['english', 'elicos', 'esl', 'ielts prep']
                for term in undergrad_terms:
                    if term in cl:
                        return 'undergraduate'
                for term in postgrad_terms:
                    if term in cl:
                        return 'postgraduate'
                for term in phd_terms:
                    if term in cl:
                        return 'phd'
                for term in vet_terms:
                    if term in cl:
                        return 'vet'
                for term in diploma_terms:
                    if term in cl:
                        return 'diploma'
                for term in foundation_terms:
                    if term in cl:
                        return 'foundation'
                for term in english_terms:
                    if term in cl:
                        return 'english'
                return cl

            def get_max_payable_terms(provider, course_level):
                prov_key = (provider or '').strip().lower()
                level_key = normalize_course_level(course_level)
                rule = None
                if prov_key in provider_comm_rules:
                    rule = provider_comm_rules[prov_key].get(level_key) or provider_comm_rules[prov_key].get('any')
                if not rule:
                    return 99
                basis = rule['basis']
                if basis == 'one_time':
                    return 1
                if basis == '1_year':
                    return 2
                if basis == '2_semesters':
                    return 2
                if basis == 'per_trimester':
                    return 99
                if basis == 'per_semester':
                    return 99
                if basis == 'per_year':
                    return 99
                if basis == 'full_course':
                    return 99
                return 99

            relevant_student_ids = set(current_year_students.keys()) | past_student_ids | target_student_ids
            all_entries_by_student = {}
            for e in CommissionEntry.objects.filter(commission_student_id__in=relevant_student_ids):
                sid = e.commission_student_id
                if sid not in all_entries_by_student:
                    all_entries_by_student[sid] = []
                all_entries_by_student[sid].append(e)

            def count_terms_already_paid(sid):
                entries = all_entries_by_student.get(sid, [])
                terms_paid = set()
                for e in entries:
                    if float(e.commission_amount or 0) > 0:
                        terms_paid.add(e.term_name)
                return len(terms_paid)

            eligible_students = {}
            ineligible_students = {}
            ineligible_reasons = {}
            for sid, s in current_year_students.items():
                status = (s.status or 'Under Enquiry').lower()
                if status == 'withdrawn':
                    ineligible_students[sid] = s
                    ineligible_reasons[sid] = 'Withdrawn'
                    continue
                provider = (s.provider or '').strip()
                course_level = (s.course_level or '').strip()
                max_terms = get_max_payable_terms(provider, course_level)
                terms_paid = count_terms_already_paid(sid)
                if status == 'complete' and terms_paid >= max_terms:
                    ineligible_students[sid] = s
                    ineligible_reasons[sid] = 'Complete - commission basis exhausted'
                    continue
                if terms_paid >= max_terms and max_terms < 99:
                    ineligible_students[sid] = s
                    ineligible_reasons[sid] = f'Commission basis exhausted ({terms_paid}/{max_terms} terms paid)'
                    continue
                eligible_students[sid] = s

            hist_by_year_term = {}
            hist_student_bonus_years = {}
            for e in past_entries:
                s = all_students.get(e.commission_student_id)
                if not s:
                    continue
                tn = term_number_map.get(e.term_name, 0)
                yr = term_year_map.get(e.term_name, 0)
                provider = (s.provider or '').strip()
                course = (s.course_name or '').strip()
                country = (s.country or '').strip()
                level = (s.course_level or '').strip()
                comm = float(e.commission_amount or 0)
                bonus = float(e.bonus or 0)

                key = (yr, e.commission_student_id)
                if key not in hist_by_year_term:
                    hist_by_year_term[key] = {}
                if tn not in hist_by_year_term[key]:
                    hist_by_year_term[key][tn] = {'comm': 0, 'bonus': 0}
                hist_by_year_term[key][tn]['comm'] += comm
                hist_by_year_term[key][tn]['bonus'] += bonus

                if bonus > 0:
                    if e.commission_student_id not in hist_student_bonus_years:
                        hist_student_bonus_years[e.commission_student_id] = set()
                    hist_student_bonus_years[e.commission_student_id].add(yr)

            hist_term_ratios = {}
            for (yr, sid), term_data in hist_by_year_term.items():
                s = all_students.get(sid)
                if not s:
                    continue
                provider = (s.provider or '').strip()
                course = (s.course_name or '').strip()
                sorted_terms = sorted(term_data.keys())
                for i, from_tn in enumerate(sorted_terms):
                    for to_tn in sorted_terms[i+1:]:
                        from_comm = term_data[from_tn]['comm']
                        to_comm = term_data[to_tn]['comm']
                        if from_comm > 0:
                            ratio = to_comm / from_comm
                            ratio_key_pc = (provider, course, from_tn, to_tn) if provider and course else None
                            ratio_key_p = (provider, from_tn, to_tn) if provider else None
                            ratio_key_g = (from_tn, to_tn)
                            for rk in [ratio_key_pc, ratio_key_p, ratio_key_g]:
                                if rk:
                                    if rk not in hist_term_ratios:
                                        hist_term_ratios[rk] = []
                                    hist_term_ratios[rk].append(ratio)

            def _median(lst):
                if not lst:
                    return None
                s = sorted(lst)
                n = len(s)
                if n % 2 == 0:
                    return (s[n//2 - 1] + s[n//2]) / 2
                return s[n//2]

            hist_bonus_rate_by_prov = {}
            hist_bonus_rate_global = {'total_students': 0, 'bonus_students': 0}
            for (yr, sid), term_data in hist_by_year_term.items():
                s = all_students.get(sid)
                if not s:
                    continue
                provider = (s.provider or '').strip()
                got_bonus = any(td['bonus'] > 0 for td in term_data.values())
                total_bonus_in_year = sum(td['bonus'] for td in term_data.values())
                if provider:
                    if provider not in hist_bonus_rate_by_prov:
                        hist_bonus_rate_by_prov[provider] = {'total': 0, 'with_bonus': 0, 'avg_bonus': []}
                    hist_bonus_rate_by_prov[provider]['total'] += 1
                    if got_bonus:
                        hist_bonus_rate_by_prov[provider]['with_bonus'] += 1
                        hist_bonus_rate_by_prov[provider]['avg_bonus'].append(total_bonus_in_year)
                hist_bonus_rate_global['total_students'] += 1
                if got_bonus:
                    hist_bonus_rate_global['bonus_students'] += 1

            hist_prov_course = {}
            hist_prov = {}
            hist_country_level = {}
            hist_country = {}
            hist_global = {}
            for e in past_entries:
                s = all_students.get(e.commission_student_id)
                if not s:
                    continue
                provider = (s.provider or '').strip()
                course = (s.course_name or '').strip()
                country = (s.country or '').strip()
                level = (s.course_level or '').strip()
                tn = term_number_map.get(e.term_name, 0)
                comm = float(e.commission_amount or 0)
                bonus = float(e.bonus or 0)

                def _add(d, k):
                    if k not in d:
                        d[k] = {'comm': 0, 'bonus': 0, 'n': 0}
                    d[k]['comm'] += comm
                    d[k]['bonus'] += bonus
                    d[k]['n'] += 1

                if provider and course:
                    _add(hist_prov_course, (provider, course, tn))
                if provider:
                    _add(hist_prov, (provider, tn))
                if country and level:
                    _add(hist_country_level, (country, level, tn))
                if country:
                    _add(hist_country, (country, tn))
                _add(hist_global, tn)

            def _avg(d, k):
                v = d.get(k)
                if v and v['n'] > 0:
                    return v['comm'] / v['n'], v['bonus'] / v['n'], v['n']
                return None, None, 0

            actual_by_term = {}
            actual_students_by_term = {}
            actual_student_comm_by_term = {}
            for t in target_terms:
                actual_by_term[t.term_number] = {'comm': 0, 'bonus': 0, 'count': 0}
                actual_students_by_term[t.term_number] = set()
                actual_student_comm_by_term[t.term_number] = {}
            for e in target_entries:
                tn = term_number_map.get(e.term_name, 0)
                if tn in actual_by_term:
                    comm = float(e.commission_amount or 0)
                    bonus = float(e.bonus or 0)
                    actual_by_term[tn]['comm'] += comm
                    actual_by_term[tn]['bonus'] += bonus
                    actual_by_term[tn]['count'] += 1
                    actual_students_by_term[tn].add(e.commission_student_id)
                    if e.commission_student_id not in actual_student_comm_by_term[tn]:
                        actual_student_comm_by_term[tn][e.commission_student_id] = 0
                    actual_student_comm_by_term[tn][e.commission_student_id] += comm

            students_with_bonus_this_year = set()
            for e in target_entries:
                if float(e.bonus or 0) > 0:
                    students_with_bonus_this_year.add(e.commission_student_id)

            target_term_numbers = sorted([t.term_number for t in target_terms])

            term_predictions = []
            total_predicted_comm = 0
            total_predicted_bonus = 0
            total_actual_comm = 0
            total_actual_bonus = 0
            provider_predicted = {}
            course_predicted = {}
            country_predicted = {}
            level_predicted = {}
            provider_student_ids = {}
            course_student_ids = {}
            country_student_ids = {}
            level_student_ids = {}

            for t in target_terms:
                tn = t.term_number
                actual = actual_by_term.get(tn, {'comm': 0, 'bonus': 0, 'count': 0})
                total_actual_comm += actual['comm']
                total_actual_bonus += actual['bonus']

                pred_comm = actual['comm']
                pred_bonus = actual['bonus']
                student_count = actual['count']
                has_actual_students = actual_students_by_term.get(tn, set())

                students_to_predict = set()
                for sid in eligible_students:
                    if sid not in has_actual_students:
                        s = eligible_students[sid]
                        prov = (s.provider or '').strip()
                        cl = (s.course_level or '').strip()
                        max_t = get_max_payable_terms(prov, cl)
                        paid_so_far = count_terms_already_paid(sid)
                        if paid_so_far < max_t or max_t >= 99:
                            students_to_predict.add(sid)

                source = 'actual' if not students_to_predict and actual['count'] > 0 else 'mixed' if actual['count'] > 0 else 'estimated'

                prior_actual_terms = [pt for pt in target_term_numbers if pt < tn]

                for sid in students_to_predict:
                    s = eligible_students.get(sid) or all_students.get(sid)
                    if not s:
                        continue
                    provider = (s.provider or '').strip()
                    course = (s.course_name or '').strip()
                    country = (s.country or '').strip()
                    level = (s.course_level or '').strip()

                    ratio_pred = None
                    for from_tn in reversed(prior_actual_terms):
                        if sid in actual_student_comm_by_term.get(from_tn, {}):
                            from_comm = actual_student_comm_by_term[from_tn][sid]
                            if from_comm > 0:
                                ratio = None
                                if provider and course:
                                    r = _median(hist_term_ratios.get((provider, course, from_tn, tn), []))
                                    if r is not None:
                                        ratio = r
                                if ratio is None and provider:
                                    r = _median(hist_term_ratios.get((provider, from_tn, tn), []))
                                    if r is not None:
                                        ratio = r
                                if ratio is None:
                                    r = _median(hist_term_ratios.get((from_tn, tn), []))
                                    if r is not None:
                                        ratio = r
                                if ratio is not None:
                                    ratio_pred = from_comm * ratio
                                    break

                    if ratio_pred is not None:
                        pred_c = ratio_pred
                    else:
                        avg_c, avg_b, matched = None, None, 0
                        if provider and course:
                            avg_c, avg_b, matched = _avg(hist_prov_course, (provider, course, tn))
                        if avg_c is None and provider:
                            avg_c, avg_b, matched = _avg(hist_prov, (provider, tn))
                        if avg_c is None and country and level:
                            avg_c, avg_b, matched = _avg(hist_country_level, (country, level, tn))
                        if avg_c is None and country:
                            avg_c, avg_b, matched = _avg(hist_country, (country, tn))
                        if avg_c is None:
                            avg_c, avg_b, matched = _avg(hist_global, tn)
                        pred_c = avg_c if avg_c and avg_c > 0 else 0

                    pred_b = 0
                    if tn == min(target_term_numbers):
                        student_entries = all_entries_by_student.get(sid, [])
                        bonus_paid_target_year = sum(
                            float(e.bonus or 0) for e in student_entries
                            if e.term_name in target_term_names
                        )
                        bonus_paid_prior_years = sum(
                            float(e.bonus or 0) for e in student_entries
                            if e.term_name not in target_term_names
                        )
                        if provider and provider in hist_bonus_rate_by_prov:
                            bp = hist_bonus_rate_by_prov[provider]
                            if bp['total'] > 0 and bp['with_bonus'] > 0:
                                avg_bonus_val = sum(bp['avg_bonus']) / len(bp['avg_bonus']) if bp['avg_bonus'] else 0
                                if bonus_paid_target_year > 0 and avg_bonus_val > 0:
                                    remaining = max(avg_bonus_val - bonus_paid_target_year, 0)
                                    pred_b = remaining
                                elif bonus_paid_target_year == 0:
                                    if bonus_paid_prior_years > 0 and avg_bonus_val > 0:
                                        remaining = max(avg_bonus_val - bonus_paid_prior_years, 0)
                                        pred_b = remaining
                                    else:
                                        bonus_prob = bp['with_bonus'] / bp['total']
                                        pred_b = avg_bonus_val * bonus_prob

                    if pred_c > 0:
                        pred_comm += pred_c
                        pred_bonus += pred_b
                        student_count += 1

                        prov_key = provider or 'Unknown'
                        if prov_key not in provider_predicted:
                            provider_predicted[prov_key] = {'predicted': 0, 'actual': 0}
                            provider_student_ids[prov_key] = set()
                        provider_predicted[prov_key]['predicted'] += pred_c
                        provider_student_ids[prov_key].add(sid)

                        course_key = course or 'Unknown'
                        if course_key not in course_predicted:
                            course_predicted[course_key] = {'predicted': 0, 'actual': 0}
                            course_student_ids[course_key] = set()
                        course_predicted[course_key]['predicted'] += pred_c
                        course_student_ids[course_key].add(sid)

                        country_key = country or 'Unknown'
                        if country_key not in country_predicted:
                            country_predicted[country_key] = {'predicted': 0, 'actual': 0}
                            country_student_ids[country_key] = set()
                        country_predicted[country_key]['predicted'] += pred_c
                        country_student_ids[country_key].add(sid)

                        level_key = level or 'Unknown'
                        if level_key not in level_predicted:
                            level_predicted[level_key] = {'predicted': 0, 'actual': 0}
                            level_student_ids[level_key] = set()
                        level_predicted[level_key]['predicted'] += pred_c
                        level_student_ids[level_key].add(sid)

                for sid in has_actual_students:
                    s = all_students.get(sid)
                    if not s:
                        continue
                    prov_key = (s.provider or 'Unknown').strip()
                    if prov_key not in provider_predicted:
                        provider_predicted[prov_key] = {'predicted': 0, 'actual': 0}
                        provider_student_ids[prov_key] = set()
                    sid_actual_comm = sum(float(e.commission_amount or 0) for e in target_entries
                                         if e.commission_student_id == sid and term_number_map.get(e.term_name) == tn)
                    provider_predicted[prov_key]['actual'] += sid_actual_comm

                    course_key = (s.course_name or 'Unknown').strip()
                    if course_key not in course_predicted:
                        course_predicted[course_key] = {'predicted': 0, 'actual': 0}
                        course_student_ids[course_key] = set()
                    course_predicted[course_key]['actual'] += sid_actual_comm

                    country_key = (s.country or 'Unknown').strip()
                    if country_key not in country_predicted:
                        country_predicted[country_key] = {'predicted': 0, 'actual': 0}
                        country_student_ids[country_key] = set()
                    country_predicted[country_key]['actual'] += sid_actual_comm

                    level_key = (s.course_level or 'Unknown').strip()
                    if level_key not in level_predicted:
                        level_predicted[level_key] = {'predicted': 0, 'actual': 0}
                        level_student_ids[level_key] = set()
                    level_predicted[level_key]['actual'] += sid_actual_comm

                total_predicted_comm += pred_comm
                total_predicted_bonus += pred_bonus

                term_predictions.append({
                    'termNumber': tn,
                    'termName': t.term_name,
                    'termLabel': t.term_label,
                    'expectedCommission': round2(pred_comm),
                    'expectedBonus': round2(pred_bonus),
                    'expectedTotal': round2(pred_comm + pred_bonus),
                    'actualCommission': round2(actual['comm']),
                    'actualBonus': round2(actual['bonus']),
                    'studentCount': student_count,
                    'estimatedStudents': len(students_to_predict),
                    'actualStudents': actual['count'],
                    'source': source,
                })

            provider_list = []
            for prov, data in sorted(provider_predicted.items(), key=lambda x: -(x[1]['predicted'] + x[1]['actual'])):
                provider_list.append({
                    'provider': prov,
                    'estimatedCommission': round2(data['predicted']),
                    'actualCommission': round2(data['actual']),
                    'totalExpected': round2(data['predicted'] + data['actual']),
                    'eligibleStudents': len(provider_student_ids.get(prov, set())),
                })

            course_list = []
            for crs, data in sorted(course_predicted.items(), key=lambda x: -(x[1]['predicted'] + x[1]['actual'])):
                course_list.append({
                    'course': crs,
                    'estimatedCommission': round2(data['predicted']),
                    'actualCommission': round2(data['actual']),
                    'totalExpected': round2(data['predicted'] + data['actual']),
                    'eligibleStudents': len(course_student_ids.get(crs, set())),
                })

            hist_provider_summary = {}
            hist_course_summary = {}
            hist_country_summary = {}
            hist_level_summary = {}
            for e in past_entries:
                s = all_students.get(e.commission_student_id)
                if not s:
                    continue
                prov = (s.provider or 'Unknown').strip()
                crs = (s.course_name or 'Unknown').strip()
                country = (s.country or 'Unknown').strip()
                level = (s.course_level or 'Unknown').strip()
                comm = float(e.commission_amount or 0)

                if prov not in hist_provider_summary:
                    hist_provider_summary[prov] = {'total': 0, 'n': 0}
                hist_provider_summary[prov]['total'] += comm
                hist_provider_summary[prov]['n'] += 1

                if crs not in hist_course_summary:
                    hist_course_summary[crs] = {'total': 0, 'n': 0}
                hist_course_summary[crs]['total'] += comm
                hist_course_summary[crs]['n'] += 1

                if country not in hist_country_summary:
                    hist_country_summary[country] = {'total': 0, 'n': 0}
                hist_country_summary[country]['total'] += comm
                hist_country_summary[country]['n'] += 1

                if level not in hist_level_summary:
                    hist_level_summary[level] = {'total': 0, 'n': 0}
                hist_level_summary[level]['total'] += comm
                hist_level_summary[level]['n'] += 1

            def _build_hist_list(summary, key_name):
                return sorted([
                    {key_name: k, 'avgCommission': round2(v['total'] / v['n']), 'entries': v['n'], 'totalCommission': round2(v['total'])}
                    for k, v in summary.items() if v['n'] > 0
                ], key=lambda x: -x['totalCommission'])

            country_list = sorted([
                {'country': k, 'estimatedCommission': round2(v['predicted']), 'actualCommission': round2(v['actual']),
                 'totalExpected': round2(v['predicted'] + v['actual']), 'eligibleStudents': len(country_student_ids.get(k, set()))}
                for k, v in country_predicted.items() if v['predicted'] + v['actual'] > 0
            ], key=lambda x: -x['totalExpected'])

            level_list = sorted([
                {'studyLevel': k, 'estimatedCommission': round2(v['predicted']), 'actualCommission': round2(v['actual']),
                 'totalExpected': round2(v['predicted'] + v['actual']), 'eligibleStudents': len(level_student_ids.get(k, set()))}
                for k, v in level_predicted.items() if v['predicted'] + v['actual'] > 0
            ], key=lambda x: -x['totalExpected'])

            ineligible_summary = {}
            for sid, reason in ineligible_reasons.items():
                if reason not in ineligible_summary:
                    ineligible_summary[reason] = 0
                ineligible_summary[reason] += 1
            if excluded_ids:
                ineligible_summary['Excluded (prior-year terminal)'] = len(excluded_ids)

            return Response({
                'prediction': {
                    'year': target_year,
                    'basedOnYears': past_years,
                    'totalExpectedCommission': round2(total_predicted_comm),
                    'totalExpectedBonus': round2(total_predicted_bonus),
                    'totalExpectedReceivable': round2(total_predicted_comm + total_predicted_bonus),
                    'totalActualCommission': round2(total_actual_comm),
                    'totalActualBonus': round2(total_actual_bonus),
                    'terms': term_predictions,
                    'byProvider': provider_list,
                    'byCourse': course_list,
                    'byCountry': country_list,
                    'byStudyLevel': level_list,
                    'histByProvider': _build_hist_list(hist_provider_summary, 'provider'),
                    'histByCourse': _build_hist_list(hist_course_summary, 'course'),
                    'histByCountry': _build_hist_list(hist_country_summary, 'country'),
                    'histByStudyLevel': _build_hist_list(hist_level_summary, 'studyLevel'),
                    'eligibleStudents': len(eligible_students),
                    'excludedStudents': len(ineligible_students) + len(excluded_ids),
                    'excludedReasons': ineligible_summary,
                    'studentsConsidered': len(current_year_students),
                },
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'prediction': None, 'message': f'Prediction error: {str(e)}'}, status=200)


def _parse_bulk_row(row, i, col_map=None):
    def g(key, default=''):
        if col_map:
            mapped_key = col_map.get(key, key)
            val = row.get(mapped_key)
        else:
            val = row.get(key)
        if val is None or (isinstance(val, str) and val.strip() == ''):
            return str(default)
        return str(val).strip()

    rate_raw = g('Commission Rate (%)')
    if rate_raw:
        try:
            rate_val = float(rate_raw)
            if rate_val > 1:
                rate_raw = str(rate_val)
            else:
                rate_raw = str(rate_val * 100)
        except (ValueError, TypeError):
            pass

    gst_raw = g('GST Rate (%)', '10')
    if gst_raw:
        try:
            gst_val = float(gst_raw)
            if gst_val <= 1:
                gst_raw = str(gst_val * 100)
        except (ValueError, TypeError):
            pass

    return {
        'agentName': g('Agent Name'),
        'studentId': g('Student ID'),
        'agentsicId': g('Agentsic ID') or g('AgentSIC ID'),
        'studentName': g('Student Name'),
        'provider': g('Provider'),
        'country': g('Country', 'AU'),
        'startIntake': g('Start Intake'),
        'courseLevel': g('Course Level'),
        'courseName': g('Course Name'),
        'courseDurationYears': g('Course Duration (Years)'),
        'commissionRatePct': rate_raw,
        'gstRatePct': gst_raw,
        'gstApplicable': g('GST Applicable', 'Yes'),
        'scholarshipType': g('Scholarship Type', 'None'),
        'scholarshipValue': g('Scholarship Value', '0'),
        'status': g('Status', 'Under Enquiry'),
        'notes': g('Notes'),
        'rowIndex': i,
    }


class SampleSheetView(APIView):
    @require_auth
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'MASTER'
        headers = [
            'S.No', 'Agent Name', 'Student ID', 'Agentsic ID', 'Student Name',
            'Provider', 'Country', 'Start Intake',
            'Course Level (Diploma/Bachelor/Master)', 'Course Name',
            'Course Duration (Years)', 'Commission Rate (%)',
            'GST Rate (%)', 'GST Applicable', 'Scholarship Type',
            'Scholarship Value', 'Status', 'Notes',
        ]
        ws.append(headers)
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.append([
            1, 'Sample Agent', 'STU001', 'ASIC100001', 'John Doe',
            'University of Melbourne', 'AU', 'T1 2025',
            'Bachelor', 'Bachelor of Engineering', 4, 20,
            10, 'Yes', 'None', 0, 'Under Enquiry', '',
        ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="commission_tracker_sample.xlsx"'
        return response


class BulkUploadPreviewView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    @require_permission("commission_tracker.student.add")
    def post(self, request):
        try:
            file = request.FILES.get('file')
            if not file:
                return Response({'message': 'No file provided'}, status=400)

            filename = getattr(file, 'name', '')
            ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
            if ext not in ('xlsx', 'xls', 'csv'):
                return Response({'message': 'Only .xlsx, .xls, or .csv files are accepted'}, status=400)

            from core.file_security import validate_uploaded_file
            is_safe, security_msg = validate_uploaded_file(
                file, file.content_type or 'application/octet-stream',
                filename=filename,
                user_id=request.session.get('userId'),
                ip_address=request.META.get('REMOTE_ADDR', ''),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            if not is_safe:
                return Response({'message': f'File rejected: {security_msg}'}, status=400)

            file.seek(0)

            rows = []
            errors = []

            if filename.lower().endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                sheet_name = 'MASTER' if 'MASTER' in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet_name]

                header_row = None
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    header_row = [str(c or '').strip() for c in row]
                if not header_row:
                    return Response({'message': 'Empty spreadsheet'}, status=400)

                col_map = {}
                for h in header_row:
                    hl = h.lower()
                    if 'agent name' in hl:
                        col_map['Agent Name'] = h
                    elif hl == 'student id':
                        col_map['Student ID'] = h
                    elif 'agentsic' in hl:
                        col_map['Agentsic ID'] = h
                    elif 'student name' in hl:
                        col_map['Student Name'] = h
                    elif hl == 'provider' or 'provider' in hl and 'auto' not in hl:
                        col_map['Provider'] = h
                    elif hl == 'country' or 'country' in hl and 'auto' not in hl:
                        col_map['Country'] = h
                    elif 'start intake' in hl:
                        col_map['Start Intake'] = h
                    elif 'course level' in hl:
                        col_map['Course Level'] = h
                    elif 'course name' in hl:
                        col_map['Course Name'] = h
                    elif 'course duration' in hl:
                        col_map['Course Duration (Years)'] = h
                    elif 'commission rate' in hl and 'override' not in hl and 'used' not in hl and 'auto' not in hl:
                        col_map['Commission Rate (%)'] = h
                    elif 'gst rate' in hl:
                        col_map['GST Rate (%)'] = h
                    elif 'gst applicable' in hl:
                        col_map['GST Applicable'] = h
                    elif hl == 'scholarship type' or ('scholarship type' in hl and 'override' not in hl and 'used' not in hl and 'auto' not in hl):
                        col_map['Scholarship Type'] = h
                    elif hl == 'scholarship value' or ('scholarship value' in hl and 'override' not in hl and 'used' not in hl and 'auto' not in hl):
                        col_map['Scholarship Value'] = h
                    elif hl == 'status':
                        col_map['Status'] = h
                    elif hl == 'notes':
                        col_map['Notes'] = h

                for i, data_row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    row_dict = {}
                    for idx, val in enumerate(data_row):
                        if idx < len(header_row):
                            row_dict[header_row[idx]] = val
                    r = _parse_bulk_row(row_dict, i, col_map)
                    if not r['studentName'] and not r['studentId']:
                        continue
                    if not r['studentName']:
                        errors.append({'row': i + 2, 'message': 'Student Name is required'})
                    if not r['provider']:
                        errors.append({'row': i + 2, 'message': 'Provider is required'})
                    rows.append(r)
                wb.close()
            else:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for i, row in enumerate(reader):
                    r = _parse_bulk_row(row, i)
                    if not r['studentName'] and not r['studentId']:
                        continue
                    if not r['studentName']:
                        errors.append({'row': i + 2, 'message': 'Student Name is required'})
                    if not r['provider']:
                        errors.append({'row': i + 2, 'message': 'Provider is required'})
                    rows.append(r)

            return Response({'rows': rows, 'errors': errors, 'totalRows': len(rows)})
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class BulkUploadConfirmView(APIView):
    @require_permission("commission_tracker.student.add")
    def post(self, request):
        try:
            rows = request.data.get('rows', [])
            user_id = request.session.get('userId')
            created = 0
            skipped = 0
            providers_added = 0
            errors = []
            for i, row in enumerate(rows):
                try:
                    student_id_val = _clean_id((row.get('studentId') or '').strip())
                    provider_val = (row.get('provider') or '').strip()
                    agentsic_id_val = _clean_id((row.get('agentsicId') or '').strip())
                    student_name_val = (row.get('studentName') or '').strip()

                    existing_student = None
                    if student_name_val and agentsic_id_val:
                        existing_student = CommissionStudent.objects.filter(
                            student_name__iexact=student_name_val,
                            agentsic_id__iexact=agentsic_id_val,
                        ).first()

                    if not existing_student and student_id_val:
                        existing_student = CommissionStudent.objects.filter(
                            student_id=student_id_val,
                            student_name__iexact=student_name_val,
                        ).first()

                    if existing_student:
                        already_has_primary = existing_student.provider and existing_student.provider.lower() == provider_val.lower()
                        already_has_secondary = StudentProvider.objects.filter(
                            commission_student_id=existing_student.id,
                            provider__iexact=provider_val,
                        ).exists()
                        if already_has_primary or already_has_secondary:
                            skipped += 1
                            errors.append({'row': i + 1, 'message': f'Duplicate: "{student_name_val}" already has provider "{provider_val}"'})
                            continue

                        dur = row.get('courseDurationYears') or None
                        rate = row.get('commissionRatePct') or None
                        gst = row.get('gstRatePct', 10)
                        schol_val = row.get('scholarshipValue', 0)

                        StudentProvider.objects.create(
                            commission_student_id=existing_student.id,
                            provider=provider_val,
                            student_id=student_id_val or existing_student.student_id,
                            country=row.get('country', 'AU'),
                            course_level=row.get('courseLevel'),
                            course_name=row.get('courseName'),
                            course_duration_years=dur if dur != '' else None,
                            start_intake=row.get('startIntake'),
                            commission_rate_pct=rate if rate != '' else None,
                            gst_rate_pct=gst if gst != '' else 10,
                            gst_applicable=row.get('gstApplicable', 'Yes'),
                            scholarship_type=row.get('scholarshipType', 'None'),
                            scholarship_value=schol_val if schol_val != '' else 0,
                            status=row.get('status', 'Under Enquiry'),
                            created_by_user_id=user_id,
                        )
                        providers_added += 1
                        continue

                    if student_id_val and provider_val:
                        if CommissionStudent.objects.filter(student_id=student_id_val, provider=provider_val).exists():
                            skipped += 1
                            errors.append({'row': i + 1, 'message': f'Duplicate: Student ID "{student_id_val}" already exists for provider "{provider_val}"'})
                            continue

                    dur = row.get('courseDurationYears') or None
                    rate = row.get('commissionRatePct') or None
                    gst = row.get('gstRatePct', 10)
                    schol_val = row.get('scholarshipValue', 0)

                    CommissionStudent.objects.create(
                        agent_name=row.get('agentName', ''),
                        student_id=student_id_val,
                        agentsic_id=agentsic_id_val,
                        student_name=student_name_val,
                        provider=provider_val,
                        country=row.get('country', 'AU'),
                        start_intake=row.get('startIntake'),
                        course_level=row.get('courseLevel'),
                        course_name=row.get('courseName'),
                        course_duration_years=dur if dur != '' else None,
                        commission_rate_pct=rate if rate != '' else None,
                        gst_rate_pct=gst if gst != '' else 10,
                        gst_applicable=row.get('gstApplicable', 'Yes'),
                        scholarship_type=row.get('scholarshipType', 'None'),
                        scholarship_value=schol_val if schol_val != '' else 0,
                        status=row.get('status', 'Under Enquiry'),
                        notes=row.get('notes') or None,
                        created_by_user_id=user_id,
                        updated_by_user_id=user_id,
                    )
                    created += 1
                except Exception as e:
                    errors.append({'row': i + 1, 'message': str(e)})
            return Response({
                'created': created,
                'skipped': skipped,
                'providersAdded': providers_added,
                'errors': errors,
            })
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class ProviderAgreementsMapView(APIView):
    @require_auth
    def get(self, request):
        from agreements.models import Agreement
        from providers.models import Provider
        agreements = Agreement.objects.exclude(status='terminated').values_list('university_id', 'id')
        provider_ids = set(a[0] for a in agreements if a[0])
        providers = {p.id: p.name for p in Provider.objects.filter(id__in=provider_ids)}
        result = {}
        for uid, aid in agreements:
            if uid and uid in providers:
                name = providers[uid]
                if name not in result:
                    result[name] = aid
        return Response(result)
