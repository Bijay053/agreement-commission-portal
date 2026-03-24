import io
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from core.permissions import require_auth, require_permission
from providers.models import Provider
from commissions.models import AgreementCommissionRule
from agreements.models import Agreement
from .models import ProviderCommissionEntry, ProviderCommissionConfig
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def entry_to_dict(e, global_sub_pct=None):
    pct = e.sub_agent_percentage if e.sub_agent_percentage is not None else global_sub_pct
    d = {
        'id': e.id,
        'providerName': e.provider_name,
        'degreeLevel': e.degree_level,
        'territory': e.territory,
        'commissionValue': str(e.commission_value),
        'commissionType': e.commission_type,
        'currency': e.currency,
        'commissionBasis': e.commission_basis,
        'notes': e.notes,
        'isActive': e.is_active,
        'subAgentPercentage': str(e.sub_agent_percentage) if e.sub_agent_percentage is not None else None,
        'copiedFromRuleId': e.copied_from_rule_id,
        'createdAt': e.created_at.isoformat() if e.created_at else None,
        'updatedAt': e.updated_at.isoformat() if e.updated_at else None,
    }
    if pct is not None:
        d['subAgentCommission'] = str(round(e.commission_value * pct / Decimal('100'), 2))
    else:
        d['subAgentCommission'] = None
    d['effectiveSubAgentPercentage'] = str(pct) if pct is not None else None
    return d


class ProviderCommissionConfigView(APIView):
    @require_auth
    def get(self, request):
        config = ProviderCommissionConfig.objects.first()
        if not config:
            config = ProviderCommissionConfig.objects.create(sub_agent_percentage=Decimal('70.00'))
        return Response({
            'subAgentPercentage': str(config.sub_agent_percentage),
            'updatedAt': config.updated_at.isoformat() if config.updated_at else None,
        })

    @require_permission("provider_commission.manage")
    def patch(self, request):
        try:
            pct = request.data.get('subAgentPercentage')
            if pct is None:
                return Response({'message': 'subAgentPercentage is required'}, status=400)
            pct_val = Decimal(str(pct))
            if pct_val < 0 or pct_val > 100:
                return Response({'message': 'Percentage must be between 0 and 100'}, status=400)
            config = ProviderCommissionConfig.objects.first()
            if not config:
                config = ProviderCommissionConfig(sub_agent_percentage=pct_val, updated_by=request.session.get('userId'))
                config.save()
            else:
                config.sub_agent_percentage = pct_val
                config.updated_by = request.session.get('userId')
                config.save()
            return Response({
                'subAgentPercentage': str(config.sub_agent_percentage),
                'updatedAt': config.updated_at.isoformat(),
            })
        except (InvalidOperation, ValueError):
            return Response({'message': 'Invalid percentage value'}, status=400)


class ProviderCommissionListView(APIView):
    @require_permission("provider_commission.view")
    def get(self, request):
        entries = ProviderCommissionEntry.objects.all().order_by('provider_name', 'degree_level')
        search = request.query_params.get('search', '').strip()
        degree_level = request.query_params.get('degreeLevel')
        basis = request.query_params.get('basis')
        active_only = request.query_params.get('activeOnly', 'true')

        if active_only == 'true':
            entries = entries.filter(is_active=True)
        if degree_level:
            entries = entries.filter(degree_level=degree_level)
        if basis:
            entries = entries.filter(commission_basis=basis)

        config = ProviderCommissionConfig.objects.first()
        sub_pct = config.sub_agent_percentage if config else Decimal('70.00')

        result = []
        for e in entries:
            if search and search.lower() not in e.provider_name.lower() and search.lower() not in e.territory.lower():
                continue
            result.append(entry_to_dict(e, sub_pct))
        return Response(result)

    @require_permission("provider_commission.add")
    def post(self, request):
        try:
            data = request.data
            provider_name = (data.get('providerName') or '').strip()
            if not provider_name:
                return Response({'message': 'Provider name is required'}, status=400)

            commission_value = data.get('commissionValue')
            if commission_value is None:
                return Response({'message': 'Commission value is required'}, status=400)

            territory = data.get('territory', '')
            if isinstance(territory, list):
                territory = ','.join(territory)

            sub_pct_raw = data.get('subAgentPercentage')
            sub_pct_val = None
            if sub_pct_raw is not None and str(sub_pct_raw).strip() != '':
                sub_pct_val = Decimal(str(sub_pct_raw))
                if sub_pct_val < 0 or sub_pct_val > 100:
                    return Response({'message': 'Sub-agent percentage must be between 0 and 100'}, status=400)

            entry = ProviderCommissionEntry.objects.create(
                provider_name=provider_name,
                degree_level=data.get('degreeLevel', 'any'),
                territory=territory,
                commission_value=Decimal(str(commission_value)),
                commission_type=data.get('commissionType', 'percentage'),
                currency=data.get('currency', 'AUD'),
                commission_basis=data.get('commissionBasis', 'full_course'),
                notes=data.get('notes', ''),
                is_active=True,
                sub_agent_percentage=sub_pct_val,
                created_by=request.session.get('userId'),
            )
            config = ProviderCommissionConfig.objects.first()
            sub_pct = config.sub_agent_percentage if config else Decimal('70.00')
            return Response(entry_to_dict(entry, sub_pct), status=201)
        except (InvalidOperation, ValueError) as e:
            return Response({'message': f'Invalid value: {e}'}, status=400)
        except Exception as e:
            return Response({'message': str(e)}, status=500)


class ProviderCommissionDetailView(APIView):
    @require_permission("provider_commission.edit")
    def patch(self, request, entry_id):
        try:
            entry = ProviderCommissionEntry.objects.get(id=entry_id)
        except ProviderCommissionEntry.DoesNotExist:
            return Response({'message': 'Entry not found'}, status=404)

        data = request.data
        if 'providerName' in data:
            entry.provider_name = data['providerName']
        if 'degreeLevel' in data:
            entry.degree_level = data['degreeLevel']
        if 'territory' in data:
            territory = data['territory']
            if isinstance(territory, list):
                territory = ','.join(territory)
            entry.territory = territory
        if 'commissionValue' in data:
            entry.commission_value = Decimal(str(data['commissionValue']))
        if 'commissionType' in data:
            entry.commission_type = data['commissionType']
        if 'currency' in data:
            entry.currency = data['currency']
        if 'commissionBasis' in data:
            entry.commission_basis = data['commissionBasis']
        if 'notes' in data:
            entry.notes = data['notes']
        if 'isActive' in data:
            entry.is_active = data['isActive']
        if 'subAgentPercentage' in data:
            sub_pct_raw = data['subAgentPercentage']
            if sub_pct_raw is None or str(sub_pct_raw).strip() == '':
                entry.sub_agent_percentage = None
            else:
                pct_val = Decimal(str(sub_pct_raw))
                if pct_val < 0 or pct_val > 100:
                    return Response({'message': 'Sub-agent percentage must be between 0 and 100'}, status=400)
                entry.sub_agent_percentage = pct_val
        entry.save()

        config = ProviderCommissionConfig.objects.first()
        sub_pct = config.sub_agent_percentage if config else Decimal('70.00')
        return Response(entry_to_dict(entry, sub_pct))

    @require_permission("provider_commission.delete")
    def delete(self, request, entry_id):
        try:
            entry = ProviderCommissionEntry.objects.get(id=entry_id)
            entry.delete()
            return Response({'message': 'Deleted'})
        except ProviderCommissionEntry.DoesNotExist:
            return Response({'message': 'Entry not found'}, status=404)


class ProviderSubAgentPercentageView(APIView):
    @require_permission("provider_commission.edit")
    def patch(self, request):
        provider_name = (request.data.get('providerName') or '').strip()
        if not provider_name:
            return Response({'message': 'Provider name is required'}, status=400)
        sub_pct_raw = request.data.get('subAgentPercentage')
        sub_pct_val = None
        if sub_pct_raw is not None and str(sub_pct_raw).strip() != '':
            try:
                sub_pct_val = Decimal(str(sub_pct_raw))
                if sub_pct_val < 0 or sub_pct_val > 100:
                    return Response({'message': 'Percentage must be between 0 and 100'}, status=400)
            except (InvalidOperation, ValueError):
                return Response({'message': 'Invalid percentage value'}, status=400)
        updated = ProviderCommissionEntry.objects.filter(provider_name=provider_name).update(
            sub_agent_percentage=sub_pct_val
        )
        return Response({'updated': updated, 'providerName': provider_name, 'subAgentPercentage': str(sub_pct_val) if sub_pct_val is not None else None})


class CopyFromCommissionRulesView(APIView):
    @require_permission("provider_commission.add")
    def get(self, request):
        rules = AgreementCommissionRule.objects.filter(is_active=True).order_by('agreement_id')
        result = []
        for r in rules:
            try:
                agr = Agreement.objects.get(id=r.agreement_id)
                prov = Provider.objects.get(id=agr.university_id)
            except (Agreement.DoesNotExist, Provider.DoesNotExist):
                continue

            already_copied = ProviderCommissionEntry.objects.filter(copied_from_rule_id=r.id).exists()
            result.append({
                'ruleId': r.id,
                'providerName': prov.name,
                'studyLevel': r.study_level or 'Any',
                'commissionMode': r.commission_mode,
                'percentageValue': str(r.percentage_value) if r.percentage_value else None,
                'flatAmount': str(r.flat_amount) if r.flat_amount else None,
                'currency': r.currency,
                'basis': r.basis,
                'agreementCode': agr.agreement_code,
                'agreementTitle': agr.title,
                'alreadyCopied': already_copied,
            })
        return Response(result)

    @require_permission("provider_commission.add")
    def post(self, request):
        rule_ids = request.data.get('ruleIds', [])
        if not rule_ids:
            return Response({'message': 'No rules selected'}, status=400)

        created = 0
        skipped = 0
        for rule_id in rule_ids:
            if ProviderCommissionEntry.objects.filter(copied_from_rule_id=rule_id).exists():
                skipped += 1
                continue
            try:
                rule = AgreementCommissionRule.objects.get(id=rule_id)
                agr = Agreement.objects.get(id=rule.agreement_id)
                prov = Provider.objects.get(id=agr.university_id)
                prov_name = prov.name
            except (AgreementCommissionRule.DoesNotExist, Agreement.DoesNotExist, Provider.DoesNotExist):
                skipped += 1
                continue

            level_map = {
                'Undergraduate': 'undergraduate',
                'Postgraduate': 'postgraduate',
                'VET': 'vet',
                'Foundation': 'foundation',
                'Diploma': 'diploma',
                'PhD': 'phd',
                'English Language': 'english',
            }
            degree = level_map.get(rule.study_level, 'any')

            if rule.commission_mode == 'percentage' and rule.percentage_value:
                cval = rule.percentage_value
                ctype = 'percentage'
            elif rule.flat_amount:
                cval = rule.flat_amount
                ctype = 'flat'
            else:
                cval = rule.percentage_value or 0
                ctype = 'percentage'

            basis_map = {
                'first_year': '1_year',
                'per_semester': 'per_semester',
                'per_year': 'per_year',
                'full_course': 'full_course',
                'per_subject': 'one_time',
                'per_trimester': 'per_trimester',
                'one_time': 'one_time',
            }
            basis = basis_map.get(rule.basis, 'full_course')

            ProviderCommissionEntry.objects.create(
                provider_name=prov_name,
                degree_level=degree,
                territory='',
                commission_value=cval,
                commission_type=ctype,
                currency=rule.currency or 'AUD',
                commission_basis=basis,
                notes=f'Copied from agreement {agr.agreement_code} rule #{rule.id}',
                is_active=True,
                copied_from_rule_id=rule.id,
                created_by=request.session.get('userId'),
            )
            created += 1

        return Response({'created': created, 'skipped': skipped})


VALID_DEGREE_LEVELS = {'any', 'undergraduate', 'postgraduate', 'vet', 'foundation', 'diploma', 'phd', 'english'}
VALID_COMMISSION_TYPES = {'percentage', 'flat'}
VALID_BASES = {'1_year', '2_semesters', 'full_course', 'per_semester', 'per_year', 'per_trimester', 'one_time'}

DEGREE_LABEL_TO_VALUE = {
    'any': 'any', 'undergraduate': 'undergraduate', 'ug': 'undergraduate',
    'postgraduate': 'postgraduate', 'pg': 'postgraduate',
    'vet': 'vet', 'foundation': 'foundation', 'diploma': 'diploma',
    'phd': 'phd', 'english': 'english', 'english language': 'english',
}

BASIS_LABEL_TO_VALUE = {
    '1 year': '1_year', '1_year': '1_year', 'first year': '1_year', 'first_year': '1_year',
    '2 semesters': '2_semesters', '2_semesters': '2_semesters',
    'full course': 'full_course', 'full_course': 'full_course',
    'per semester': 'per_semester', 'per_semester': 'per_semester',
    'per year': 'per_year', 'per_year': 'per_year',
    'per trimester': 'per_trimester', 'per_trimester': 'per_trimester',
    'one time': 'one_time', 'one_time': 'one_time',
    'per intake': 'one_time', 'per_intake': 'one_time',
    'per subject': 'one_time', 'per_subject': 'one_time',
}


class SampleDownloadView(APIView):
    @require_auth
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commission Distribution"

        headers = [
            "Provider Name*", "Degree Level", "Territory",
            "Commission Type*", "Value*", "Currency", "Commission Basis", "Notes"
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        sample_data = [
            ["University of Melbourne", "Undergraduate", "Australia, Global", "Percentage", "15", "", "Full Course", "Standard commission"],
            ["Deakin University", "Postgraduate", "South Asia, Nepal", "Percentage", "12.5", "", "Per Year", ""],
            ["University of Sydney", "Any", "Global", "Flat", "2500", "AUD", "Per Semester", "Fixed amount per semester"],
            ["Monash University", "Diploma", "Nepal, India, Bangladesh", "Percentage", "10", "", "1 Year", "First year only"],
            ["RMIT University", "VET", "Australia", "Flat", "1500", "AUD", "One Time", "One-time payment"],
        ]

        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 30

        ws_help = wb.create_sheet("Instructions")
        ws_help.column_dimensions['A'].width = 25
        ws_help.column_dimensions['B'].width = 60

        help_title_font = Font(bold=True, size=12, color="2563EB")
        help_header_font = Font(bold=True, size=10)
        help_content_font = Font(size=10)

        ws_help.cell(row=1, column=1, value="Bulk Upload Instructions").font = help_title_font
        ws_help.merge_cells('A1:B1')

        instructions = [
            ("Column", "Description"),
            ("Provider Name*", "Required. Name of the provider/university."),
            ("Degree Level", "Optional. Values: Any, Undergraduate (UG), Postgraduate (PG), VET, Foundation, Diploma, PhD, English Language. Default: Any"),
            ("Territory", "Optional. Comma-separated country names. e.g. 'Australia, Nepal' or 'Global' or 'South Asia'."),
            ("Commission Type*", "Required. Values: Percentage or Flat."),
            ("Value*", "Required. Numeric value. e.g. 15 for 15% or 2500 for flat amount."),
            ("Currency", "Optional. Values: AUD, USD, NPR, GBP. Default: AUD. Only used when Commission Type is Flat."),
            ("Commission Basis", "Optional. Values: Full Course, 1 Year, 2 Semesters, Per Semester, Per Year, Per Trimester, One Time. Default: Full Course."),
            ("Notes", "Optional. Any additional notes."),
        ]

        for row_idx, (col_a, col_b) in enumerate(instructions, 3):
            cell_a = ws_help.cell(row=row_idx, column=1, value=col_a)
            cell_b = ws_help.cell(row=row_idx, column=2, value=col_b)
            if row_idx == 3:
                cell_a.font = help_header_font
                cell_b.font = help_header_font
            else:
                cell_a.font = help_content_font
                cell_b.font = help_content_font

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="commission_distribution_sample.xlsx"'
        return response


class BulkUploadView(APIView):
    parser_classes = [MultiPartParser]

    @require_permission("provider_commission.add")
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No file uploaded'}, status=400)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'message': 'Please upload an Excel file (.xlsx)'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            if not rows:
                return Response({'message': 'No data rows found in the file'}, status=400)

            created = 0
            errors = []
            user_id = request.session.get('userId')

            for idx, row in enumerate(rows, 2):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                provider_name = str(row[0]).strip() if row[0] else ''
                if not provider_name:
                    errors.append(f"Row {idx}: Provider Name is required")
                    continue

                degree_raw = str(row[1]).strip().lower() if len(row) > 1 and row[1] else 'any'
                degree_level = DEGREE_LABEL_TO_VALUE.get(degree_raw, None)
                if not degree_level:
                    errors.append(f"Row {idx}: Invalid degree level '{row[1]}'")
                    continue

                territory = str(row[2]).strip() if len(row) > 2 and row[2] else ''

                comm_type_raw = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ''
                if comm_type_raw in ('percentage', '%', 'percent'):
                    commission_type = 'percentage'
                elif comm_type_raw in ('flat', 'flat amount', 'fixed'):
                    commission_type = 'flat'
                else:
                    errors.append(f"Row {idx}: Commission Type must be 'Percentage' or 'Flat', got '{row[3]}'")
                    continue

                try:
                    commission_value = Decimal(str(row[4]).strip()) if len(row) > 4 and row[4] else None
                    if commission_value is None:
                        errors.append(f"Row {idx}: Value is required")
                        continue
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {idx}: Invalid numeric value '{row[4]}'")
                    continue

                currency = str(row[5]).strip().upper() if len(row) > 5 and row[5] else 'AUD'
                if currency not in ('AUD', 'USD', 'NPR', 'GBP', ''):
                    currency = 'AUD'
                if not currency:
                    currency = 'AUD'

                basis_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'full course'
                commission_basis = BASIS_LABEL_TO_VALUE.get(basis_raw, 'full_course')

                notes = str(row[7]).strip() if len(row) > 7 and row[7] else ''

                ProviderCommissionEntry.objects.create(
                    provider_name=provider_name,
                    degree_level=degree_level,
                    territory=territory,
                    commission_value=commission_value,
                    commission_type=commission_type,
                    currency=currency,
                    commission_basis=commission_basis,
                    notes=notes,
                    is_active=True,
                    created_by=user_id,
                )
                created += 1

            wb.close()
            return Response({
                'created': created,
                'errors': errors,
                'totalRows': len([r for r in rows if r and any(cell is not None and str(cell).strip() != '' for cell in r)]),
            })

        except Exception as e:
            return Response({'message': f'Error processing file: {str(e)}'}, status=400)
