import os
import secrets
import hashlib
from datetime import datetime, date, timedelta
from decimal import Decimal
from django.conf import settings
from django.utils import timezone
from django.db.models import Q, Sum
from rest_framework.views import APIView
from rest_framework.response import Response

from core.permissions import require_auth, require_permission
from employees.models import Employee
from .models import (
    Organization, Department, FiscalYear, LeaveType,
    DepartmentLeaveAllocation, LeavePolicy, Holiday,
    LeaveBalance, LeaveRequest, AttendanceRecord,
    OnlineCheckInPermission, DeviceMapping, SalaryStructure,
    PayrollRun, Payslip, NotificationSetting,
    Bonus, TravelExpense, AdvancePayment, TaxSlab,
    CountryTaxLabel, BONUS_TYPE_CHOICES,
)


def get_presigned_url(s3_url, expiry=3600):
    if not s3_url:
        return None
    import boto3
    from urllib.parse import unquote
    try:
        bucket = settings.AWS_S3_BUCKET_NAME
        region = settings.AWS_S3_REGION_NAME
        key = None
        prefix_vhost = f'https://{bucket}.s3.{region}.amazonaws.com/'
        prefix_vhost_no_region = f'https://{bucket}.s3.amazonaws.com/'
        prefix_path = f'https://s3.{region}.amazonaws.com/{bucket}/'
        prefix_path_no_region = f'https://s3.amazonaws.com/{bucket}/'
        if s3_url.startswith(prefix_vhost):
            key = s3_url[len(prefix_vhost):]
        elif s3_url.startswith(prefix_vhost_no_region):
            key = s3_url[len(prefix_vhost_no_region):]
        elif s3_url.startswith(prefix_path):
            key = s3_url[len(prefix_path):]
        elif s3_url.startswith(prefix_path_no_region):
            key = s3_url[len(prefix_path_no_region):]
        elif f'{bucket}.s3.amazonaws.com/' in s3_url:
            key = s3_url.split(f'{bucket}.s3.amazonaws.com/')[-1]
        elif f'{bucket}/' in s3_url and 's3' in s3_url and 'amazonaws.com' in s3_url:
            key = s3_url.split(f'{bucket}/')[-1]
        if key is None:
            return s3_url
        key = unquote(key)
        s3 = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=region,
        )
        return s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket, 'Key': key},
            ExpiresIn=expiry,
        )
    except Exception:
        return s3_url


COUNTRY_TAX_LABELS = {
    'Nepal': 'PAN No.',
    'Australia': 'TFN',
    'Bangladesh': 'TIN',
    'India': 'PAN',
    'United Kingdom': 'NI Number',
    'United States': 'SSN',
    'Canada': 'SIN',
    'New Zealand': 'IRD Number',
    'Pakistan': 'NTN',
    'Sri Lanka': 'TIN',
    'Philippines': 'TIN',
    'Malaysia': 'TIN',
    'Singapore': 'NRIC/FIN',
    'Japan': 'My Number',
    'South Korea': 'RRN',
    'Germany': 'Tax ID',
    'France': 'NIF',
    'UAE': 'TRN',
    'Saudi Arabia': 'TIN',
    'Qatar': 'QID',
    'China': 'Tax ID',
}


def _get_tax_id_label(country):
    if not country:
        return 'Tax ID No.'
    try:
        ctl = CountryTaxLabel.objects.get(country=country, is_active=True)
        return ctl.tax_id_label
    except CountryTaxLabel.DoesNotExist:
        return COUNTRY_TAX_LABELS.get(country, 'Tax ID No.')


WEEKDAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


def serialize_org(org):
    return {
        'id': str(org.id),
        'name': org.name,
        'short_code': org.short_code,
        'address': org.address,
        'country': org.country,
        'phone': org.phone,
        'email': org.email,
        'registration_number': org.registration_number,
        'registration_label': org.registration_label or 'Registration No.',
        'pan_number': org.pan_number,
        'pan_label': org.pan_label or 'PAN No.',
        'logo_url': org.logo_url,
        'currency': org.currency or 'NPR',
        'week_off_day': getattr(org, 'week_off_day', 6),
        'week_off_day_name': WEEKDAY_NAMES[getattr(org, 'week_off_day', 6)] if 0 <= getattr(org, 'week_off_day', 6) <= 6 else 'Sunday',
        'status': org.status,
        'created_at': org.created_at.isoformat() if org.created_at else None,
    }


def serialize_dept(dept):
    return {
        'id': str(dept.id),
        'organization_id': str(dept.organization_id),
        'organization_name': dept.organization.name if dept.organization else None,
        'name': dept.name,
        'head_employee_id': str(dept.head_employee_id) if dept.head_employee_id else None,
        'working_days_per_week': dept.working_days_per_week,
        'work_start_time': dept.work_start_time.strftime('%H:%M') if dept.work_start_time else None,
        'work_end_time': dept.work_end_time.strftime('%H:%M') if dept.work_end_time else None,
        'late_threshold_minutes': dept.late_threshold_minutes,
        'early_leave_threshold_minutes': dept.early_leave_threshold_minutes,
        'status': dept.status,
    }


def _safe_date_str(val):
    if val is None:
        return None
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    return str(val)

def serialize_fiscal_year(fy):
    return {
        'id': str(fy.id),
        'organization_id': str(fy.organization_id),
        'name': fy.name,
        'start_date': _safe_date_str(fy.start_date),
        'end_date': _safe_date_str(fy.end_date),
        'is_current': fy.is_current,
    }


def serialize_leave_type(lt):
    dept_allocations = []
    for da in DepartmentLeaveAllocation.objects.filter(leave_type_id=lt.id).select_related('department'):
        dept_allocations.append({
            'id': str(da.id),
            'department_id': str(da.department_id),
            'department_name': da.department.name,
            'allocated_days': float(da.allocated_days),
        })
    return {
        'id': str(lt.id),
        'organization_id': str(lt.organization_id),
        'name': lt.name,
        'code': lt.code,
        'default_days': float(lt.default_days),
        'is_paid': lt.is_paid,
        'is_carry_forward': lt.is_carry_forward,
        'max_carry_forward_days': float(lt.max_carry_forward_days),
        'min_advance_days': lt.min_advance_days,
        'requires_document': lt.requires_document,
        'document_required_after_days': lt.document_required_after_days,
        'color': lt.color,
        'status': lt.status,
        'department_allocations': dept_allocations,
    }


def serialize_leave_policy(lp):
    return {
        'id': str(lp.id),
        'organization_id': str(lp.organization_id),
        'min_days_advance_notice': lp.min_days_advance_notice,
        'max_consecutive_days': lp.max_consecutive_days,
        'require_document_after_days': lp.require_document_after_days,
        'allow_half_day': lp.allow_half_day,
        'allow_negative_balance': lp.allow_negative_balance,
        'max_negative_days': float(lp.max_negative_days),
        'require_approval': lp.require_approval,
        'auto_approve_if_balance': lp.auto_approve_if_balance,
        'weekend_days': lp.weekend_days,
        'advance_notice_rules': lp.advance_notice_rules or [],
        'require_cover_person': lp.require_cover_person,
        'require_cover_after_days': lp.require_cover_after_days,
    }


def serialize_holiday(h):
    return {
        'id': str(h.id),
        'organization_id': str(h.organization_id),
        'name': h.name,
        'date': h.date.isoformat() if h.date else None,
        'is_optional': h.is_optional,
        'fiscal_year_id': str(h.fiscal_year_id) if h.fiscal_year_id else None,
    }


def serialize_leave_balance(lb):
    return {
        'id': str(lb.id),
        'employee_id': str(lb.employee_id),
        'leave_type_id': str(lb.leave_type_id),
        'leave_type_name': lb.leave_type.name if lb.leave_type else None,
        'leave_type_code': lb.leave_type.code if lb.leave_type else None,
        'min_advance_days': lb.leave_type.min_advance_days if lb.leave_type else 0,
        'fiscal_year_id': str(lb.fiscal_year_id),
        'fiscal_year_name': lb.fiscal_year.name if lb.fiscal_year else None,
        'allocated_days': float(lb.allocated_days),
        'used_days': float(lb.used_days),
        'carried_forward_days': float(lb.carried_forward_days),
        'remaining_days': float(lb.remaining_days),
    }


def _get_cover_person_name(cover_person_id):
    if not cover_person_id:
        return None
    try:
        emp = Employee.objects.get(id=cover_person_id)
        return emp.full_name
    except Employee.DoesNotExist:
        return None


def serialize_leave_request(lr):
    emp_name = None
    try:
        emp = Employee.objects.get(id=lr.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass

    approver_name = None
    if lr.approved_by:
        try:
            from accounts.models import User
            approver = User.objects.get(id=lr.approved_by)
            approver_name = approver.name
        except Exception:
            pass

    return {
        'id': str(lr.id),
        'employee_id': str(lr.employee_id),
        'employee_name': emp_name,
        'leave_type_id': str(lr.leave_type_id),
        'leave_type_name': lr.leave_type.name if lr.leave_type else None,
        'leave_type_color': lr.leave_type.color if lr.leave_type else None,
        'start_date': lr.start_date.isoformat() if lr.start_date else None,
        'end_date': lr.end_date.isoformat() if lr.end_date else None,
        'days_count': float(lr.days_count),
        'is_half_day': lr.is_half_day,
        'half_day_period': lr.half_day_period,
        'reason': lr.reason,
        'status': lr.status,
        'approved_by': str(lr.approved_by) if lr.approved_by else None,
        'approver_name': approver_name,
        'approved_at': lr.approved_at.isoformat() if lr.approved_at else None,
        'rejection_reason': lr.rejection_reason,
        'document_url': lr.document_url,
        'cover_person_id': str(lr.cover_person_id) if lr.cover_person_id else None,
        'cover_person_name': _get_cover_person_name(lr.cover_person_id),
        'created_at': lr.created_at.isoformat() if lr.created_at else None,
    }


def serialize_attendance(att):
    emp_name = None
    try:
        emp = Employee.objects.get(id=att.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass

    return {
        'id': str(att.id),
        'employee_id': str(att.employee_id),
        'employee_name': emp_name,
        'date': att.date.isoformat() if att.date else None,
        'check_in': att.check_in.isoformat() if att.check_in else None,
        'check_out': att.check_out.isoformat() if att.check_out else None,
        'check_in_method': att.check_in_method,
        'check_out_method': att.check_out_method,
        'check_in_location': att.check_in_location,
        'check_out_location': att.check_out_location,
        'check_in_photo_url': get_presigned_url(att.check_in_photo_url),
        'check_out_photo_url': get_presigned_url(att.check_out_photo_url),
        'is_late': att.is_late,
        'is_early_leave': att.is_early_leave,
        'late_minutes': att.late_minutes,
        'early_leave_minutes': att.early_leave_minutes,
        'status': att.status,
        'work_hours': float(att.work_hours),
        'overtime_hours': float(att.overtime_hours),
        'notes': att.notes,
    }


def serialize_salary_structure(ss):
    emp_name = None
    try:
        emp = Employee.objects.get(id=ss.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass

    return {
        'id': str(ss.id),
        'employee_id': str(ss.employee_id),
        'employee_name': emp_name,
        'basic_salary': float(ss.basic_salary),
        'allowances': ss.allowances,
        'deductions': ss.deductions,
        'cit_type': ss.cit_type,
        'cit_value': float(ss.cit_value),
        'ssf_applicable': ss.ssf_applicable,
        'ssf_employee_percentage': float(ss.ssf_employee_percentage),
        'ssf_employer_percentage': float(ss.ssf_employer_percentage),
        'tax_applicable': ss.tax_applicable,
        'effective_from': ss.effective_from.isoformat() if ss.effective_from and hasattr(ss.effective_from, 'isoformat') else str(ss.effective_from) if ss.effective_from else None,
        'effective_to': ss.effective_to.isoformat() if ss.effective_to and hasattr(ss.effective_to, 'isoformat') else str(ss.effective_to) if ss.effective_to else None,
        'status': ss.status,
    }


def serialize_payslip(ps):
    emp_name = None
    emp_pan = None
    emp_join_date = None
    emp_marital_status = None
    emp_gender = None
    emp_position = None
    try:
        emp = Employee.objects.get(id=ps.employee_id)
        emp_name = emp.full_name
        emp_pan = emp.pan_no
        emp_join_date = emp.join_date.isoformat() if emp.join_date else None
        emp_marital_status = emp.marital_status
        emp_gender = emp.gender
        emp_position = emp.position
    except Employee.DoesNotExist:
        pass

    paid_leave_days = 0
    unpaid_leave_days_count = 0
    from datetime import date as dt_date
    month_start = dt_date(ps.year, ps.month, 1)
    if ps.month == 12:
        month_end = dt_date(ps.year, 12, 31)
    else:
        month_end = dt_date(ps.year, ps.month + 1, 1) - timedelta(days=1)
    approved_leaves = LeaveRequest.objects.filter(
        employee_id=ps.employee_id,
        status='approved',
        start_date__lte=month_end,
        end_date__gte=month_start,
    ).select_related('leave_type').only(
        'id', 'employee_id', 'start_date', 'end_date',
        'days_count', 'status', 'is_half_day',
        'leave_type__is_paid', 'leave_type__id', 'leave_type__name',
    )
    for lr in approved_leaves:
        overlap_start = max(lr.start_date, month_start)
        overlap_end = min(lr.end_date, month_end)
        days = float(lr.days_count) if lr.is_half_day else (overlap_end - overlap_start).days + 1
        if lr.leave_type and lr.leave_type.is_paid:
            paid_leave_days += days
        else:
            unpaid_leave_days_count += days

    total_calendar_days = (month_end - month_start).days + 1

    total_income = float(ps.gross_salary)
    leave_amount = float(getattr(ps, 'unpaid_leave_deduction', 0) or 0)
    cit = float(ps.cit_deduction)
    bonus = float(getattr(ps, 'bonus_amount', 0) or 0)
    sst = float(ps.ssf_employee_deduction)
    tds = float(ps.tax_deduction)
    total_tax = cit + sst + tds
    taxable_amount_yearly = (total_income - cit + bonus) * 12
    advance = float(getattr(ps, 'advance_deduction', 0) or 0)
    payable_salary = float(ps.net_salary) - advance

    return {
        'id': str(ps.id),
        'payroll_run_id': str(ps.payroll_run_id),
        'employee_id': str(ps.employee_id),
        'employee_name': emp_name,
        'employee_pan': emp_pan,
        'employee_position': emp_position,
        'employee_join_date': emp_join_date,
        'employee_marital_status': emp_marital_status,
        'employee_gender': emp_gender,
        'month': ps.month,
        'year': ps.year,
        'basic_salary': float(ps.basic_salary),
        'allowances': ps.allowances,
        'gross_salary': float(ps.gross_salary),
        'cit_deduction': float(ps.cit_deduction),
        'ssf_employee_deduction': sst,
        'ssf_employer_contribution': float(ps.ssf_employer_contribution),
        'tax_deduction': tds,
        'bonus_amount': float(getattr(ps, 'bonus_amount', 0) or 0),
        'travel_reimbursement': float(getattr(ps, 'travel_reimbursement', 0) or 0),
        'advance_deduction': advance,
        'unpaid_leave_deduction': leave_amount,
        'other_deductions': ps.other_deductions,
        'total_deductions': float(ps.total_deductions),
        'net_salary': float(ps.net_salary),
        'working_days': ps.working_days,
        'present_days': ps.present_days,
        'absent_days': ps.absent_days,
        'leave_days': ps.leave_days,
        'paid_leave_days': paid_leave_days,
        'unpaid_leave_days': unpaid_leave_days_count,
        'total_days': total_calendar_days,
        'late_count': ps.late_count,
        'early_leave_count': ps.early_leave_count,
        'total_income': total_income,
        'leave_deduction_amount': leave_amount,
        'taxable_amount_yearly': taxable_amount_yearly,
        'sst': sst,
        'tds': tds,
        'total_tax': total_tax,
        'payable_salary': payable_salary,
        'status': ps.status,
        'view_token': ps.view_token,
    }


def serialize_bonus(b):
    emp_name = None
    try:
        emp = Employee.objects.get(id=b.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass
    return {
        'id': str(b.id),
        'employee_id': str(b.employee_id),
        'employee_name': emp_name,
        'bonus_type': b.bonus_type,
        'amount': float(b.amount),
        'reason': b.reason,
        'month': b.month,
        'year': b.year,
        'is_taxable': b.is_taxable,
        'status': b.status,
        'approved_by': str(b.approved_by) if b.approved_by else None,
        'approved_at': b.approved_at.isoformat() if b.approved_at else None,
        'created_at': b.created_at.isoformat() if b.created_at else None,
    }


def serialize_travel_expense(te):
    emp_name = None
    try:
        emp = Employee.objects.get(id=te.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass
    return {
        'id': str(te.id),
        'employee_id': str(te.employee_id),
        'employee_name': emp_name,
        'category': te.category,
        'description': te.description,
        'amount': float(te.amount),
        'expense_date': te.expense_date.isoformat() if te.expense_date else None,
        'receipt_url': get_presigned_url(te.receipt_url) if te.receipt_url else None,
        'month': te.month,
        'year': te.year,
        'include_in_salary': te.include_in_salary,
        'status': te.status,
        'approved_by': str(te.approved_by) if te.approved_by else None,
        'approved_at': te.approved_at.isoformat() if te.approved_at else None,
        'rejection_reason': te.rejection_reason,
        'created_at': te.created_at.isoformat() if te.created_at else None,
    }


def serialize_advance_payment(ap):
    emp_name = None
    try:
        emp = Employee.objects.get(id=ap.employee_id)
        emp_name = emp.full_name
    except Employee.DoesNotExist:
        pass
    return {
        'id': str(ap.id),
        'employee_id': str(ap.employee_id),
        'employee_name': emp_name,
        'amount': float(ap.amount),
        'reason': ap.reason,
        'request_date': ap.request_date.isoformat() if ap.request_date else None,
        'monthly_deduction': float(ap.monthly_deduction),
        'deduction_start_month': ap.deduction_start_month,
        'deduction_start_year': ap.deduction_start_year,
        'total_deducted': float(ap.total_deducted),
        'remaining_balance': float(ap.remaining_balance),
        'status': ap.status,
        'approved_by': str(ap.approved_by) if ap.approved_by else None,
        'approved_at': ap.approved_at.isoformat() if ap.approved_at else None,
        'created_at': ap.created_at.isoformat() if ap.created_at else None,
    }


def get_employee_for_user(user_id):
    if not user_id:
        return None
    try:
        return Employee.objects.get(user_id=user_id)
    except Employee.DoesNotExist:
        pass
    try:
        from accounts.models import User as AccountUser
        account = AccountUser.objects.get(id=user_id)
        if account.email:
            emp = Employee.objects.get(email__iexact=account.email)
            emp.user_id = user_id
            emp.save(update_fields=['user_id'])
            return emp
    except (AccountUser.DoesNotExist, Employee.DoesNotExist):
        pass
    return None


def _attendance_email_wrap(title, body_html):
    return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background-color:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f1f5f9;padding:40px 20px;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background-color:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
  <tr><td style="background:linear-gradient(135deg,#1e40af 0%,#3b82f6 100%);padding:28px 40px;text-align:center;">
    <h1 style="color:#ffffff;margin:0;font-size:20px;font-weight:700;">HRMS - Attendance Alert</h1>
    <p style="color:#bfdbfe;margin:4px 0 0;font-size:12px;">Study Info Centre</p>
  </td></tr>
  <tr><td style="padding:32px 40px;">
    <h2 style="color:#1e293b;margin:0 0 16px;font-size:18px;font-weight:600;">{title}</h2>
    {body_html}
  </td></tr>
  <tr><td style="background-color:#f8fafc;padding:20px 40px;border-top:1px solid #e2e8f0;text-align:center;">
    <p style="color:#94a3b8;font-size:11px;margin:0;">&copy; {timezone.now().year} Study Info Centre. All rights reserved.</p>
  </td></tr>
</table>
</td></tr></table>
</body></html>'''


def _get_attendance_cc_list(org):
    cc_list = []
    if org:
        ns = NotificationSetting.objects.filter(organization_id=org.id).first()
        if ns and ns.cc_emails:
            for email in ns.cc_emails.split(','):
                email = email.strip()
                if email and email not in cc_list:
                    cc_list.append(email)
        if org.email and org.email not in cc_list:
            cc_list.append(org.email)
    return cc_list


def _send_attendance_email(to_email, subject, html_body, cc_list=None):
    from django.core.mail import EmailMessage
    msg = EmailMessage(
        subject=subject,
        body=html_body,
        from_email=f'"{settings.FROM_NAME}" <{settings.DEFAULT_FROM_EMAIL}>',
        to=[to_email],
        cc=cc_list or [],
    )
    msg.content_subtype = 'html'
    msg.send(fail_silently=True)


def send_late_notification(employee, attendance_record, department):
    try:
        org = Organization.objects.get(id=employee.organization_id) if employee.organization_id else None
        if not org:
            return
        ns = NotificationSetting.objects.filter(organization=org).first()
        if not ns or not ns.late_arrival_notify_employee:
            return

        check_in_time = attendance_record.check_in.strftime('%I:%M %p') if attendance_record.check_in else 'N/A'
        att_date = attendance_record.date.strftime('%B %d, %Y')
        start_time = department.work_start_time.strftime('%I:%M %p') if department else 'N/A'
        late_mins = attendance_record.late_minutes or 0

        body_html = f'''
        <p style="color:#475569;font-size:15px;line-height:1.6;">Dear <strong>{employee.full_name}</strong>,</p>
        <div style="background-color:#fef2f2;border-left:4px solid #ef4444;padding:16px 20px;border-radius:0 8px 8px 0;margin:16px 0;">
          <p style="color:#991b1b;font-size:14px;margin:0;font-weight:600;">You were late today.</p>
        </div>
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f8fafc;border-radius:8px;border:1px solid #e2e8f0;margin:16px 0;">
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;width:160px;">Date</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{att_date}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Your Check-In</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;color:#ef4444;">{check_in_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Office Start Time</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{start_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;">Late By</td>
              <td style="padding:10px 16px;font-size:14px;font-weight:600;color:#ef4444;">{late_mins} minutes</td></tr>
        </table>
        <p style="color:#475569;font-size:14px;line-height:1.6;">Please be punctual and ensure you arrive on time. Repeated late arrivals may be subject to review.</p>
        <p style="color:#64748b;font-size:13px;margin-top:20px;">Regards,<br><strong>HR Department</strong><br>{org.name}</p>
        '''

        html = _attendance_email_wrap('Late Arrival Notice', body_html)
        cc_list = _get_attendance_cc_list(org)
        _send_attendance_email(employee.email, ns.late_email_subject, html, cc_list)
    except Exception as e:
        print(f'[Late Notification] Error: {e}')


def send_early_leave_notification(employee, attendance_record, department):
    try:
        org = Organization.objects.get(id=employee.organization_id) if employee.organization_id else None
        if not org:
            return
        ns = NotificationSetting.objects.filter(organization=org).first()
        if not ns or not ns.early_leave_notify_employee:
            return

        check_out_time = attendance_record.check_out.strftime('%I:%M %p') if attendance_record.check_out else 'N/A'
        att_date = attendance_record.date.strftime('%B %d, %Y')
        end_time = department.work_end_time.strftime('%I:%M %p') if department else 'N/A'
        early_mins = attendance_record.early_leave_minutes or 0

        body_html = f'''
        <p style="color:#475569;font-size:15px;line-height:1.6;">Dear <strong>{employee.full_name}</strong>,</p>
        <div style="background-color:#fff7ed;border-left:4px solid #f97316;padding:16px 20px;border-radius:0 8px 8px 0;margin:16px 0;">
          <p style="color:#9a3412;font-size:14px;margin:0;font-weight:600;">You checked out early today.</p>
        </div>
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f8fafc;border-radius:8px;border:1px solid #e2e8f0;margin:16px 0;">
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;width:160px;">Date</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{att_date}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Your Check-Out</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;color:#f97316;">{check_out_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Office End Time</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{end_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;">Left Early By</td>
              <td style="padding:10px 16px;font-size:14px;font-weight:600;color:#f97316;">{early_mins} minutes</td></tr>
        </table>
        <p style="color:#475569;font-size:14px;line-height:1.6;">Please ensure you complete your full working hours. If you had a valid reason, please inform your supervisor.</p>
        <p style="color:#64748b;font-size:13px;margin-top:20px;">Regards,<br><strong>HR Department</strong><br>{org.name}</p>
        '''

        html = _attendance_email_wrap('Early Departure Notice', body_html)
        cc_list = _get_attendance_cc_list(org)
        _send_attendance_email(employee.email, ns.early_leave_email_subject, html, cc_list)
    except Exception as e:
        print(f'[Early Leave Notification] Error: {e}')


def send_no_checkout_notification(employee, attendance_record, department):
    try:
        org = Organization.objects.get(id=employee.organization_id) if employee.organization_id else None
        if not org:
            return
        ns = NotificationSetting.objects.filter(organization=org).first()
        if not ns or not ns.no_checkout_notify_employee:
            return

        check_in_time = attendance_record.check_in.strftime('%I:%M %p') if attendance_record.check_in else 'N/A'
        att_date = attendance_record.date.strftime('%B %d, %Y')
        end_time = department.work_end_time.strftime('%I:%M %p') if department else 'N/A'

        body_html = f'''
        <p style="color:#475569;font-size:15px;line-height:1.6;">Dear <strong>{employee.full_name}</strong>,</p>
        <div style="background-color:#fefce8;border-left:4px solid #eab308;padding:16px 20px;border-radius:0 8px 8px 0;margin:16px 0;">
          <p style="color:#854d0e;font-size:14px;margin:0;font-weight:600;">You did not check out today.</p>
        </div>
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f8fafc;border-radius:8px;border:1px solid #e2e8f0;margin:16px 0;">
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;width:160px;">Date</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{att_date}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Your Check-In</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{check_in_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;border-bottom:1px solid #f1f5f9;">Office End Time</td>
              <td style="padding:10px 16px;font-size:14px;border-bottom:1px solid #f1f5f9;font-weight:500;">{end_time}</td></tr>
          <tr><td style="padding:10px 16px;color:#64748b;font-size:14px;">Check-Out</td>
              <td style="padding:10px 16px;font-size:14px;font-weight:600;color:#eab308;">Missing</td></tr>
        </table>
        <p style="color:#475569;font-size:14px;line-height:1.6;">Please remember to check out at the end of your working hours. Missing check-outs affect attendance records and may require manual correction.</p>
        <p style="color:#64748b;font-size:13px;margin-top:20px;">Regards,<br><strong>HR Department</strong><br>{org.name}</p>
        '''

        html = _attendance_email_wrap('Missing Check-Out Notice', body_html)
        cc_list = _get_attendance_cc_list(org)
        _send_attendance_email(employee.email, ns.no_checkout_email_subject, html, cc_list)
    except Exception as e:
        print(f'[No Checkout Notification] Error: {e}')


class OrganizationListView(APIView):
    @require_permission('hrms.organization.read')
    def get(self, request):
        orgs = Organization.objects.all()
        status = request.GET.get('status')
        if status:
            orgs = orgs.filter(status=status)
        return Response([serialize_org(o) for o in orgs])

    @require_permission('hrms.organization.add')
    def post(self, request):
        data = request.data
        org = Organization.objects.create(
            name=data.get('name', ''),
            short_code=data.get('short_code', ''),
            address=data.get('address'),
            country=data.get('country'),
            phone=data.get('phone'),
            email=data.get('email'),
            registration_number=data.get('registration_number'),
            registration_label=data.get('registration_label', 'Registration No.'),
            pan_number=data.get('pan_number'),
            pan_label=data.get('pan_label', 'PAN No.'),
            logo_url=data.get('logo_url'),
            currency=data.get('currency', 'NPR'),
            week_off_day=int(data.get('week_off_day', 6)),
        )
        return Response(serialize_org(org), status=201)


class OrganizationDetailView(APIView):
    @require_permission('hrms.organization.read')
    def get(self, request, org_id):
        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            return Response({'message': 'Organization not found'}, status=404)
        return Response(serialize_org(org))

    @require_permission('hrms.organization.update')
    def put(self, request, org_id):
        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            return Response({'message': 'Organization not found'}, status=404)
        data = request.data
        for field in ['name', 'short_code', 'address', 'country', 'phone', 'email',
                      'registration_number', 'registration_label', 'pan_number', 'pan_label',
                      'logo_url', 'currency', 'status', 'week_off_day']:
            if field in data:
                val = data[field]
                if field == 'week_off_day':
                    val = int(val)
                setattr(org, field, val)
        org.save()
        return Response(serialize_org(org))

    @require_permission('hrms.organization.delete')
    def delete(self, request, org_id):
        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            return Response({'message': 'Organization not found'}, status=404)
        org.delete()
        return Response({'message': 'Organization deleted'})


class DepartmentListView(APIView):
    @require_permission('hrms.department.read')
    def get(self, request):
        depts = Department.objects.select_related('organization').all()
        org_id = request.GET.get('organization_id')
        if org_id:
            depts = depts.filter(organization_id=org_id)
        status = request.GET.get('status')
        if status:
            depts = depts.filter(status=status)
        return Response([serialize_dept(d) for d in depts])

    @require_permission('hrms.department.add')
    def post(self, request):
        data = request.data
        try:
            dept = Department.objects.create(
                organization_id=data.get('organization_id'),
                name=data.get('name', ''),
                head_employee_id=data.get('head_employee_id'),
                working_days_per_week=data.get('working_days_per_week', 6),
                work_start_time=data.get('work_start_time', '10:00'),
                work_end_time=data.get('work_end_time', '18:00'),
                late_threshold_minutes=data.get('late_threshold_minutes', 15),
                early_leave_threshold_minutes=data.get('early_leave_threshold_minutes', 15),
            )
            dept = Department.objects.select_related('organization').get(id=dept.id)
            return Response(serialize_dept(dept), status=201)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'message': str(e)}, status=500)


class DepartmentDetailView(APIView):
    @require_permission('hrms.department.read')
    def get(self, request, dept_id):
        try:
            dept = Department.objects.select_related('organization').get(id=dept_id)
        except Department.DoesNotExist:
            return Response({'message': 'Department not found'}, status=404)
        return Response(serialize_dept(dept))

    @require_permission('hrms.department.update')
    def put(self, request, dept_id):
        try:
            dept = Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return Response({'message': 'Department not found'}, status=404)
        data = request.data
        for field in ['name', 'head_employee_id', 'working_days_per_week', 'work_start_time',
                      'work_end_time', 'late_threshold_minutes', 'early_leave_threshold_minutes', 'status']:
            if field in data:
                setattr(dept, field, data[field])
        if 'organization_id' in data:
            dept.organization_id = data['organization_id']
        dept.save()
        return Response(serialize_dept(dept))

    @require_permission('hrms.department.delete')
    def delete(self, request, dept_id):
        try:
            dept = Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return Response({'message': 'Department not found'}, status=404)
        dept.delete()
        return Response({'message': 'Department deleted'})


class FiscalYearListView(APIView):
    @require_permission('hrms.fiscal_year.read')
    def get(self, request):
        fys = FiscalYear.objects.all()
        org_id = request.GET.get('organization_id')
        if org_id:
            fys = fys.filter(organization_id=org_id)
        return Response([serialize_fiscal_year(fy) for fy in fys])

    @require_permission('hrms.fiscal_year.add')
    def post(self, request):
        data = request.data
        org_id = data.get('organization_id')
        name = (data.get('name') or '').strip()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        existing = FiscalYear.objects.filter(
            organization_id=org_id, name=name, start_date=start_date, end_date=end_date
        )
        if existing.exists():
            return Response({'message': f'Fiscal year "{name}" with these dates already exists for this organization.'}, status=400)
        if data.get('is_current'):
            FiscalYear.objects.filter(organization_id=org_id, is_current=True).update(is_current=False)
        fy = FiscalYear.objects.create(
            organization_id=org_id,
            name=name,
            start_date=start_date,
            end_date=end_date,
            is_current=data.get('is_current', False),
        )
        return Response(serialize_fiscal_year(fy), status=201)


class FiscalYearDetailView(APIView):
    @require_permission('hrms.fiscal_year.read')
    def get(self, request, fy_id):
        try:
            fy = FiscalYear.objects.get(id=fy_id)
        except FiscalYear.DoesNotExist:
            return Response({'message': 'Fiscal year not found'}, status=404)
        return Response(serialize_fiscal_year(fy))

    @require_permission('hrms.fiscal_year.update')
    def put(self, request, fy_id):
        try:
            fy = FiscalYear.objects.get(id=fy_id)
        except FiscalYear.DoesNotExist:
            return Response({'message': 'Fiscal year not found'}, status=404)
        data = request.data
        if data.get('is_current'):
            FiscalYear.objects.filter(organization_id=fy.organization_id, is_current=True).exclude(id=fy_id).update(is_current=False)
        for field in ['name', 'start_date', 'end_date', 'is_current']:
            if field in data:
                setattr(fy, field, data[field])
        fy.save()
        return Response(serialize_fiscal_year(fy))

    @require_permission('hrms.fiscal_year.delete')
    def delete(self, request, fy_id):
        try:
            fy = FiscalYear.objects.get(id=fy_id)
        except FiscalYear.DoesNotExist:
            return Response({'message': 'Fiscal year not found'}, status=404)
        fy.delete()
        return Response({'message': 'Fiscal year deleted'})


class LeaveTypeListView(APIView):
    @require_permission('hrms.leave_type.read')
    def get(self, request):
        lts = LeaveType.objects.all()
        org_id = request.GET.get('organization_id')
        if org_id:
            lts = lts.filter(organization_id=org_id)
        return Response([serialize_leave_type(lt) for lt in lts])

    @require_permission('hrms.leave_type.add')
    def post(self, request):
        data = request.data
        min_adv = data.get('min_advance_days', 0)
        if not isinstance(min_adv, int) or min_adv < 0:
            min_adv = max(0, int(min_adv or 0))
        lt = LeaveType.objects.create(
            organization_id=data.get('organization_id'),
            name=data.get('name', ''),
            code=data.get('code', ''),
            default_days=data.get('default_days', 0),
            is_paid=data.get('is_paid', True),
            is_carry_forward=data.get('is_carry_forward', False),
            max_carry_forward_days=data.get('max_carry_forward_days', 0),
            min_advance_days=min_adv,
            requires_document=data.get('requires_document', False),
            document_required_after_days=data.get('document_required_after_days', 0),
            color=data.get('color', '#3B82F6'),
        )
        return Response(serialize_leave_type(lt), status=201)


class LeaveTypeDetailView(APIView):
    @require_permission('hrms.leave_type.read')
    def get(self, request, lt_id):
        try:
            lt = LeaveType.objects.get(id=lt_id)
        except LeaveType.DoesNotExist:
            return Response({'message': 'Leave type not found'}, status=404)
        return Response(serialize_leave_type(lt))

    @require_permission('hrms.leave_type.update')
    def put(self, request, lt_id):
        try:
            lt = LeaveType.objects.get(id=lt_id)
        except LeaveType.DoesNotExist:
            return Response({'message': 'Leave type not found'}, status=404)
        data = request.data
        if 'min_advance_days' in data:
            min_adv = data['min_advance_days']
            if not isinstance(min_adv, int) or min_adv < 0:
                data['min_advance_days'] = max(0, int(min_adv or 0))
        for field in ['name', 'code', 'default_days', 'is_paid', 'is_carry_forward',
                      'max_carry_forward_days', 'min_advance_days', 'requires_document', 'document_required_after_days', 'color', 'status']:
            if field in data:
                setattr(lt, field, data[field])
        lt.save()
        return Response(serialize_leave_type(lt))

    @require_permission('hrms.leave_type.delete')
    def delete(self, request, lt_id):
        try:
            lt = LeaveType.objects.get(id=lt_id)
        except LeaveType.DoesNotExist:
            return Response({'message': 'Leave type not found'}, status=404)
        lt.delete()
        return Response({'message': 'Leave type deleted'})


class LeavePolicyListView(APIView):
    @require_permission('hrms.leave_policy.read')
    def get(self, request):
        policies = LeavePolicy.objects.all()
        org_id = request.GET.get('organization_id')
        if org_id:
            policies = policies.filter(organization_id=org_id)
        return Response([serialize_leave_policy(lp) for lp in policies])

    @require_permission('hrms.leave_policy.add')
    def post(self, request):
        data = request.data
        lp = LeavePolicy.objects.create(
            organization_id=data.get('organization_id'),
            min_days_advance_notice=data.get('min_days_advance_notice', 1),
            max_consecutive_days=data.get('max_consecutive_days', 14),
            require_document_after_days=data.get('require_document_after_days', 3),
            allow_half_day=data.get('allow_half_day', True),
            allow_negative_balance=data.get('allow_negative_balance', False),
            max_negative_days=data.get('max_negative_days', 0),
            require_approval=data.get('require_approval', True),
            auto_approve_if_balance=data.get('auto_approve_if_balance', False),
            weekend_days=data.get('weekend_days', [6]),
            advance_notice_rules=data.get('advance_notice_rules', []),
            require_cover_person=data.get('require_cover_person', False),
            require_cover_after_days=data.get('require_cover_after_days', 1),
        )
        return Response(serialize_leave_policy(lp), status=201)


class LeavePolicyDetailView(APIView):
    @require_permission('hrms.leave_policy.read')
    def get(self, request, lp_id):
        try:
            lp = LeavePolicy.objects.get(id=lp_id)
        except LeavePolicy.DoesNotExist:
            return Response({'message': 'Leave policy not found'}, status=404)
        return Response(serialize_leave_policy(lp))

    @require_permission('hrms.leave_policy.update')
    def put(self, request, lp_id):
        try:
            lp = LeavePolicy.objects.get(id=lp_id)
        except LeavePolicy.DoesNotExist:
            return Response({'message': 'Leave policy not found'}, status=404)
        data = request.data
        for field in ['min_days_advance_notice', 'max_consecutive_days', 'require_document_after_days',
                      'allow_half_day', 'allow_negative_balance', 'max_negative_days',
                      'require_approval', 'auto_approve_if_balance', 'weekend_days',
                      'advance_notice_rules', 'require_cover_person', 'require_cover_after_days']:
            if field in data:
                setattr(lp, field, data[field])
        lp.save()
        return Response(serialize_leave_policy(lp))

    @require_permission('hrms.leave_policy.delete')
    def delete(self, request, lp_id):
        try:
            lp = LeavePolicy.objects.get(id=lp_id)
        except LeavePolicy.DoesNotExist:
            return Response({'message': 'Leave policy not found'}, status=404)
        lp.delete()
        return Response({'message': 'Leave policy deleted'})


class HolidayListView(APIView):
    @require_permission('hrms.holiday.read')
    def get(self, request):
        holidays = Holiday.objects.all()
        org_id = request.GET.get('organization_id')
        if org_id:
            holidays = holidays.filter(organization_id=org_id)
        fy_id = request.GET.get('fiscal_year_id')
        if fy_id:
            holidays = holidays.filter(fiscal_year_id=fy_id)
        return Response([serialize_holiday(h) for h in holidays])

    @require_permission('hrms.holiday.add')
    def post(self, request):
        data = request.data
        h = Holiday.objects.create(
            organization_id=data.get('organization_id'),
            name=data.get('name', ''),
            date=data.get('date'),
            is_optional=data.get('is_optional', False),
            fiscal_year_id=data.get('fiscal_year_id'),
        )
        return Response(serialize_holiday(h), status=201)


class HolidayDetailView(APIView):
    @require_permission('hrms.holiday.read')
    def get(self, request, h_id):
        try:
            h = Holiday.objects.get(id=h_id)
        except Holiday.DoesNotExist:
            return Response({'message': 'Holiday not found'}, status=404)
        return Response(serialize_holiday(h))

    @require_permission('hrms.holiday.update')
    def put(self, request, h_id):
        try:
            h = Holiday.objects.get(id=h_id)
        except Holiday.DoesNotExist:
            return Response({'message': 'Holiday not found'}, status=404)
        data = request.data
        for field in ['name', 'date', 'is_optional', 'fiscal_year_id']:
            if field in data:
                setattr(h, field, data[field])
        h.save()
        return Response(serialize_holiday(h))

    @require_permission('hrms.holiday.delete')
    def delete(self, request, h_id):
        try:
            h = Holiday.objects.get(id=h_id)
        except Holiday.DoesNotExist:
            return Response({'message': 'Holiday not found'}, status=404)
        h.delete()
        return Response({'message': 'Holiday deleted'})


class LeaveBalanceListView(APIView):
    @require_permission('hrms.leave_balance.read')
    def get(self, request):
        balances = LeaveBalance.objects.select_related('leave_type', 'fiscal_year').all()
        emp_id = request.GET.get('employee_id')
        if emp_id:
            balances = balances.filter(employee_id=emp_id)
        fy_id = request.GET.get('fiscal_year_id')
        if fy_id:
            balances = balances.filter(fiscal_year_id=fy_id)
        return Response([serialize_leave_balance(lb) for lb in balances])


class LeaveBalanceAllocateView(APIView):
    @require_permission('hrms.leave_balance.add')
    def post(self, request):
        data = request.data
        employee_ids = data.get('employee_ids', [])
        fiscal_year_id = data.get('fiscal_year_id')
        organization_id = data.get('organization_id')

        if not fiscal_year_id or not organization_id:
            return Response({'message': 'fiscal_year_id and organization_id are required'}, status=400)

        leave_types = LeaveType.objects.filter(organization_id=organization_id, status='active')

        prev_fiscal_years = FiscalYear.objects.filter(
            organization_id=organization_id,
        ).exclude(id=fiscal_year_id).order_by('-end_date')
        prev_fy = prev_fiscal_years.first()

        created = 0
        for emp_id in employee_ids:
            emp = Employee.objects.filter(id=emp_id).first()
            for lt in leave_types:
                allocated = lt.default_days
                if emp and emp.department_id:
                    dept_alloc = DepartmentLeaveAllocation.objects.filter(
                        department_id=emp.department_id, leave_type=lt
                    ).first()
                    if dept_alloc:
                        allocated = dept_alloc.allocated_days

                carry_forward = Decimal('0')
                if lt.is_carry_forward and prev_fy:
                    prev_balance = LeaveBalance.objects.filter(
                        employee_id=emp_id, leave_type=lt, fiscal_year_id=prev_fy.id,
                    ).first()
                    if prev_balance:
                        remaining = prev_balance.allocated_days + prev_balance.carried_forward_days - prev_balance.used_days
                        if remaining > 0:
                            max_cf = lt.max_carry_forward_days if lt.max_carry_forward_days > 0 else remaining
                            carry_forward = min(remaining, max_cf)

                _, was_created = LeaveBalance.objects.get_or_create(
                    employee_id=emp_id,
                    leave_type=lt,
                    fiscal_year_id=fiscal_year_id,
                    defaults={
                        'allocated_days': allocated,
                        'carried_forward_days': carry_forward,
                    },
                )
                if was_created:
                    created += 1

        return Response({'message': f'{created} leave balances allocated'})


class LeaveRequestListView(APIView):
    @require_permission('hrms.leave_request.read')
    def get(self, request):
        reqs = LeaveRequest.objects.select_related('leave_type').all()
        emp_id = request.GET.get('employee_id')
        if emp_id:
            reqs = reqs.filter(employee_id=emp_id)
        status_filter = request.GET.get('status')
        if status_filter:
            reqs = reqs.filter(status=status_filter)
        org_id = request.GET.get('organization_id')
        if org_id:
            emp_ids = Employee.objects.filter(organization_id=org_id).values_list('id', flat=True)
            reqs = reqs.filter(employee_id__in=emp_ids)
        return Response([serialize_leave_request(lr) for lr in reqs[:200]])

    @require_permission('hrms.leave_request.add')
    def post(self, request):
        data = request.data
        employee_id = data.get('employee_id')
        leave_type_id = data.get('leave_type_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        is_half_day = data.get('is_half_day', False)

        if not all([employee_id, leave_type_id, start_date, end_date]):
            return Response({'message': 'employee_id, leave_type_id, start_date, end_date are required'}, status=400)

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        if end < start:
            return Response({'message': 'End date cannot be before start date'}, status=400)

        days_count = Decimal('0.5') if is_half_day else Decimal(str((end - start).days + 1))

        emp = Employee.objects.filter(id=employee_id).first()
        if emp and emp.organization_id:
            policy = LeavePolicy.objects.filter(organization_id=emp.organization_id).first()
            if policy:
                if policy.weekend_days:
                    business_days = Decimal('0')
                    current = start
                    while current <= end:
                        if current.weekday() not in policy.weekend_days:
                            business_days += 1
                        current += timedelta(days=1)
                    if not is_half_day:
                        days_count = business_days

                if days_count > policy.max_consecutive_days:
                    return Response({'message': f'Cannot request more than {policy.max_consecutive_days} consecutive days'}, status=400)

                if policy.min_days_advance_notice > 0:
                    days_ahead = (start - date.today()).days
                    if days_ahead < policy.min_days_advance_notice:
                        return Response({'message': f'Leave must be requested at least {policy.min_days_advance_notice} days in advance'}, status=400)

                if not policy.allow_half_day and is_half_day:
                    return Response({'message': 'Half-day leave is not allowed'}, status=400)

        lr = LeaveRequest.objects.create(
            employee_id=employee_id,
            leave_type_id=leave_type_id,
            start_date=start,
            end_date=end,
            days_count=days_count,
            is_half_day=is_half_day,
            half_day_period=data.get('half_day_period'),
            reason=data.get('reason'),
            document_url=data.get('document_url'),
        )
        return Response(serialize_leave_request(lr), status=201)


class LeaveRequestDetailView(APIView):
    @require_permission('hrms.leave_request.read')
    def get(self, request, lr_id):
        try:
            lr = LeaveRequest.objects.select_related('leave_type').get(id=lr_id)
        except LeaveRequest.DoesNotExist:
            return Response({'message': 'Leave request not found'}, status=404)
        return Response(serialize_leave_request(lr))

    @require_permission('hrms.leave_request.update')
    def put(self, request, lr_id):
        try:
            lr = LeaveRequest.objects.get(id=lr_id)
        except LeaveRequest.DoesNotExist:
            return Response({'message': 'Leave request not found'}, status=404)
        if lr.status != 'pending':
            return Response({'message': 'Only pending requests can be edited'}, status=400)
        data = request.data
        for field in ['start_date', 'end_date', 'reason', 'is_half_day', 'half_day_period', 'document_url']:
            if field in data:
                setattr(lr, field, data[field])
        lr.save()
        return Response(serialize_leave_request(lr))

    @require_permission('hrms.leave_request.delete')
    def delete(self, request, lr_id):
        try:
            lr = LeaveRequest.objects.get(id=lr_id)
        except LeaveRequest.DoesNotExist:
            return Response({'message': 'Leave request not found'}, status=404)
        if lr.status not in ('pending', 'rejected'):
            return Response({'message': 'Cannot delete an approved leave request'}, status=400)
        lr.delete()
        return Response({'message': 'Leave request deleted'})


class LeaveRequestApproveView(APIView):
    @require_permission('hrms.leave_request.approve')
    def post(self, request, lr_id):
        try:
            lr = LeaveRequest.objects.get(id=lr_id)
        except LeaveRequest.DoesNotExist:
            return Response({'message': 'Leave request not found'}, status=404)
        if lr.status != 'pending':
            return Response({'message': 'Only pending requests can be approved'}, status=400)

        user_id = request.session.get('userId')
        lr.status = 'approved'
        lr.approved_by = user_id
        lr.approved_at = timezone.now()
        lr.save()

        current_fy = FiscalYear.objects.filter(is_current=True).first()
        if current_fy:
            balance = LeaveBalance.objects.filter(
                employee_id=lr.employee_id,
                leave_type=lr.leave_type,
                fiscal_year=current_fy,
            ).first()
            if balance:
                balance.used_days += lr.days_count
                balance.save()

        start_d = lr.start_date
        end_d = lr.end_date
        current_d = start_d
        while current_d <= end_d:
            AttendanceRecord.objects.update_or_create(
                employee_id=lr.employee_id,
                date=current_d,
                defaults={'status': 'on_leave', 'notes': f'Leave: {lr.leave_type.name}'},
            )
            current_d += timedelta(days=1)

        return Response(serialize_leave_request(lr))


class LeaveRequestRejectView(APIView):
    @require_permission('hrms.leave_request.approve')
    def post(self, request, lr_id):
        try:
            lr = LeaveRequest.objects.get(id=lr_id)
        except LeaveRequest.DoesNotExist:
            return Response({'message': 'Leave request not found'}, status=404)
        if lr.status != 'pending':
            return Response({'message': 'Only pending requests can be rejected'}, status=400)

        lr.status = 'rejected'
        lr.rejection_reason = request.data.get('rejection_reason', '')
        lr.approved_by = request.session.get('userId')
        lr.approved_at = timezone.now()
        lr.save()
        return Response(serialize_leave_request(lr))


class AttendanceListView(APIView):
    @require_permission('hrms.attendance.read')
    def get(self, request):
        records = AttendanceRecord.objects.all()
        emp_id = request.GET.get('employee_id')
        if emp_id:
            records = records.filter(employee_id=emp_id)
        date_from = request.GET.get('date_from')
        if date_from:
            records = records.filter(date__gte=date_from)
        date_to = request.GET.get('date_to')
        if date_to:
            records = records.filter(date__lte=date_to)
        dept_id = request.GET.get('department_id')
        if dept_id:
            emp_ids = Employee.objects.filter(department_id=dept_id).values_list('id', flat=True)
            records = records.filter(employee_id__in=emp_ids)
        org_id = request.GET.get('organization_id')
        if org_id:
            emp_ids = Employee.objects.filter(organization_id=org_id).values_list('id', flat=True)
            records = records.filter(employee_id__in=emp_ids)
        the_date = request.GET.get('date')
        if the_date:
            records = records.filter(date=the_date)
        return Response([serialize_attendance(a) for a in records[:500]])

    @require_permission('hrms.attendance.add')
    def post(self, request):
        data = request.data
        att, created = AttendanceRecord.objects.update_or_create(
            employee_id=data.get('employee_id'),
            date=data.get('date'),
            defaults={
                'check_in': data.get('check_in'),
                'check_out': data.get('check_out'),
                'check_in_method': data.get('check_in_method', 'manual'),
                'check_out_method': data.get('check_out_method'),
                'status': data.get('status', 'present'),
                'notes': data.get('notes'),
            },
        )
        return Response(serialize_attendance(att), status=201 if created else 200)


class AttendanceDetailView(APIView):
    @require_permission('hrms.attendance.read')
    def get(self, request, att_id):
        try:
            att = AttendanceRecord.objects.get(id=att_id)
        except AttendanceRecord.DoesNotExist:
            return Response({'message': 'Attendance record not found'}, status=404)
        return Response(serialize_attendance(att))

    @require_permission('hrms.attendance.update')
    def put(self, request, att_id):
        try:
            att = AttendanceRecord.objects.get(id=att_id)
        except AttendanceRecord.DoesNotExist:
            return Response({'message': 'Attendance record not found'}, status=404)
        data = request.data
        for field in ['check_in', 'check_out', 'check_in_method', 'check_out_method',
                      'status', 'notes', 'is_late', 'is_early_leave', 'late_minutes', 'early_leave_minutes']:
            if field in data:
                setattr(att, field, data[field])
        att.save()
        return Response(serialize_attendance(att))


class OnlineCheckInView(APIView):
    @require_auth
    def post(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked to your account'}, status=404)

        perm = OnlineCheckInPermission.objects.filter(employee_id=employee.id, is_allowed=True).first()
        if not perm:
            return Response({'message': 'You do not have permission for online check-in'}, status=403)

        data = request.data
        today = date.today()
        now = timezone.now()

        existing = AttendanceRecord.objects.filter(employee_id=employee.id, date=today).first()
        if existing and existing.check_in:
            return Response({'message': 'You have already checked in today'}, status=400)

        dept = None
        is_late = False
        late_mins = 0
        if employee.department_id:
            dept = Department.objects.filter(id=employee.department_id).first()
            if dept:
                work_start = datetime.combine(today, dept.work_start_time)
                work_start = timezone.make_aware(work_start)
                diff = (now - work_start).total_seconds() / 60
                if diff > dept.late_threshold_minutes:
                    is_late = True
                    late_mins = int(diff)

        att, _ = AttendanceRecord.objects.update_or_create(
            employee_id=employee.id,
            date=today,
            defaults={
                'check_in': now,
                'check_in_method': 'online',
                'check_in_location': data.get('location'),
                'check_in_photo_url': data.get('photo_url'),
                'is_late': is_late,
                'late_minutes': late_mins,
                'status': 'present',
            },
        )

        if is_late:
            send_late_notification(employee, att, dept)

        return Response(serialize_attendance(att))


class OnlineCheckOutView(APIView):
    @require_auth
    def post(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked to your account'}, status=404)

        perm = OnlineCheckInPermission.objects.filter(employee_id=employee.id, is_allowed=True).first()
        if not perm:
            return Response({'message': 'You do not have permission for online check-out'}, status=403)

        data = request.data
        today = date.today()
        now = timezone.now()

        att = AttendanceRecord.objects.filter(employee_id=employee.id, date=today).first()
        if not att or not att.check_in:
            return Response({'message': 'You have not checked in today'}, status=400)
        if att.check_out:
            return Response({'message': 'You have already checked out today'}, status=400)

        dept = None
        is_early = False
        early_mins = 0
        if employee.department_id:
            dept = Department.objects.filter(id=employee.department_id).first()
            if dept:
                work_end = datetime.combine(today, dept.work_end_time)
                work_end = timezone.make_aware(work_end)
                diff = (work_end - now).total_seconds() / 60
                if diff > dept.early_leave_threshold_minutes:
                    is_early = True
                    early_mins = int(diff)

        if att.check_in:
            work_seconds = (now - att.check_in).total_seconds()
            att.work_hours = Decimal(str(round(work_seconds / 3600, 2)))

        att.check_out = now
        att.check_out_method = 'online'
        att.check_out_location = data.get('location')
        att.check_out_photo_url = data.get('photo_url')
        att.is_early_leave = is_early
        att.early_leave_minutes = early_mins
        att.save()

        if is_early:
            send_early_leave_notification(employee, att, dept)

        return Response(serialize_attendance(att))


class LeaveDocumentUploadView(APIView):
    @require_auth
    def post(self, request):
        import boto3
        import uuid as uuid_mod
        doc = request.FILES.get('document')
        if not doc:
            return Response({'message': 'No document provided'}, status=400)
        ext = doc.name.rsplit('.', 1)[-1] if '.' in doc.name else 'pdf'
        key = f'leave-documents/{uuid_mod.uuid4()}.{ext}'
        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            s3.upload_fileobj(
                doc, settings.AWS_S3_BUCKET_NAME, key,
                ExtraArgs={'ContentType': doc.content_type or 'application/pdf'},
            )
            url = f'https://{settings.AWS_S3_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{key}'
            return Response({'url': url, 'filename': doc.name})
        except Exception as e:
            return Response({'message': f'Upload failed: {str(e)}'}, status=500)


class AttendancePhotoUploadView(APIView):
    @require_auth
    def post(self, request):
        import boto3
        import uuid as uuid_mod
        photo = request.FILES.get('photo')
        if not photo:
            return Response({'message': 'No photo provided'}, status=400)
        ext = photo.name.rsplit('.', 1)[-1] if '.' in photo.name else 'jpg'
        key = f'attendance-photos/{uuid_mod.uuid4()}.{ext}'
        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            s3.upload_fileobj(
                photo, settings.AWS_S3_BUCKET_NAME, key,
                ExtraArgs={'ContentType': photo.content_type or 'image/jpeg'},
            )
            url = f'https://{settings.AWS_S3_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{key}'
            return Response({'url': url})
        except Exception as e:
            return Response({'message': f'Upload failed: {str(e)}'}, status=500)


class AttendancePhotoProxyView(APIView):
    @require_auth
    def get(self, request, att_id):
        import boto3
        from urllib.parse import unquote
        from django.http import HttpResponse

        photo_type = request.GET.get('type', 'check_in')
        try:
            att = AttendanceRecord.objects.get(id=att_id)
        except AttendanceRecord.DoesNotExist:
            return Response({'message': 'Attendance record not found'}, status=404)

        s3_url = att.check_in_photo_url if photo_type == 'check_in' else att.check_out_photo_url
        if not s3_url:
            return Response({'message': 'No photo available'}, status=404)

        bucket = settings.AWS_S3_BUCKET_NAME
        region = settings.AWS_S3_REGION_NAME
        key = None
        prefix_vhost = f'https://{bucket}.s3.{region}.amazonaws.com/'
        prefix_vhost_no_region = f'https://{bucket}.s3.amazonaws.com/'
        prefix_path = f'https://s3.{region}.amazonaws.com/{bucket}/'
        if s3_url.startswith(prefix_vhost):
            key = s3_url[len(prefix_vhost):]
        elif s3_url.startswith(prefix_vhost_no_region):
            key = s3_url[len(prefix_vhost_no_region):]
        elif s3_url.startswith(prefix_path):
            key = s3_url[len(prefix_path):]
        elif f'{bucket}.s3.amazonaws.com/' in s3_url:
            key = s3_url.split(f'{bucket}.s3.amazonaws.com/')[-1]
        elif f'{bucket}/' in s3_url and 's3' in s3_url:
            key = s3_url.split(f'{bucket}/')[-1]
        if not key:
            return Response({'message': 'Cannot resolve S3 key'}, status=400)
        key = unquote(key)

        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=region,
            )
            obj = s3.get_object(Bucket=bucket, Key=key)
            content_type = obj.get('ContentType', 'image/jpeg')
            body = obj['Body'].read()
            response = HttpResponse(body, content_type=content_type)
            response['Cache-Control'] = 'private, max-age=3600'
            return response
        except Exception as e:
            return Response({'message': f'Failed to fetch photo: {str(e)}'}, status=500)


class ExpenseReceiptUploadView(APIView):
    @require_auth
    def post(self, request):
        import boto3
        import uuid as uuid_mod
        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No file provided'}, status=400)
        ext = file.name.rsplit('.', 1)[-1] if '.' in file.name else 'pdf'
        key = f'expense-receipts/{uuid_mod.uuid4()}.{ext}'
        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            s3.upload_fileobj(
                file, settings.AWS_S3_BUCKET_NAME, key,
                ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'},
            )
            url = f'https://{settings.AWS_S3_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{key}'
            return Response({'url': url, 'filename': file.name})
        except Exception as e:
            return Response({'message': f'Upload failed: {str(e)}'}, status=500)


class DeviceSyncView(APIView):
    def post(self, request):
        api_key = request.headers.get('X-Sync-Key', '')
        sync_secret = os.getenv('DEVICE_SYNC_KEY', '')
        if api_key and sync_secret and api_key == sync_secret:
            pass
        else:
            user_id = request.session.get('userId')
            if not user_id:
                return Response({'message': 'Authentication required'}, status=401)
            user_perms = request.session.get('userPermissions', [])
            if 'hrms.attendance.add' not in user_perms:
                return Response({'message': 'Insufficient permissions'}, status=403)

        records = request.data.get('records', [])
        synced = 0
        errors = []
        for rec in records:
            device_user_id = rec.get('user_id')
            punch_time_str = rec.get('punch_time')
            punch_type = rec.get('punch_type', 'in')

            mapping = DeviceMapping.objects.filter(device_user_id=device_user_id).first()
            if not mapping:
                emp_match = Employee.objects.filter(
                    employee_id_number=str(device_user_id), status='active'
                ).first()
                if emp_match:
                    if not DeviceMapping.objects.filter(employee_id=emp_match.id).exists():
                        mapping = DeviceMapping.objects.create(
                            employee_id=emp_match.id,
                            device_user_id=str(device_user_id),
                            device_name='ZKT K40 (auto)',
                        )
                    else:
                        errors.append(f'No mapping for device user {device_user_id} (employee {emp_match.employee_id_number} already mapped to different device)')
                        continue
                else:
                    errors.append(f'No mapping for device user {device_user_id}')
                    continue

            try:
                import zoneinfo
                nepal_tz = zoneinfo.ZoneInfo('Asia/Kathmandu')
                punch_time = datetime.fromisoformat(punch_time_str)
                if timezone.is_naive(punch_time):
                    punch_time = punch_time.replace(tzinfo=nepal_tz)
            except (ValueError, TypeError):
                errors.append(f'Invalid punch_time: {punch_time_str}')
                continue

            punch_date = punch_time.astimezone(nepal_tz).date()
            att, created = AttendanceRecord.objects.get_or_create(
                employee_id=mapping.employee_id,
                date=punch_date,
                defaults={'device_user_id': device_user_id}
            )

            employee = Employee.objects.filter(id=mapping.employee_id).first()
            dept = Department.objects.filter(id=employee.department_id).first() if employee and employee.department_id else None

            is_checkin = False
            is_checkout = False

            if punch_type == 'out':
                is_checkout = True
            elif not att.check_in:
                is_checkin = True
            elif att.check_in and punch_time > att.check_in:
                is_checkout = True

            if is_checkin:
                att.check_in = punch_time
                att.check_in_method = 'device'
                att.status = 'present'
                if dept:
                    work_start = datetime.combine(punch_date, dept.work_start_time)
                    work_start = work_start.replace(tzinfo=nepal_tz)
                    diff = (punch_time - work_start).total_seconds() / 60
                    if diff > dept.late_threshold_minutes:
                        att.is_late = True
                        att.late_minutes = int(diff)
                att.save()
                if att.is_late and employee:
                    send_late_notification(employee, att, dept)
            elif is_checkout:
                if not att.check_out or punch_time > att.check_out:
                    att.check_out = punch_time
                    att.check_out_method = 'device'
                if att.check_in:
                    att.work_hours = Decimal(str(round((att.check_out - att.check_in).total_seconds() / 3600, 2)))
                if dept:
                    work_end = datetime.combine(punch_date, dept.work_end_time)
                    work_end = work_end.replace(tzinfo=nepal_tz)
                    diff = (work_end - punch_time).total_seconds() / 60
                    if diff > dept.early_leave_threshold_minutes:
                        att.is_early_leave = True
                        att.early_leave_minutes = int(diff)
                att.save()
                if att.is_early_leave and employee:
                    send_early_leave_notification(employee, att, dept)

            synced += 1

        return Response({'synced': synced, 'errors': errors})


class AttendanceDashboardView(APIView):
    @require_permission('hrms.attendance.read')
    def get(self, request):
        today = request.GET.get('date', date.today().isoformat())
        org_id = request.GET.get('organization_id')
        dept_id = request.GET.get('department_id')

        employees = Employee.objects.filter(status='active')
        if org_id:
            employees = employees.filter(organization_id=org_id)
        if dept_id:
            employees = employees.filter(department_id=dept_id)

        emp_ids = list(employees.values_list('id', flat=True))
        attendance = AttendanceRecord.objects.filter(employee_id__in=emp_ids, date=today)
        att_map = {str(a.employee_id): a for a in attendance}

        present = []
        absent = []
        on_leave = []
        late = []

        for emp in employees:
            emp_data = {
                'id': str(emp.id),
                'full_name': emp.full_name,
                'department': emp.department,
                'position': emp.position,
            }
            att = att_map.get(str(emp.id))
            if att:
                if att.status == 'on_leave':
                    on_leave.append(emp_data)
                else:
                    present.append({
                        **emp_data,
                        'check_in': att.check_in.isoformat() if att.check_in else None,
                        'check_out': att.check_out.isoformat() if att.check_out else None,
                        'is_late': att.is_late,
                        'late_minutes': att.late_minutes,
                    })
                    if att.is_late:
                        late.append(emp_data)
            else:
                absent.append(emp_data)

        leave_reqs = LeaveRequest.objects.filter(
            employee_id__in=emp_ids,
            status='approved',
            start_date__lte=today,
            end_date__gte=today,
        ).only('id', 'employee_id', 'start_date', 'end_date', 'status')
        for lr in leave_reqs:
            emp = employees.filter(id=lr.employee_id).first()
            if emp and str(emp.id) not in [o['id'] for o in on_leave]:
                on_leave.append({
                    'id': str(emp.id),
                    'full_name': emp.full_name,
                    'department': emp.department,
                    'position': emp.position,
                    'leave_type': lr.leave_type.name if lr.leave_type else None,
                })

        return Response({
            'date': today,
            'total_employees': len(emp_ids),
            'present_count': len(present),
            'absent_count': len(absent),
            'on_leave_count': len(on_leave),
            'late_count': len(late),
            'present': present,
            'absent': absent,
            'on_leave': on_leave,
            'late': late,
        })


class HRMSDashboardView(APIView):
    @require_auth
    def get(self, request):
        today = date.today()
        employees = Employee.objects.filter(status='active')
        total = employees.count()

        gender_counts = {}
        marital_counts = {}
        upcoming_birthdays = []
        for emp in employees:
            g = (emp.gender or 'unknown').lower()
            gender_counts[g] = gender_counts.get(g, 0) + 1
            ms = (emp.marital_status or 'unknown').lower()
            marital_counts[ms] = marital_counts.get(ms, 0) + 1
            if emp.date_of_birth:
                this_year_bday = emp.date_of_birth.replace(year=today.year)
                if this_year_bday < today:
                    this_year_bday = emp.date_of_birth.replace(year=today.year + 1)
                days_until = (this_year_bday - today).days
                if 0 <= days_until <= 30:
                    upcoming_birthdays.append({
                        'id': str(emp.id),
                        'name': emp.full_name,
                        'email': emp.email,
                        'date_of_birth': emp.date_of_birth.isoformat(),
                        'days_until': days_until,
                        'birthday_date': this_year_bday.isoformat(),
                    })
        upcoming_birthdays.sort(key=lambda x: x['days_until'])

        emp_ids = list(employees.values_list('id', flat=True))

        att_records = AttendanceRecord.objects.filter(employee_id__in=emp_ids, date=today)
        att_map = {str(a.employee_id): a for a in att_records}
        on_leave_today = []
        on_leave_ids = set()

        approved_leaves_today = LeaveRequest.objects.filter(
            employee_id__in=emp_ids,
            status='approved',
            start_date__lte=today,
            end_date__gte=today,
        ).select_related('leave_type')
        for lr in approved_leaves_today:
            if str(lr.employee_id) not in on_leave_ids:
                emp_obj = employees.filter(id=lr.employee_id).first()
                if emp_obj:
                    on_leave_ids.add(str(lr.employee_id))
                    on_leave_today.append({
                        'id': str(emp_obj.id),
                        'name': emp_obj.full_name,
                        'department': emp_obj.department or '',
                        'leave_type': lr.leave_type.name if lr.leave_type else 'Leave',
                        'start_date': lr.start_date.isoformat(),
                        'end_date': lr.end_date.isoformat(),
                        'days': float(lr.days_count),
                    })

        for emp in employees:
            att = att_map.get(str(emp.id))
            if att and att.status == 'on_leave' and str(emp.id) not in on_leave_ids:
                on_leave_ids.add(str(emp.id))
                on_leave_today.append({
                    'id': str(emp.id),
                    'name': emp.full_name,
                    'department': emp.department or '',
                    'leave_type': 'On Leave',
                    'start_date': today.isoformat(),
                    'end_date': today.isoformat(),
                    'days': 1,
                })

        pending_leaves = LeaveRequest.objects.filter(
            employee_id__in=emp_ids,
            status='pending',
        ).select_related('leave_type').order_by('-created_at')[:10]
        pending_leave_list = []
        for lr in pending_leaves:
            emp_obj = employees.filter(id=lr.employee_id).first()
            pending_leave_list.append({
                'id': str(lr.id),
                'employee_id': str(lr.employee_id),
                'employee_name': emp_obj.full_name if emp_obj else 'Unknown',
                'employee_email': emp_obj.email if emp_obj else '',
                'leave_type': lr.leave_type.name if lr.leave_type else '',
                'start_date': lr.start_date.isoformat(),
                'end_date': lr.end_date.isoformat(),
                'days': float(lr.days_count),
                'reason': lr.reason or '',
                'status': lr.status,
                'created_at': lr.created_at.isoformat() if lr.created_at else None,
            })

        org_ids = set(employees.values_list('organization_id', flat=True))
        org_ids.discard(None)
        upcoming_holidays = []
        if org_ids:
            holidays = Holiday.objects.filter(
                organization_id__in=org_ids,
                date__gte=today,
                date__lte=today + timedelta(days=60),
            ).order_by('date')[:10]
            for h in holidays:
                org_name = ''
                try:
                    org_name = Organization.objects.get(id=h.organization_id).name
                except Organization.DoesNotExist:
                    pass
                upcoming_holidays.append({
                    'id': str(h.id),
                    'name': h.name,
                    'date': h.date.isoformat(),
                    'is_optional': h.is_optional,
                    'organization': org_name,
                    'days_until': (h.date - today).days,
                })

        present_count = sum(1 for a in att_records if a.status in ('present', 'half_day'))
        absent_count = total - present_count - len(on_leave_ids)
        late_count = sum(1 for a in att_records if a.is_late)

        contracts_ending = []
        for emp in employees:
            if emp.contract_end_date:
                days_left = (emp.contract_end_date - today).days
                if 0 <= days_left <= 60:
                    contracts_ending.append({
                        'id': str(emp.id),
                        'name': emp.full_name,
                        'email': emp.email,
                        'department': emp.department or '',
                        'contract_end_date': emp.contract_end_date.isoformat(),
                        'days_remaining': days_left,
                    })
        contracts_ending.sort(key=lambda x: x['days_remaining'])

        return Response({
            'total_employees': total,
            'present_today': present_count,
            'absent_today': max(absent_count, 0),
            'on_leave_today': len(on_leave_ids),
            'late_today': late_count,
            'gender_distribution': gender_counts,
            'marital_status_distribution': marital_counts,
            'upcoming_birthdays': upcoming_birthdays[:10],
            'on_leave_list': on_leave_today,
            'pending_leave_requests': pending_leave_list,
            'upcoming_holidays': upcoming_holidays,
            'contracts_ending_soon': contracts_ending[:10],
        })


class AttendanceGridView(APIView):
    @require_permission('hrms.attendance.read')
    def get(self, request):
        view_mode = request.GET.get('mode', 'monthly')
        target_date = request.GET.get('date', date.today().isoformat())
        org_id = request.GET.get('organization_id')
        dept_id = request.GET.get('department_id')

        try:
            target = date.fromisoformat(target_date)
        except ValueError:
            target = date.today()

        if view_mode == 'daily':
            date_from = target
            date_to = target
        elif view_mode == 'weekly':
            date_from = target - timedelta(days=target.weekday())
            date_to = date_from + timedelta(days=6)
        else:
            date_from = date(target.year, target.month, 1)
            if target.month == 12:
                date_to = date(target.year + 1, 1, 1) - timedelta(days=1)
            else:
                date_to = date(target.year, target.month + 1, 1) - timedelta(days=1)

        employees = Employee.objects.filter(status='active').order_by('full_name')
        if org_id:
            employees = employees.filter(organization_id=org_id)
        if dept_id:
            employees = employees.filter(department_id=dept_id)

        days = []
        d = date_from
        while d <= date_to:
            days.append(d.isoformat())
            d += timedelta(days=1)

        emp_ids = [emp.id for emp in employees]
        records = AttendanceRecord.objects.filter(
            employee_id__in=emp_ids,
            date__gte=date_from,
            date__lte=date_to,
        )
        att_map = {}
        for r in records:
            key = f"{r.employee_id}_{r.date.isoformat()}"
            att_map[key] = {
                'id': str(r.id),
                'status': r.status,
                'check_in': r.check_in.isoformat() if r.check_in else None,
                'check_out': r.check_out.isoformat() if r.check_out else None,
                'is_late': r.is_late,
                'late_minutes': r.late_minutes,
                'check_in_method': r.check_in_method,
                'check_out_method': r.check_out_method,
                'check_in_photo_url': get_presigned_url(r.check_in_photo_url) if r.check_in_photo_url else None,
                'check_out_photo_url': get_presigned_url(r.check_out_photo_url) if r.check_out_photo_url else None,
                'check_in_location': r.check_in_location or None,
                'notes': r.notes or '',
            }

        rows = []
        for emp in employees:
            dept_name = ''
            if emp.department_id:
                try:
                    dept_name = Department.objects.get(id=emp.department_id).name
                except Department.DoesNotExist:
                    pass
            attendance_data = {}
            summary = {'present': 0, 'absent': 0, 'on_leave': 0, 'late': 0}
            for day in days:
                key = f"{emp.id}_{day}"
                entry = att_map.get(key)
                if entry:
                    attendance_data[day] = entry
                    if entry['status'] == 'on_leave':
                        summary['on_leave'] += 1
                    elif entry['status'] in ('present', 'half_day'):
                        summary['present'] += 1
                        if entry['is_late']:
                            summary['late'] += 1
                    elif entry['status'] == 'absent':
                        summary['absent'] += 1
                else:
                    attendance_data[day] = None
                    if date.fromisoformat(day) <= date.today():
                        summary['absent'] += 1

            rows.append({
                'employee_id': str(emp.id),
                'full_name': emp.full_name,
                'department': dept_name,
                'position': emp.position or '',
                'attendance': attendance_data,
                'summary': summary,
            })

        return Response({
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'days': days,
            'employees': rows,
            'mode': view_mode,
        })


class AttendanceSummaryView(APIView):
    @require_permission('hrms.attendance.read')
    def get(self, request):
        month = request.GET.get('month')
        year = request.GET.get('year')
        org_id = request.GET.get('organization_id')
        dept_id = request.GET.get('department_id')

        if not month or not year:
            today = date.today()
            month = month or str(today.month)
            year = year or str(today.year)

        month_int = int(month)
        year_int = int(year)

        date_from = date(year_int, month_int, 1)
        if month_int == 12:
            date_to = date(year_int + 1, 1, 1) - timedelta(days=1)
        else:
            date_to = date(year_int, month_int + 1, 1) - timedelta(days=1)

        today_date = date.today()
        effective_to = min(date_to, today_date)

        total_days_in_month = (date_to - date_from).days + 1

        employees = Employee.objects.filter(status='active').order_by('full_name')
        if org_id:
            employees = employees.filter(organization_id=org_id)
        if dept_id:
            employees = employees.filter(department_id=dept_id)

        emp_ids = [emp.id for emp in employees]

        records = AttendanceRecord.objects.filter(
            employee_id__in=emp_ids,
            date__gte=date_from,
            date__lte=date_to,
        )
        att_map = {}
        for r in records:
            att_map.setdefault(str(r.employee_id), []).append(r)

        org_ids = set()
        dept_ids_set = set()
        for emp in employees:
            if emp.organization_id:
                org_ids.add(emp.organization_id)
            if emp.department_id:
                dept_ids_set.add(emp.department_id)

        holidays_by_org = {}
        if org_ids:
            all_holidays = Holiday.objects.filter(
                organization_id__in=org_ids,
                date__gte=date_from,
                date__lte=date_to,
                is_optional=False,
            )
            for h in all_holidays:
                holidays_by_org.setdefault(str(h.organization_id), set()).add(h.date)

        dept_cache = {}
        if dept_ids_set:
            for d in Department.objects.filter(id__in=dept_ids_set):
                dept_cache[str(d.id)] = d

        org_cache = {}
        if org_ids:
            for o in Organization.objects.filter(id__in=org_ids):
                org_cache[str(o.id)] = o

        leave_requests = LeaveRequest.objects.filter(
            employee_id__in=emp_ids,
            status='approved',
            start_date__lte=date_to,
            end_date__gte=date_from,
        ).select_related('leave_type')
        leave_map = {}
        for lr in leave_requests:
            leave_map.setdefault(str(lr.employee_id), []).append(lr)

        rows = []
        for idx, emp in enumerate(employees):
            dept = dept_cache.get(str(emp.department_id)) if emp.department_id else None
            emp_wdpw = emp.working_days_per_week if emp.working_days_per_week else (dept.working_days_per_week if dept else 6)
            org_obj = org_cache.get(str(emp.organization_id)) if emp.organization_id else None
            week_off_day = getattr(org_obj, 'week_off_day', 6) if org_obj else 6
            org_holidays = holidays_by_org.get(str(emp.organization_id), set()) if emp.organization_id else set()

            std_work_hours_per_day = 0
            if dept and dept.work_start_time and dept.work_end_time:
                from datetime import datetime as dt_mod
                start_dt = dt_mod.combine(date_from, dept.work_start_time)
                end_dt = dt_mod.combine(date_from, dept.work_end_time)
                std_work_hours_per_day = (end_dt - start_dt).total_seconds() / 3600
            if std_work_hours_per_day <= 0:
                std_work_hours_per_day = 8

            week_off_count = 0
            public_holiday_count = 0
            total_working_days = 0

            second_off_day = 5 if emp_wdpw == 5 else None

            d = date_from
            while d <= date_to:
                if d in org_holidays:
                    public_holiday_count += 1
                elif d.weekday() == week_off_day:
                    week_off_count += 1
                elif second_off_day is not None and d.weekday() == second_off_day:
                    week_off_count += 1
                else:
                    total_working_days += 1
                d += timedelta(days=1)

            total_working_hours = round(total_working_days * std_work_hours_per_day, 2)

            emp_records = att_map.get(str(emp.id), [])
            worked_days = 0
            total_worked_hours = 0
            absent_days = 0
            system_overtime = 0
            actual_overtime = 0

            att_dates_with_status = {}
            for r in emp_records:
                att_dates_with_status[r.date] = r
                if r.status in ('present', 'half_day'):
                    worked_days += 1
                    wh = float(r.work_hours) if r.work_hours else 0
                    total_worked_hours += wh
                    if wh > std_work_hours_per_day:
                        system_overtime += round(wh - std_work_hours_per_day, 2)
                    actual_overtime += float(r.overtime_hours) if r.overtime_hours else 0
                elif r.status == 'absent':
                    absent_days += 1

            d = date_from
            while d <= effective_to:
                if d not in att_dates_with_status and d not in org_holidays and d.weekday() != week_off_day:
                    if second_off_day is None or d.weekday() != second_off_day:
                        absent_days += 1
                d += timedelta(days=1)

            emp_leaves = leave_map.get(str(emp.id), [])
            total_leave_taken = 0
            total_paid_leave = 0
            total_unpaid_leave = 0
            for lr in emp_leaves:
                ls = max(lr.start_date, date_from)
                le = min(lr.end_date, date_to)
                if ls <= le:
                    leave_days = float(lr.days_count) if lr.is_half_day else (le - ls).days + 1
                    total_leave_taken += leave_days
                    if lr.leave_type and lr.leave_type.is_paid:
                        total_paid_leave += leave_days
                    else:
                        total_unpaid_leave += leave_days

            dept_name = dept.name if dept else ''
            rows.append({
                'sn': idx + 1,
                'employee_id': str(emp.id),
                'employee_name': emp.full_name,
                'department': dept_name,
                'total_days': total_days_in_month,
                'total_working_days': total_working_days,
                'week_off': week_off_count,
                'public_holidays': public_holiday_count,
                'total_working_hours': total_working_hours,
                'total_worked_days': worked_days,
                'total_worked_hours': round(total_worked_hours, 2),
                'total_leave_taken': total_leave_taken,
                'total_paid_leave': total_paid_leave,
                'total_unpaid_leave': total_unpaid_leave,
                'absent_days': absent_days,
                'system_overtime': round(system_overtime, 2),
                'actual_overtime': round(actual_overtime, 2),
            })

        return Response({
            'month': month_int,
            'year': year_int,
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'employees': rows,
        })


class DeviceMappingListView(APIView):
    @require_permission('hrms.device_mapping.read')
    def get(self, request):
        mappings = DeviceMapping.objects.all()
        data = []
        for m in mappings:
            emp_name = None
            try:
                emp = Employee.objects.get(id=m.employee_id)
                emp_name = emp.full_name
            except Employee.DoesNotExist:
                pass
            data.append({
                'id': str(m.id),
                'employee_id': str(m.employee_id),
                'employee_name': emp_name,
                'device_user_id': m.device_user_id,
                'device_name': m.device_name,
            })
        return Response(data)

    @require_permission('hrms.device_mapping.add')
    def post(self, request):
        data = request.data
        dm = DeviceMapping.objects.create(
            employee_id=data.get('employee_id'),
            device_user_id=data.get('device_user_id'),
            device_name=data.get('device_name', 'ZKT K40'),
        )
        return Response({'id': str(dm.id), 'message': 'Device mapping created'}, status=201)


class DeviceMappingDetailView(APIView):
    @require_permission('hrms.device_mapping.delete')
    def delete(self, request, dm_id):
        try:
            dm = DeviceMapping.objects.get(id=dm_id)
        except DeviceMapping.DoesNotExist:
            return Response({'message': 'Device mapping not found'}, status=404)
        dm.delete()
        return Response({'message': 'Device mapping deleted'})


class OnlineCheckInPermissionListView(APIView):
    @require_permission('hrms.online_checkin.read')
    def get(self, request):
        perms = OnlineCheckInPermission.objects.all()
        data = []
        for p in perms:
            emp_name = None
            try:
                emp = Employee.objects.get(id=p.employee_id)
                emp_name = emp.full_name
            except Employee.DoesNotExist:
                pass
            data.append({
                'id': str(p.id),
                'employee_id': str(p.employee_id),
                'employee_name': emp_name,
                'is_allowed': p.is_allowed,
                'require_photo': p.require_photo,
                'require_location': p.require_location,
            })
        return Response(data)

    @require_permission('hrms.online_checkin.add')
    def post(self, request):
        data = request.data
        perm, created = OnlineCheckInPermission.objects.update_or_create(
            employee_id=data.get('employee_id'),
            defaults={
                'is_allowed': data.get('is_allowed', True),
                'require_photo': data.get('require_photo', True),
                'require_location': data.get('require_location', True),
                'allowed_by': request.session.get('userId'),
            },
        )
        return Response({
            'id': str(perm.id),
            'is_allowed': perm.is_allowed,
            'message': 'Permission updated' if not created else 'Permission created',
        }, status=201 if created else 200)


class OnlineCheckInPermissionDetailView(APIView):
    @require_permission('hrms.online_checkin.delete')
    def delete(self, request, perm_id):
        try:
            perm = OnlineCheckInPermission.objects.get(id=perm_id)
        except OnlineCheckInPermission.DoesNotExist:
            return Response({'message': 'Permission not found'}, status=404)
        perm.delete()
        return Response({'message': 'Permission deleted'})


class SalaryStructureListView(APIView):
    @require_permission('hrms.salary.read')
    def get(self, request):
        structs = SalaryStructure.objects.all()
        emp_id = request.GET.get('employee_id')
        if emp_id:
            structs = structs.filter(employee_id=emp_id)
        status = request.GET.get('status')
        if status:
            structs = structs.filter(status=status)
        return Response([serialize_salary_structure(ss) for ss in structs])

    @require_permission('hrms.salary.add')
    def post(self, request):
        data = request.data
        ss = SalaryStructure.objects.create(
            employee_id=data.get('employee_id'),
            basic_salary=data.get('basic_salary', 0),
            allowances=data.get('allowances', {}),
            deductions=data.get('deductions', {}),
            cit_type=data.get('cit_type', 'none'),
            cit_value=data.get('cit_value', 0),
            ssf_applicable=data.get('ssf_applicable', False),
            ssf_employee_percentage=data.get('ssf_employee_percentage', 11.0),
            ssf_employer_percentage=data.get('ssf_employer_percentage', 20.0),
            tax_applicable=data.get('tax_applicable', True),
            effective_from=data.get('effective_from'),
            effective_to=data.get('effective_to'),
        )
        return Response(serialize_salary_structure(ss), status=201)


class SalaryStructureDetailView(APIView):
    @require_permission('hrms.salary.read')
    def get(self, request, ss_id):
        try:
            ss = SalaryStructure.objects.get(id=ss_id)
        except SalaryStructure.DoesNotExist:
            return Response({'message': 'Salary structure not found'}, status=404)
        return Response(serialize_salary_structure(ss))

    @require_permission('hrms.salary.update')
    def put(self, request, ss_id):
        try:
            ss = SalaryStructure.objects.get(id=ss_id)
        except SalaryStructure.DoesNotExist:
            return Response({'message': 'Salary structure not found'}, status=404)
        data = request.data
        for field in ['basic_salary', 'allowances', 'deductions', 'cit_type', 'cit_value',
                      'ssf_applicable', 'ssf_employee_percentage', 'ssf_employer_percentage',
                      'tax_applicable', 'effective_from', 'effective_to', 'status']:
            if field in data:
                setattr(ss, field, data[field])
        ss.save()
        return Response(serialize_salary_structure(ss))

    def patch(self, request, ss_id):
        return self.put(request, ss_id)


class PayrollRunListView(APIView):
    @require_permission('hrms.payroll.read')
    def get(self, request):
        runs = PayrollRun.objects.all()
        org_id = request.GET.get('organization_id')
        if org_id:
            runs = runs.filter(organization_id=org_id)
        return Response([{
            'id': str(r.id),
            'organization_id': str(r.organization_id),
            'organization_name': r.organization.name if r.organization else None,
            'fiscal_year_id': str(r.fiscal_year_id) if r.fiscal_year_id else None,
            'month': r.month,
            'year': r.year,
            'status': r.status,
            'currency': r.organization.currency if r.organization else 'NPR',
            'total_gross': float(r.total_gross),
            'total_deductions': float(r.total_deductions),
            'total_net': float(r.total_net),
            'total_employer_contribution': float(r.total_employer_contribution),
            'payslip_count': r.payslips.count(),
            'processed_at': r.processed_at.isoformat() if r.processed_at else None,
            'created_at': r.created_at.isoformat() if r.created_at else None,
        } for r in runs])

    @require_permission('hrms.payroll.add')
    def post(self, request):
        data = request.data
        org_id = data.get('organization_id')
        month = data.get('month')
        year = data.get('year')
        if not all([org_id, month, year]):
            return Response({'message': 'organization_id, month, year are required'}, status=400)

        existing = PayrollRun.objects.filter(organization_id=org_id, month=month, year=year).first()
        if existing:
            return Response({'message': 'Payroll run already exists for this period'}, status=400)

        pr = PayrollRun.objects.create(
            organization_id=org_id,
            month=month,
            year=year,
            fiscal_year_id=data.get('fiscal_year_id'),
            created_by=request.session.get('userId'),
            notes=data.get('notes'),
        )
        return Response({
            'id': str(pr.id),
            'month': pr.month,
            'year': pr.year,
            'status': pr.status,
        }, status=201)


class PayrollRunDetailView(APIView):
    @require_permission('hrms.payroll.read')
    def get(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)
        payslips = Payslip.objects.filter(payroll_run=pr)
        org_currency = 'NPR'
        try:
            org_obj = Organization.objects.get(id=pr.organization_id)
            org_currency = org_obj.currency or 'NPR'
        except Organization.DoesNotExist:
            pass
        return Response({
            'id': str(pr.id),
            'organization_id': str(pr.organization_id),
            'month': pr.month,
            'year': pr.year,
            'status': pr.status,
            'currency': org_currency,
            'total_gross': float(pr.total_gross),
            'total_deductions': float(pr.total_deductions),
            'total_net': float(pr.total_net),
            'total_employer_contribution': float(pr.total_employer_contribution),
            'notes': pr.notes,
            'payslips': [serialize_payslip(ps) for ps in payslips],
        })

    @require_permission('hrms.payroll.delete')
    def delete(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)
        if pr.status == 'paid':
            return Response({'message': 'Cannot delete a paid payroll run'}, status=400)
        pr.delete()
        return Response({'message': 'Payroll run deleted'})


class PayrollExportView(APIView):
    @require_permission('hrms.payroll.read')
    def get(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)

        payslips = Payslip.objects.filter(payroll_run=pr)
        rows_data = [serialize_payslip(ps) for ps in payslips]

        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payroll Report'

        month_name = ['January','February','March','April','May','June','July','August','September','October','November','December'][(pr.month or 1) - 1]
        org_name = ''
        try:
            org_name = Organization.objects.get(id=pr.organization_id).name
        except Organization.DoesNotExist:
            pass

        ws.merge_cells('A1:AD1')
        title_cell = ws['A1']
        title_cell.value = f'Payroll Report — {org_name} — {month_name} {pr.year}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        header_font = Font(bold=True, size=9)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')

        row1_headers = [
            ('S.N', 1, 1), ('Employee Name', 1, 1), ('Position', 1, 1),
            ('Join Date', 1, 1), ('Marital Status', 1, 1), ('Gender', 1, 1),
            ('Total Days', 1, 1), ('Total Worked Days', 1, 1),
            ('Extra Working Days', 1, 1), ('Total Paid Leave Days', 1, 1),
            ('Total Unpaid Leave Days', 1, 1),
        ]

        row = 3
        col = 1
        for label, rs, cs in row1_headers:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            if rs > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + rs - 1, end_column=col)
            col += 1

        income_start = col
        ws.merge_cells(start_row=row, start_column=income_start, end_row=row, end_column=income_start + 2)
        cell = ws.cell(row=row, column=income_start, value='Income')
        cell.font = header_font
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        cell.border = thin_border
        cell.alignment = center_align
        col = income_start + 3

        ded_start = col
        ws.merge_cells(start_row=row, start_column=ded_start, end_row=row, end_column=ded_start + 2)
        cell = ws.cell(row=row, column=ded_start, value='Deduction')
        cell.font = header_font
        cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        cell.border = thin_border
        cell.alignment = center_align
        col = ded_start + 3

        remaining_headers = [
            'Total Salary', 'Festival Bonus', 'CIT',
            'Taxable Amount (Yearly)',
            'SST', 'TDS', 'Single Women Tax Credit\n(10% of SST + TDS)',
            'Total Tax', 'Net Salary', 'Adjustment', 'Advance',
            'Payable Salary', 'Remarks',
        ]
        for label in remaining_headers:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            col += 1

        row2 = row + 1
        income_sub = ['Arrear Amount', 'Over-Time Pay', 'Total']
        for i, label in enumerate(income_sub):
            cell = ws.cell(row=row2, column=income_start + i, value=label)
            cell.font = header_font
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            cell.border = thin_border
            cell.alignment = center_align

        ded_sub = ['Leave Amount', 'Fine & Penalty', 'Total']
        for i, label in enumerate(ded_sub):
            cell = ws.cell(row=row2, column=ded_start + i, value=label)
            cell.font = header_font
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            cell.border = thin_border
            cell.alignment = center_align

        for i in range(11):
            ws.merge_cells(start_row=row, start_column=i + 1, end_row=row + 1, end_column=i + 1)

        data_row = row + 2
        for sn, ps in enumerate(rows_data, start=1):
            c = 1
            vals = [
                sn,
                ps.get('employee_name') or 'Unknown',
                ps.get('employee_position') or '',
                ps.get('employee_join_date') or '',
                (ps.get('employee_marital_status') or '').title(),
                (ps.get('employee_gender') or '').title(),
                ps.get('total_days', 0),
                ps.get('present_days', 0),
                0,
                ps.get('paid_leave_days', 0),
                ps.get('unpaid_leave_days', 0),
                0,
                0,
                ps.get('total_income', 0),
                ps.get('leave_deduction_amount', 0),
                0,
                ps.get('total_deductions', 0),
                ps.get('gross_salary', 0),
                ps.get('bonus_amount', 0),
                ps.get('cit_deduction', 0),
                ps.get('taxable_amount_yearly', 0),
                ps.get('sst', 0),
                ps.get('tds', 0),
                0,
                ps.get('total_tax', 0),
                ps.get('net_salary', 0),
                ps.get('travel_reimbursement', 0),
                ps.get('advance_deduction', 0),
                ps.get('payable_salary', 0),
                '',
            ]
            for v in vals:
                cell = ws.cell(row=data_row, column=c, value=v)
                cell.border = thin_border
                if isinstance(v, (int, float)) and c > 6:
                    cell.alignment = right_align
                    cell.number_format = '#,##0'
                c += 1
            data_row += 1

        total_row = data_row
        ws.cell(row=total_row, column=1, value='').border = thin_border
        total_cell = ws.cell(row=total_row, column=2, value=f'TOTAL ({len(rows_data)} staff)')
        total_cell.font = Font(bold=True, size=9)
        total_cell.border = thin_border
        for c in range(3, 7):
            ws.cell(row=total_row, column=c, value='').border = thin_border

        sum_cols = {
            7: 'total_days', 8: 'present_days', 10: 'paid_leave_days', 11: 'unpaid_leave_days',
            14: 'total_income', 15: 'leave_deduction_amount', 17: 'total_deductions',
            18: 'gross_salary', 19: 'bonus_amount', 20: 'cit_deduction',
            21: 'taxable_amount_yearly', 22: 'sst', 23: 'tds',
            25: 'total_tax', 26: 'net_salary', 27: 'travel_reimbursement',
            28: 'advance_deduction', 29: 'payable_salary',
        }
        for col_idx in range(7, 31):
            key = sum_cols.get(col_idx)
            val = sum(ps.get(key, 0) for ps in rows_data) if key else 0
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font = Font(bold=True, size=9)
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0'

        col_widths = [5, 22, 14, 12, 12, 10, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 10, 14, 10, 10, 18, 10, 12, 12, 10, 12, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="payroll_report_{month_name}_{pr.year}.xlsx"'
        return response


def calculate_nepal_tax(annual_income, marital_status='single', organization_id=None, country=None):
    db_slabs = TaxSlab.objects.filter(
        marital_status=marital_status, is_active=True
    ).order_by('slab_order')
    if country:
        country_slabs = db_slabs.filter(country=country)
        if country_slabs.exists():
            db_slabs = country_slabs
        else:
            db_slabs = db_slabs.filter(country__isnull=True)
    else:
        db_slabs = db_slabs.filter(country__isnull=True)
    if organization_id:
        org_slabs = db_slabs.filter(organization_id=organization_id)
        if org_slabs.exists():
            db_slabs = org_slabs
        else:
            db_slabs = db_slabs.filter(organization__isnull=True)

    if db_slabs.exists():
        slabs = []
        for s in db_slabs:
            if s.upper_limit is not None:
                slab_width = s.upper_limit - s.lower_limit
                slabs.append((slab_width, s.rate / Decimal('100')))
            else:
                slabs.append((None, s.rate / Decimal('100')))
    else:
        if marital_status == 'married':
            slabs = [
                (Decimal('600000'), Decimal('0.01')),
                (Decimal('200000'), Decimal('0.10')),
                (Decimal('300000'), Decimal('0.20')),
                (Decimal('1000000'), Decimal('0.30')),
                (None, Decimal('0.36')),
            ]
        else:
            slabs = [
                (Decimal('500000'), Decimal('0.01')),
                (Decimal('200000'), Decimal('0.10')),
                (Decimal('300000'), Decimal('0.20')),
                (Decimal('1000000'), Decimal('0.30')),
                (None, Decimal('0.36')),
            ]

    remaining = Decimal(str(annual_income))
    total_tax = Decimal('0')
    for limit, rate in slabs:
        if limit is None:
            total_tax += remaining * rate
            break
        if remaining <= 0:
            break
        taxable = min(remaining, Decimal(str(limit)))
        total_tax += taxable * rate
        remaining -= taxable

    return total_tax


class PayrollRunProcessView(APIView):
    @require_permission('hrms.payroll.process')
    def post(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)

        if pr.status in ('approved', 'paid'):
            return Response({'message': f'Cannot reprocess a payroll run that is already {pr.status}'}, status=400)

        pr.status = 'processing'
        pr.save()

        employees = Employee.objects.filter(organization_id=pr.organization_id, status='active')
        total_gross = Decimal('0')
        total_deductions = Decimal('0')
        total_net = Decimal('0')
        total_employer = Decimal('0')
        skipped_employees = []

        month_start = date(pr.year, pr.month, 1)
        if pr.month == 12:
            month_end = date(pr.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(pr.year, pr.month + 1, 1) - timedelta(days=1)

        if employees.count() == 0:
            pr.status = 'completed'
            pr.processed_at = timezone.now()
            pr.save()
            return Response({
                'message': 'No active employees found in this organization. Add employees and assign them to this organization first.',
                'payslip_count': 0,
                'total_employees': 0,
                'skipped_employees': [],
            })

        for emp in employees:
            sal = SalaryStructure.objects.filter(
                employee_id=emp.id, status='active',
                effective_from__lte=month_end,
            ).order_by('-effective_from').first()

            if not sal:
                skipped_employees.append({
                    'name': emp.full_name,
                    'reason': 'No active salary structure',
                })
                continue

            basic = sal.basic_salary
            allowances_total = sum(Decimal(str(v)) for v in sal.allowances.values()) if sal.allowances else Decimal('0')
            gross = basic + allowances_total

            att_records = AttendanceRecord.objects.filter(
                employee_id=emp.id, date__gte=month_start, date__lte=month_end
            )
            present_count = att_records.filter(status='present').count()
            absent_count = att_records.filter(status='absent').count()
            leave_count = att_records.filter(status='on_leave').count()
            late_count = att_records.filter(is_late=True).count()
            early_count = att_records.filter(is_early_leave=True).count()

            dept = None
            if emp.department_id:
                try:
                    dept = Department.objects.get(id=emp.department_id)
                except Department.DoesNotExist:
                    pass
            emp_wdpw = emp.working_days_per_week if emp.working_days_per_week else (dept.working_days_per_week if dept else 6)
            org_obj = None
            if emp.organization_id:
                try:
                    org_obj = Organization.objects.get(id=emp.organization_id)
                except Organization.DoesNotExist:
                    pass
            week_off_day = getattr(org_obj, 'week_off_day', 6) if org_obj else 6
            second_off_day = 5 if emp_wdpw == 5 else None
            total_calendar_days = (month_end - month_start).days + 1
            working_days = 0
            for d in range(total_calendar_days):
                day = month_start + timedelta(days=d)
                if day.weekday() == week_off_day:
                    continue
                if second_off_day is not None and day.weekday() == second_off_day:
                    continue
                working_days += 1
            holidays_in_month = Holiday.objects.filter(
                organization_id=pr.organization_id,
                date__gte=month_start,
                date__lte=month_end,
                is_optional=False,
            ).count()
            effective_working_days = max(working_days - holidays_in_month, 1)

            unpaid_leave_days = 0
            unpaid_leave_requests = LeaveRequest.objects.filter(
                employee_id=emp.id,
                status='approved',
                start_date__lte=month_end,
                end_date__gte=month_start,
            ).select_related('leave_type').only(
                'id', 'employee_id', 'start_date', 'end_date',
                'days_count', 'status', 'leave_type__is_paid',
                'leave_type__id', 'leave_type__name',
            )
            for lr in unpaid_leave_requests:
                if not lr.leave_type.is_paid:
                    overlap_start = max(lr.start_date, month_start)
                    overlap_end = min(lr.end_date, month_end)
                    unpaid_leave_days += (overlap_end - overlap_start).days + 1

            recorded_days = present_count + absent_count + leave_count
            unrecorded_absent = max(0, effective_working_days - recorded_days)
            total_absent = absent_count + unrecorded_absent
            unpaid_leave_days += total_absent

            per_day_salary = gross / Decimal(str(effective_working_days))
            unpaid_leave_deduction = (per_day_salary * Decimal(str(unpaid_leave_days))).quantize(Decimal('0.01'))
            adjusted_gross = gross - unpaid_leave_deduction

            bonus_records = Bonus.objects.filter(
                employee_id=emp.id, month=pr.month, year=pr.year,
                status__in=['approved', 'paid'],
            )
            bonus_total = sum(b.amount for b in bonus_records)

            expense_records = TravelExpense.objects.filter(
                employee_id=emp.id, month=pr.month, year=pr.year,
                status='approved', include_in_salary=True,
            )
            travel_total = sum(te.amount for te in expense_records)

            active_advances = AdvancePayment.objects.filter(
                employee_id=emp.id,
                status__in=['approved', 'active'],
            ).filter(
                Q(deduction_start_year__lt=pr.year) |
                Q(deduction_start_year=pr.year, deduction_start_month__lte=pr.month)
            )
            advance_ded = Decimal('0')
            for adv in active_advances:
                ded_amount = min(adv.monthly_deduction, adv.remaining_balance)
                advance_ded += ded_amount
                adv.total_deducted += ded_amount
                adv.remaining_balance -= ded_amount
                if adv.remaining_balance <= 0:
                    adv.remaining_balance = 0
                    adv.status = 'completed'
                else:
                    adv.status = 'active'
                adv.save()

            cit_deduction = Decimal('0')
            if sal.cit_type == 'percentage':
                cit_deduction = adjusted_gross * sal.cit_value / 100
            elif sal.cit_type == 'flat':
                cit_deduction = sal.cit_value

            ssf_employee = Decimal('0')
            ssf_employer = Decimal('0')
            if sal.ssf_applicable:
                ssf_employee = adjusted_gross * sal.ssf_employee_percentage / 100
                ssf_employer = adjusted_gross * sal.ssf_employer_percentage / 100

            taxable_income = adjusted_gross + bonus_total if bonus_records.filter(is_taxable=True).exists() else adjusted_gross
            tax_deduction = Decimal('0')
            if sal.tax_applicable:
                annual_taxable = (taxable_income - cit_deduction - ssf_employee) * 12
                annual_tax = calculate_nepal_tax(annual_taxable, emp.marital_status or 'single', organization_id=pr.organization_id, country=emp.country)
                tax_deduction = (annual_tax / 12).quantize(Decimal('0.01'))

            other_deductions = sal.deductions or {}
            other_ded_total = sum(Decimal(str(v)) for v in other_deductions.values()) if other_deductions else Decimal('0')

            total_ded = (cit_deduction + ssf_employee + tax_deduction +
                        other_ded_total + unpaid_leave_deduction + advance_ded)
            net = adjusted_gross + bonus_total + travel_total - (
                cit_deduction + ssf_employee + tax_deduction + other_ded_total + advance_ded
            )

            for br in bonus_records:
                if br.status == 'approved':
                    br.status = 'paid'
                    br.save()
            for te in expense_records:
                te.status = 'reimbursed'
                te.save()

            Payslip.objects.update_or_create(
                payroll_run=pr,
                employee_id=emp.id,
                defaults={
                    'month': pr.month,
                    'year': pr.year,
                    'basic_salary': basic,
                    'allowances': sal.allowances or {},
                    'gross_salary': gross,
                    'cit_deduction': cit_deduction,
                    'ssf_employee_deduction': ssf_employee,
                    'ssf_employer_contribution': ssf_employer,
                    'tax_deduction': tax_deduction,
                    'bonus_amount': bonus_total,
                    'travel_reimbursement': travel_total,
                    'advance_deduction': advance_ded,
                    'unpaid_leave_deduction': unpaid_leave_deduction,
                    'other_deductions': other_deductions,
                    'total_deductions': total_ded,
                    'net_salary': net,
                    'working_days': effective_working_days,
                    'present_days': present_count,
                    'absent_days': total_absent,
                    'leave_days': leave_count,
                    'late_count': late_count,
                    'early_leave_count': early_count,
                    'status': 'generated',
                },
            )

            total_gross += gross
            total_deductions += total_ded
            total_net += net
            total_employer += ssf_employer

        pr.status = 'processed'
        pr.total_gross = total_gross
        pr.total_deductions = total_deductions
        pr.total_net = total_net
        pr.total_employer_contribution = total_employer
        pr.processed_at = timezone.now()
        pr.save()

        payslip_count = pr.payslips.count()
        msg = 'Payroll processed successfully'
        if payslip_count == 0 and skipped_employees:
            msg = f'Payroll processed but no payslips generated. {len(skipped_employees)} employee(s) skipped because they have no active salary structure. Please set up salary structures in Staff & Salary first.'
        elif skipped_employees:
            msg = f'Payroll processed. {payslip_count} payslip(s) generated. {len(skipped_employees)} employee(s) skipped (no salary structure).'

        return Response({
            'message': msg,
            'payslip_count': payslip_count,
            'total_employees': employees.count(),
            'skipped_employees': skipped_employees,
            'total_gross': float(total_gross),
            'total_net': float(total_net),
        })


class PayrollRunApproveView(APIView):
    @require_permission('hrms.payroll.process')
    def post(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)
        if pr.status not in ('processed', 'completed'):
            return Response({'message': f'Cannot approve a payroll run with status "{pr.status}". Must be processed first.'}, status=400)
        pr.status = 'approved'
        pr.save()
        pr.payslips.update(status='approved')
        return Response({'message': 'Payroll run approved by management', 'status': pr.status})


class PayrollRunMarkPaidView(APIView):
    @require_permission('hrms.payroll.process')
    def post(self, request, pr_id):
        try:
            pr = PayrollRun.objects.get(id=pr_id)
        except PayrollRun.DoesNotExist:
            return Response({'message': 'Payroll run not found'}, status=404)
        if pr.status != 'approved':
            return Response({'message': f'Cannot mark as paid. Payroll must be approved first (current: {pr.status}).'}, status=400)
        pr.status = 'paid'
        pr.save()
        pr.payslips.update(status='paid')
        return Response({'message': 'Payroll marked as paid. Payslips are now visible to employees.', 'status': pr.status})


class PayslipUpdateView(APIView):
    @require_permission('hrms.payroll.process')
    def put(self, request, ps_id):
        try:
            ps = Payslip.objects.get(id=ps_id)
        except Payslip.DoesNotExist:
            return Response({'message': 'Payslip not found'}, status=404)
        pr = ps.payroll_run
        if pr.status in ('approved', 'paid'):
            return Response({'message': 'Cannot edit payslip after approval/payment'}, status=400)
        data = request.data
        editable_fields = [
            'basic_salary', 'gross_salary', 'cit_deduction', 'ssf_employee_deduction',
            'ssf_employer_contribution', 'tax_deduction', 'bonus_amount',
            'travel_reimbursement', 'advance_deduction', 'unpaid_leave_deduction',
            'total_deductions', 'net_salary',
        ]
        for f in editable_fields:
            if f in data:
                setattr(ps, f, Decimal(str(data[f])))
        if 'allowances' in data:
            ps.allowances = data['allowances']
        if 'other_deductions' in data:
            ps.other_deductions = data['other_deductions']
        ps.save()

        payslips = Payslip.objects.filter(payroll_run=pr)
        pr.total_gross = sum(p.gross_salary for p in payslips)
        pr.total_deductions = sum(p.total_deductions for p in payslips)
        pr.total_net = sum(p.net_salary for p in payslips)
        pr.total_employer_contribution = sum(p.ssf_employer_contribution for p in payslips)
        pr.save()

        return Response(serialize_payslip(ps))


class PayslipListView(APIView):
    @require_permission('hrms.payslip.read')
    def get(self, request):
        payslips = Payslip.objects.all()
        emp_id = request.GET.get('employee_id')
        if emp_id:
            payslips = payslips.filter(employee_id=emp_id)
        pr_id = request.GET.get('payroll_run_id')
        if pr_id:
            payslips = payslips.filter(payroll_run_id=pr_id)
        month = request.GET.get('month')
        year = request.GET.get('year')
        if month:
            payslips = payslips.filter(month=month)
        if year:
            payslips = payslips.filter(year=year)
        return Response([serialize_payslip(ps) for ps in payslips])


class PayslipDetailView(APIView):
    @require_permission('hrms.payslip.read')
    def get(self, request, ps_id):
        try:
            ps = Payslip.objects.get(id=ps_id)
        except Payslip.DoesNotExist:
            return Response({'message': 'Payslip not found'}, status=404)
        return Response(serialize_payslip(ps))


class PayslipPDFView(APIView):
    @require_permission('hrms.payslip.read')
    def get(self, request, ps_id):
        try:
            ps = Payslip.objects.get(id=ps_id)
        except Payslip.DoesNotExist:
            return Response({'message': 'Payslip not found'}, status=404)
        try:
            emp = Employee.objects.get(id=ps.employee_id)
        except Employee.DoesNotExist:
            return Response({'message': 'Employee not found'}, status=404)
        pr = ps.payroll_run
        org = pr.organization if pr else None
        pdf_bytes = generate_payslip_pdf(ps, emp, org)
        from django.http import HttpResponse
        month_name = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][ps.month - 1]
        filename = f"Payslip_{emp.full_name.replace(' ', '_')}_{month_name}_{ps.year}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PayslipBulkPDFView(APIView):
    @require_permission('hrms.payslip.read')
    def post(self, request):
        payslip_ids = request.data.get('payslip_ids', [])
        payroll_run_id = request.data.get('payroll_run_id')

        if payroll_run_id:
            payslips = Payslip.objects.filter(payroll_run_id=payroll_run_id)
        elif payslip_ids:
            payslips = Payslip.objects.filter(id__in=payslip_ids)
        else:
            return Response({'message': 'Provide payslip_ids or payroll_run_id'}, status=400)

        if not payslips.exists():
            return Response({'message': 'No payslips found'}, status=404)

        import zipfile, io
        zip_buffer = io.BytesIO()
        month_name = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for ps in payslips:
                try:
                    emp = Employee.objects.get(id=ps.employee_id)
                except Employee.DoesNotExist:
                    continue
                pr = ps.payroll_run
                org = pr.organization if pr else None
                pdf_bytes = generate_payslip_pdf(ps, emp, org)
                mn = month_name[ps.month - 1]
                fname = f"Payslip_{emp.full_name.replace(' ', '_')}_{mn}_{ps.year}.pdf"
                zf.writestr(fname, pdf_bytes)
        zip_buffer.seek(0)
        from django.http import HttpResponse
        first_ps = payslips.first()
        mn = month_name[first_ps.month - 1] if first_ps else 'Unknown'
        yr = first_ps.year if first_ps else ''
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Payslips_{mn}_{yr}.zip"'
        return response


CURRENCY_SYMBOLS = {
    'NPR': 'Rs.', 'INR': '₹', 'USD': '$', 'EUR': '€', 'GBP': '£', 'AUD': 'A$',
    'CAD': 'C$', 'SGD': 'S$', 'AED': 'د.إ', 'SAR': 'ر.س', 'MYR': 'RM',
    'THB': '฿', 'PHP': '₱', 'IDR': 'Rp', 'BDT': '৳', 'PKR': '₨', 'LKR': 'Rs',
    'KRW': '₩', 'JPY': '¥', 'CNY': '¥', 'HKD': 'HK$', 'NZD': 'NZ$',
    'ZAR': 'R', 'BRL': 'R$', 'MXN': 'Mex$', 'TRY': '₺', 'RUB': '₽',
    'CHF': 'CHF', 'SEK': 'kr', 'NOK': 'kr', 'DKK': 'kr', 'PLN': 'zł',
    'QAR': 'ر.ق', 'KWD': 'د.ك', 'BHD': 'BD', 'OMR': 'ر.ع.',
}


def generate_payslip_pdf(ps, emp, org):
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=15*mm, bottomMargin=15*mm)
    page_w = A4[0] - 36*mm
    styles = getSampleStyleSheet()

    c_primary = colors.HexColor('#1e3a5f')
    c_accent = colors.HexColor('#2563eb')
    c_gold = colors.HexColor('#b8860b')
    c_gold_bg = colors.HexColor('#fdf8e8')
    c_green = colors.HexColor('#059669')
    c_green_bg = colors.HexColor('#ecfdf5')
    c_red = colors.HexColor('#dc2626')
    c_red_bg = colors.HexColor('#fef2f2')
    c_grey = colors.HexColor('#64748b')
    c_lightgrey = colors.HexColor('#f8fafc')
    c_border = colors.HexColor('#e2e8f0')
    c_white = colors.white
    c_dark = colors.HexColor('#1e293b')
    c_section_bg = colors.HexColor('#1e3a5f')
    c_section_red = colors.HexColor('#991b1b')
    c_section_green = colors.HexColor('#065f46')

    org_name_style = ParagraphStyle('OrgName', fontName='Helvetica-Bold', fontSize=16, textColor=c_primary, alignment=TA_LEFT, spaceAfter=1*mm)
    org_detail_style = ParagraphStyle('OrgDetail', fontName='Helvetica', fontSize=8, textColor=c_grey, leading=11)
    slip_title_style = ParagraphStyle('SlipTitle', fontName='Helvetica-Bold', fontSize=12, textColor=c_white, alignment=TA_LEFT)
    lbl_style = ParagraphStyle('Lbl', fontName='Helvetica', fontSize=7.5, textColor=c_grey, leading=10)
    val_style = ParagraphStyle('Val', fontName='Helvetica-Bold', fontSize=8.5, textColor=c_dark, leading=11)
    section_hdr = ParagraphStyle('SecHdr', fontName='Helvetica-Bold', fontSize=9, textColor=c_white)
    row_lbl = ParagraphStyle('RowLbl', fontName='Helvetica', fontSize=8.5, textColor=colors.HexColor('#334155'))
    row_val = ParagraphStyle('RowVal', fontName='Helvetica-Bold', fontSize=8.5, textColor=c_dark, alignment=TA_RIGHT)
    footer_style = ParagraphStyle('Footer', fontName='Helvetica-Oblique', fontSize=7, textColor=c_grey, alignment=TA_CENTER)

    month_names = ['January','February','March','April','May','June','July','August','September','October','November','December']

    cur_code = getattr(org, 'currency', 'NPR') or 'NPR'
    cur_sym = CURRENCY_SYMBOLS.get(cur_code, cur_code)

    def fmt(val):
        return f"{cur_sym}{float(val):,.2f}"

    elements = []

    org_name = org.name if org else 'Company'
    org_address = getattr(org, 'address', None) or ''
    org_phone = getattr(org, 'phone', None) or ''
    org_email = getattr(org, 'email', None) or ''
    org_pan = getattr(org, 'pan_number', None) or ''
    org_reg = getattr(org, 'registration_number', None) or ''
    org_reg_label = getattr(org, 'registration_label', None) or 'Registration No.'
    org_pan_label_val = getattr(org, 'pan_label', None) or 'PAN No.'

    period_label = f"{month_names[ps.month - 1]} {ps.year}"
    pay_date_label = ps.updated_at.strftime('%d %b %Y') if ps.updated_at else ''

    hdr_left_parts = [f"<b>{org_name}</b>"]
    if org_address:
        hdr_left_parts.append(org_address)
    detail_parts = []
    if org_reg:
        detail_parts.append(f"{org_reg_label}: {org_reg}")
    if org_pan:
        detail_parts.append(f"{org_pan_label_val}: {org_pan}")
    if org_phone:
        detail_parts.append(org_phone)
    if org_email:
        detail_parts.append(org_email)
    if detail_parts:
        hdr_left_parts.append(' | '.join(detail_parts))

    hdr_right_parts = [f"<b>Pay Period:</b> {period_label}"]
    if pay_date_label:
        hdr_right_parts.append(f"<b>Pay Date:</b> {pay_date_label}")

    hdr_left_style = ParagraphStyle('HdrL', fontName='Helvetica', fontSize=8.5, textColor=c_dark, leading=13)
    hdr_right_style = ParagraphStyle('HdrR', fontName='Helvetica', fontSize=8.5, textColor=c_dark, leading=13, alignment=TA_RIGHT)

    hdr_data = [[
        Paragraph('<br/>'.join(hdr_left_parts), hdr_left_style),
        Paragraph('<br/>'.join(hdr_right_parts), hdr_right_style),
    ]]
    hdr_tbl = RTable(hdr_data, colWidths=[page_w*0.6, page_w*0.4])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(hdr_tbl)
    elements.append(Spacer(1, 2*mm))

    title_data = [[Paragraph('PAY SLIP', slip_title_style)]]
    title_tbl = RTable(title_data, colWidths=[page_w])
    title_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), c_section_bg),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(title_tbl)
    elements.append(Spacer(1, 3*mm))

    employment_type = getattr(emp, 'employment_type', '') or ''
    emp_type_display = employment_type.replace('_', ' ').title() if employment_type else '-'
    join_date = getattr(emp, 'join_date', None)
    join_display = join_date.strftime('%d %b %Y') if join_date else '-'

    bank_acct = getattr(emp, 'bank_account_no', None) or '-'
    bank_name = getattr(emp, 'bank_name', None) or '-'
    bank_branch = getattr(emp, 'bank_branch', None) or ''

    info_data = [
        [Paragraph('Employee', lbl_style), Paragraph(emp.full_name, val_style),
         Paragraph('Employee ID', lbl_style), Paragraph(str(emp.id)[:8].upper(), val_style)],
        [Paragraph('Department', lbl_style), Paragraph(emp.department or '-', val_style),
         Paragraph('Position', lbl_style), Paragraph(emp.position or '-', val_style)],
        [Paragraph('Employment Type', lbl_style), Paragraph(emp_type_display, val_style),
         Paragraph('Date Joined', lbl_style), Paragraph(join_display, val_style)],
        [Paragraph(_get_tax_id_label(getattr(emp, 'country', None)), lbl_style), Paragraph(emp.pan_no or '-', val_style),
         Paragraph('Citizenship No', lbl_style), Paragraph(getattr(emp, 'citizenship_no', None) or '-', val_style)],
    ]
    info_tbl = RTable(info_data, colWidths=[page_w*0.17, page_w*0.33, page_w*0.17, page_w*0.33])
    info_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LINEBELOW', (0,0), (-1,-1), 0.4, c_border),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 3*mm))

    absent_days = getattr(ps, 'absent_days', 0) or 0
    leave_days = getattr(ps, 'leave_days', 0) or 0
    late_count = getattr(ps, 'late_count', 0) or 0

    att_lbl_s = ParagraphStyle('AttLbl', fontName='Helvetica', fontSize=7.5, textColor=c_grey, alignment=TA_CENTER, leading=10)
    att_val_s = ParagraphStyle('AttVal', fontName='Helvetica-Bold', fontSize=9, textColor=c_dark, alignment=TA_CENTER, leading=12)

    att_data = [
        [Paragraph('Working Days', att_lbl_s), Paragraph('Leave', att_lbl_s)],
        [Paragraph(str(ps.working_days), att_val_s), Paragraph(str(leave_days), att_val_s)],
    ]
    att_tbl = RTable(att_data, colWidths=[page_w*0.5]*2)
    att_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,-1), c_lightgrey),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('BOTTOMPADDING', (0,0), (-1,0), 1),
        ('TOPPADDING', (0,1), (-1,1), 1),
        ('BOTTOMPADDING', (0,1), (-1,1), 4),
        ('BOX', (0,0), (-1,-1), 0.5, c_border),
        ('LINEBEFORE', (1,0), (1,-1), 0.3, c_border),
    ]))
    elements.append(att_tbl)
    elements.append(Spacer(1, 5*mm))

    earn_rows = []
    earn_rows.append(('Basic Salary', float(ps.basic_salary)))
    allow_dict = ps.allowances or {}
    for k, v in allow_dict.items():
        earn_rows.append((k.replace('_', ' ').title(), float(v)))
    if float(ps.bonus_amount) > 0:
        earn_rows.append(('Bonus', float(ps.bonus_amount)))
    if float(ps.travel_reimbursement) > 0:
        earn_rows.append(('Travel Reimbursement', float(ps.travel_reimbursement)))
    gross_val = float(ps.gross_salary) + float(ps.bonus_amount) + float(ps.travel_reimbursement)

    entitle_data = []
    entitle_hdr = [
        Paragraph('Entitlements', section_hdr),
        Paragraph('Total', ParagraphStyle('EHT', fontName='Helvetica-Bold', fontSize=9, textColor=c_white, alignment=TA_RIGHT)),
    ]
    entitle_data.append(entitle_hdr)
    for desc, amt in earn_rows:
        entitle_data.append([
            Paragraph(desc, row_lbl),
            Paragraph(fmt(amt), row_val),
        ])
    entitle_data.append([
        Paragraph('Total Gross Payment', ParagraphStyle('TGP', fontName='Helvetica-Bold', fontSize=9, textColor=c_primary)),
        Paragraph(fmt(gross_val), ParagraphStyle('TGPV', fontName='Helvetica-Bold', fontSize=9.5, textColor=c_primary, alignment=TA_RIGHT)),
    ])
    entitle_tbl = RTable(entitle_data, colWidths=[page_w*0.7, page_w*0.3])
    e_style = [
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (0,0), 8),
        ('BACKGROUND', (0,0), (-1,0), c_accent),
        ('LINEBELOW', (0,-1), (-1,-1), 1.2, c_primary),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#eef2ff')),
    ]
    for i in range(1, len(entitle_data) - 1):
        bg = c_lightgrey if i % 2 == 0 else c_white
        e_style.append(('BACKGROUND', (0,i), (-1,i), bg))
        e_style.append(('LINEBELOW', (0,i), (-1,i), 0.3, c_border))
    entitle_tbl.setStyle(TableStyle(e_style))
    elements.append(entitle_tbl)
    elements.append(Spacer(1, 4*mm))

    ded_rows = []
    if float(ps.unpaid_leave_deduction) > 0:
        ded_rows.append(('Unpaid Leave Deduction', float(ps.unpaid_leave_deduction)))
    if float(ps.cit_deduction) > 0:
        ded_rows.append(('CIT (Provident Fund)', float(ps.cit_deduction)))
    if float(ps.ssf_employee_deduction) > 0:
        ded_rows.append(('SSF (Employee Contribution)', float(ps.ssf_employee_deduction)))
    if float(ps.tax_deduction) > 0:
        ded_rows.append(('Income Tax (TDS)', float(ps.tax_deduction)))
    if float(ps.advance_deduction) > 0:
        ded_rows.append(('Advance Deduction', float(ps.advance_deduction)))
    other_ded = ps.other_deductions or {}
    for k, v in other_ded.items():
        ded_rows.append((k.replace('_', ' ').title(), float(v)))

    ded_data = []
    ded_hdr = [
        Paragraph('Deductions', section_hdr),
        Paragraph('Amount', ParagraphStyle('DHT', fontName='Helvetica-Bold', fontSize=9, textColor=c_white, alignment=TA_RIGHT)),
    ]
    ded_data.append(ded_hdr)
    if ded_rows:
        for desc, amt in ded_rows:
            ded_data.append([
                Paragraph(desc, row_lbl),
                Paragraph(f"-{fmt(amt)}", ParagraphStyle('DedVal', fontName='Helvetica-Bold', fontSize=8.5, textColor=c_red, alignment=TA_RIGHT)),
            ])
    else:
        ded_data.append([Paragraph('No deductions', row_lbl), Paragraph('-', row_val)])
    ded_data.append([
        Paragraph('Total Deductions', ParagraphStyle('TDP', fontName='Helvetica-Bold', fontSize=9, textColor=c_red)),
        Paragraph(f"-{fmt(ps.total_deductions)}", ParagraphStyle('TDPV', fontName='Helvetica-Bold', fontSize=9.5, textColor=c_red, alignment=TA_RIGHT)),
    ])
    ded_tbl = RTable(ded_data, colWidths=[page_w*0.7, page_w*0.3])
    d_style = [
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (0,0), 8),
        ('BACKGROUND', (0,0), (-1,0), c_section_red),
        ('LINEBELOW', (0,-1), (-1,-1), 1.2, c_red),
        ('BACKGROUND', (0,-1), (-1,-1), c_red_bg),
    ]
    for i in range(1, len(ded_data) - 1):
        bg = c_lightgrey if i % 2 == 0 else c_white
        d_style.append(('BACKGROUND', (0,i), (-1,i), bg))
        d_style.append(('LINEBELOW', (0,i), (-1,i), 0.3, c_border))
    ded_tbl.setStyle(TableStyle(d_style))
    elements.append(ded_tbl)
    elements.append(Spacer(1, 4*mm))

    if float(ps.ssf_employer_contribution) > 0:
        super_data = []
        super_hdr = [
            Paragraph('Superannuation / SSF (Employer)', ParagraphStyle('SHdr', fontName='Helvetica-Bold', fontSize=9, textColor=c_white)),
            Paragraph('Amount', ParagraphStyle('SHT', fontName='Helvetica-Bold', fontSize=9, textColor=c_white, alignment=TA_RIGHT)),
        ]
        super_data.append(super_hdr)
        super_data.append([
            Paragraph('SSF Employer Contribution', row_lbl),
            Paragraph(fmt(ps.ssf_employer_contribution), row_val),
        ])
        super_data.append([
            Paragraph('Total Contribution', ParagraphStyle('STP', fontName='Helvetica-Bold', fontSize=9, textColor=c_gold)),
            Paragraph(fmt(ps.ssf_employer_contribution), ParagraphStyle('STPV', fontName='Helvetica-Bold', fontSize=9.5, textColor=c_gold, alignment=TA_RIGHT)),
        ])
        super_tbl = RTable(super_data, colWidths=[page_w*0.7, page_w*0.3])
        super_tbl.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (0,0), 8),
            ('BACKGROUND', (0,0), (-1,0), c_gold),
            ('BACKGROUND', (0,1), (-1,1), c_white),
            ('LINEBELOW', (0,1), (-1,1), 0.3, c_border),
            ('BACKGROUND', (0,-1), (-1,-1), c_gold_bg),
            ('LINEBELOW', (0,-1), (-1,-1), 1.2, c_gold),
        ]))
        elements.append(super_tbl)
        elements.append(Spacer(1, 4*mm))

    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_section_green, spaceAfter=2*mm))

    net_lbl_s = ParagraphStyle('NL', fontName='Helvetica-Bold', fontSize=9, textColor=c_dark)
    net_val_s = ParagraphStyle('NV', fontName='Helvetica-Bold', fontSize=9, textColor=c_dark, alignment=TA_RIGHT)
    net_val_big = ParagraphStyle('NVB', fontName='Helvetica-Bold', fontSize=13, textColor=c_green, alignment=TA_RIGHT)

    net_info = []
    bank_detail_parts = []
    if bank_name and bank_name != '-':
        bank_detail_parts.append(f"<b>Bank:</b> {bank_name}")
    if bank_branch:
        bank_detail_parts.append(f"<b>Branch:</b> {bank_branch}")
    if bank_acct and bank_acct != '-':
        bank_detail_parts.append(f"<b>Account:</b> {bank_acct}")

    if bank_detail_parts:
        bank_style = ParagraphStyle('BankD', fontName='Helvetica', fontSize=8, textColor=c_grey, leading=12)
        net_info.append([
            Paragraph('<br/>'.join(bank_detail_parts), bank_style),
            '',
        ])

    net_info.append([
        Paragraph('Total Net Payment', ParagraphStyle('TNP', fontName='Helvetica-Bold', fontSize=11, textColor=c_primary)),
        Paragraph(fmt(ps.net_salary), net_val_big),
    ])

    net_tbl = RTable(net_info, colWidths=[page_w*0.55, page_w*0.45])
    net_style_cmds = [
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,0), (-1,-1), c_green_bg),
        ('BOX', (0,0), (-1,-1), 1.5, c_green),
        ('LINEABOVE', (0,-1), (-1,-1), 0.5, c_border),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
    ]
    net_tbl.setStyle(TableStyle(net_style_cmds))
    elements.append(net_tbl)
    elements.append(Spacer(1, 6*mm))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=c_border, spaceAfter=3*mm))
    keep_style = ParagraphStyle('Keep', fontName='Helvetica-Bold', fontSize=7.5, textColor=c_grey, alignment=TA_CENTER)
    elements.append(Paragraph("PLEASE KEEP A COPY FOR YOUR RECORDS", keep_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph("This is a system-generated payslip. No signature is required.", footer_style))
    elements.append(Spacer(1, 1*mm))
    gen_date_style = ParagraphStyle('GenDate', fontName='Helvetica', fontSize=6.5, textColor=c_grey, alignment=TA_CENTER)
    from datetime import datetime as dt_now
    elements.append(Paragraph(f"Generated on {dt_now.now().strftime('%d %b %Y, %I:%M %p')}", gen_date_style))

    doc.build(elements)
    return buf.getvalue()


class NotificationSettingView(APIView):
    @require_permission('hrms.notification.read')
    def get(self, request):
        org_id = request.GET.get('organization_id')
        if not org_id:
            return Response({'message': 'organization_id is required'}, status=400)
        ns = NotificationSetting.objects.filter(organization_id=org_id).first()
        if not ns:
            return Response(None)
        return Response({
            'id': str(ns.id),
            'organization_id': str(ns.organization_id),
            'late_arrival_notify_employee': ns.late_arrival_notify_employee,
            'late_arrival_notify_manager': ns.late_arrival_notify_manager,
            'late_arrival_notify_hr': ns.late_arrival_notify_hr,
            'early_leave_notify_employee': ns.early_leave_notify_employee,
            'early_leave_notify_manager': ns.early_leave_notify_manager,
            'early_leave_notify_hr': ns.early_leave_notify_hr,
            'hr_email': ns.hr_email,
            'late_email_subject': ns.late_email_subject,
            'late_email_template': ns.late_email_template,
            'early_leave_email_subject': ns.early_leave_email_subject,
            'early_leave_email_template': ns.early_leave_email_template,
            'no_checkout_notify_employee': ns.no_checkout_notify_employee,
            'no_checkout_email_subject': ns.no_checkout_email_subject,
            'no_checkout_email_template': ns.no_checkout_email_template,
            'cc_emails': ns.cc_emails,
        })

    @require_permission('hrms.notification.update')
    def post(self, request):
        data = request.data
        org_id = data.get('organization_id')
        if not org_id:
            return Response({'message': 'organization_id is required'}, status=400)

        ns, _ = NotificationSetting.objects.update_or_create(
            organization_id=org_id,
            defaults={k: v for k, v in data.items() if k != 'organization_id'},
        )
        return Response({'message': 'Notification settings updated'})


class MyProfileView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        org = None
        dept = None
        if employee.organization_id:
            org = Organization.objects.filter(id=employee.organization_id).first()
        if employee.department_id:
            dept = Department.objects.filter(id=employee.department_id).first()

        salary_structure = None
        ss = SalaryStructure.objects.filter(employee_id=employee.id, status='active').first()
        gross_salary = 0
        if ss:
            gross_salary = float(ss.basic_salary)
            for k, v in (ss.allowances or {}).items():
                gross_salary += float(v)
            salary_structure = {
                'basic_salary': float(ss.basic_salary),
                'allowances': ss.allowances or {},
                'deductions': ss.deductions or {},
                'gross_salary': gross_salary,
                'cit_type': ss.cit_type,
                'cit_value': float(ss.cit_value),
                'ssf_applicable': ss.ssf_applicable,
                'ssf_employee_percentage': float(ss.ssf_employee_percentage),
                'ssf_employer_percentage': float(ss.ssf_employer_percentage),
                'tax_applicable': ss.tax_applicable,
                'effective_from': ss.effective_from.isoformat() if ss.effective_from else None,
            }

        now = timezone.now()
        att_records = AttendanceRecord.objects.filter(
            employee_id=employee.id, date__month=now.month, date__year=now.year
        )
        present_statuses = ['present', 'half_day']
        leave_statuses = ['on_leave', 'paid_leave', 'unpaid_leave', 'sick_leave', 'casual_leave']
        att_summary = {
            'month': now.month,
            'year': now.year,
            'present': att_records.filter(status__in=present_statuses).count(),
            'absent': att_records.filter(status='absent').count(),
            'late': att_records.filter(is_late=True).count(),
            'on_leave': att_records.filter(status__in=leave_statuses).count(),
            'total_records': att_records.count(),
        }

        tax_summary = Payslip.objects.filter(
            employee_id=employee.id, year=now.year
        ).aggregate(
            total_tax=Sum('tax_deduction'),
            total_cit=Sum('cit_deduction'),
            total_ssf=Sum('ssf_employee_deduction'),
        )

        leave_balances = []
        for lb in LeaveBalance.objects.filter(employee_id=employee.id).select_related('leave_type'):
            leave_balances.append({
                'leave_type': lb.leave_type.name,
                'leave_type_code': lb.leave_type.code,
                'color': lb.leave_type.color,
                'allocated': float(lb.allocated_days),
                'used': float(lb.used_days),
                'carried_forward': float(lb.carried_forward_days),
                'remaining': float(lb.allocated_days + lb.carried_forward_days - lb.used_days),
            })

        recent_payslips = []
        for ps in Payslip.objects.filter(employee_id=employee.id).order_by('-year', '-month')[:6]:
            recent_payslips.append({
                'id': str(ps.id),
                'month': ps.month,
                'year': ps.year,
                'basic_salary': float(ps.basic_salary),
                'gross_salary': float(ps.gross_salary),
                'total_deductions': float(ps.total_deductions),
                'net_salary': float(ps.net_salary),
                'working_days': ps.working_days,
                'present_days': ps.present_days,
                'status': ps.status,
                'view_token': ps.view_token,
            })

        expenses = []
        for e in TravelExpense.objects.filter(employee_id=employee.id).order_by('-expense_date')[:10]:
            expenses.append({
                'id': str(e.id),
                'category': e.category,
                'description': e.description,
                'amount': float(e.amount),
                'expense_date': e.expense_date.isoformat() if e.expense_date else None,
                'status': e.status,
            })

        currency = org.currency if org else 'NPR'

        user_perms = set(request.session.get('userPermissions', []))
        can_expense = 'hrms.expense.read' in user_perms or 'hrms.expense.add' in user_perms
        can_submit_expense = 'hrms.expense.add' in user_perms

        return Response({
            'id': str(employee.id),
            'full_name': employee.full_name,
            'email': employee.email,
            'phone': employee.phone,
            'position': employee.position,
            'department': dept.name if dept else employee.department,
            'organization': org.name if org else None,
            'organization_currency': currency,
            'join_date': employee.join_date.isoformat() if employee.join_date else None,
            'profile_photo_url': employee.profile_photo_url,
            'employment_type': employee.employment_type,
            'status': employee.status,
            'gender': employee.gender,
            'country': employee.country,
            'marital_status': employee.marital_status,
            'date_of_birth': employee.date_of_birth.isoformat() if employee.date_of_birth else None,
            'citizenship_no': employee.citizenship_no,
            'pan_no': employee.pan_no,
            'passport_number': employee.passport_number,
            'employee_id_number': employee.employee_id_number,
            'permanent_address': employee.permanent_address,
            'temporary_address': getattr(employee, 'temporary_address', None),
            'probation_end_date': employee.probation_end_date.isoformat() if employee.probation_end_date else None,
            'contract_end_date': employee.contract_end_date.isoformat() if employee.contract_end_date else None,
            'emergency_contact_name': employee.emergency_contact_name,
            'emergency_contact_phone': employee.emergency_contact_phone,
            'salary_amount': float(employee.salary_amount) if employee.salary_amount else None,
            'salary_currency': employee.salary_currency or currency,
            'bank_name': employee.bank_name,
            'bank_account_number': employee.bank_account_number,
            'bank_branch': employee.bank_branch,
            'salary_structure': salary_structure,
            'gross_salary': gross_salary,
            'attendance_summary': att_summary,
            'tax_summary': {
                'year': now.year,
                'total_tax': float(tax_summary['total_tax'] or 0),
                'total_cit': float(tax_summary['total_cit'] or 0),
                'total_ssf': float(tax_summary['total_ssf'] or 0),
            },
            'leave_balances': leave_balances,
            'recent_payslips': recent_payslips,
            'expenses': expenses if can_expense else [],
            'can_expense': can_expense,
            'can_submit_expense': can_submit_expense,
        })


class MyAttendanceView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        records = AttendanceRecord.objects.filter(employee_id=employee.id)
        date_from = request.GET.get('date_from')
        if date_from:
            records = records.filter(date__gte=date_from)
        date_to = request.GET.get('date_to')
        if date_to:
            records = records.filter(date__lte=date_to)

        month = request.GET.get('month')
        year = request.GET.get('year')
        if month and year:
            records = records.filter(date__month=month, date__year=year)

        online_perm = OnlineCheckInPermission.objects.filter(employee_id=employee.id, is_allowed=True).first()
        online_allowed = bool(online_perm)
        require_photo = online_perm.require_photo if online_perm else True
        require_location = online_perm.require_location if online_perm else True

        today_record = AttendanceRecord.objects.filter(employee_id=employee.id, date=date.today()).first()
        today_status = None
        if today_record:
            today_status = serialize_attendance(today_record)

        summary = {}
        filtered = records
        present_statuses = ['present', 'half_day']
        absent_statuses = ['absent']
        leave_statuses = ['on_leave', 'paid_leave', 'unpaid_leave', 'sick_leave', 'casual_leave']
        summary['present'] = filtered.filter(status__in=present_statuses).count()
        summary['absent'] = filtered.filter(status__in=absent_statuses).count()
        summary['late'] = filtered.filter(is_late=True).count()
        summary['on_leave'] = filtered.filter(status__in=leave_statuses).count()

        return Response({
            'records': [serialize_attendance(a) for a in records[:100]],
            'summary': summary,
            'online_checkin_allowed': online_allowed,
            'require_photo': require_photo,
            'require_location': require_location,
            'today': today_status,
        })


class MyLeaveBalanceView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        current_fy = FiscalYear.objects.filter(is_current=True).first()

        result = []
        seen_lt_ids = set()

        if current_fy:
            balances = LeaveBalance.objects.select_related('leave_type', 'fiscal_year').filter(
                employee_id=employee.id, fiscal_year=current_fy
            )
            for lb in balances:
                result.append(serialize_leave_balance(lb))
                seen_lt_ids.add(lb.leave_type_id)

        if employee.organization_id:
            leave_types = LeaveType.objects.filter(
                organization_id=employee.organization_id, status='active'
            )
            for lt in leave_types:
                if lt.id not in seen_lt_ids:
                    result.append({
                        'id': str(lt.id),
                        'employee_id': str(employee.id),
                        'leave_type_id': str(lt.id),
                        'leave_type_name': lt.name,
                        'leave_type_code': lt.code,
                        'min_advance_days': lt.min_advance_days,
                        'fiscal_year_id': str(current_fy.id) if current_fy else None,
                        'fiscal_year_name': current_fy.name if current_fy else None,
                        'allocated_days': float(lt.default_days),
                        'used_days': 0,
                        'carried_forward_days': 0,
                        'remaining_days': float(lt.default_days),
                    })

        return Response(result)


class MyLeavePolicyView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)
        if not employee.organization_id:
            return Response({})
        policy = LeavePolicy.objects.filter(organization_id=employee.organization_id).first()
        if not policy:
            return Response({})

        colleagues = Employee.objects.filter(
            organization_id=employee.organization_id,
            status='active',
        ).exclude(id=employee.id).values('id', 'full_name', 'department_id')

        return Response({
            'advance_notice_rules': policy.advance_notice_rules or [],
            'min_days_advance_notice': policy.min_days_advance_notice,
            'max_consecutive_days': policy.max_consecutive_days,
            'require_document_after_days': policy.require_document_after_days,
            'require_cover_person': policy.require_cover_person,
            'require_cover_after_days': policy.require_cover_after_days,
            'allow_half_day': policy.allow_half_day,
            'colleagues': [{'id': str(c['id']), 'full_name': c['full_name']} for c in colleagues],
        })


class MyLeaveRequestsView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        reqs = LeaveRequest.objects.select_related('leave_type').filter(employee_id=employee.id)
        status_filter = request.GET.get('status')
        if status_filter:
            reqs = reqs.filter(status=status_filter)
        return Response([serialize_leave_request(lr) for lr in reqs[:100]])

    @require_auth
    def post(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        data = request.data
        data['employee_id'] = str(employee.id)

        leave_type_id = data.get('leave_type_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        is_half_day = data.get('is_half_day', False)

        if not all([leave_type_id, start_date, end_date]):
            return Response({'message': 'leave_type_id, start_date, end_date are required'}, status=400)

        lt_filter = {'id': leave_type_id, 'status': 'active'}
        if employee.organization_id:
            lt_filter['organization_id'] = employee.organization_id
        try:
            lt_obj = LeaveType.objects.get(**lt_filter)
        except LeaveType.DoesNotExist:
            return Response({'message': 'Leave type not found or not available for your organization'}, status=404)

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        if end < start:
            return Response({'message': 'End date cannot be before start date'}, status=400)

        if lt_obj.min_advance_days > 0:
            days_ahead_lt = (start - date.today()).days
            if days_ahead_lt < lt_obj.min_advance_days:
                return Response({'message': f'{lt_obj.name} requires at least {lt_obj.min_advance_days} days advance notice'}, status=400)

        days_count = Decimal('0.5') if is_half_day else Decimal(str((end - start).days + 1))

        if employee.organization_id:
            policy = LeavePolicy.objects.filter(organization_id=employee.organization_id).first()
            if policy:
                if policy.weekend_days:
                    business_days = Decimal('0')
                    current = start
                    while current <= end:
                        if current.weekday() not in policy.weekend_days:
                            business_days += 1
                        current += timedelta(days=1)
                    if not is_half_day:
                        days_count = business_days

                if days_count > policy.max_consecutive_days:
                    return Response({'message': f'Cannot request more than {policy.max_consecutive_days} consecutive days'}, status=400)

                days_ahead = (start - date.today()).days
                advance_notice_rules = policy.advance_notice_rules or []
                if advance_notice_rules:
                    sorted_rules = sorted(advance_notice_rules, key=lambda r: r.get('min_leave_days', 0), reverse=True)
                    for rule in sorted_rules:
                        if float(days_count) >= rule.get('min_leave_days', 0):
                            required_notice = rule.get('advance_notice_days', 0)
                            if days_ahead < required_notice:
                                return Response({
                                    'message': f'Leave of {int(days_count)}+ days requires at least {required_notice} days advance notice (you gave {days_ahead} days)'
                                }, status=400)
                            break
                elif policy.min_days_advance_notice > 0:
                    if days_ahead < policy.min_days_advance_notice:
                        return Response({'message': f'Leave must be requested at least {policy.min_days_advance_notice} days in advance'}, status=400)

                if not policy.allow_half_day and is_half_day:
                    return Response({'message': 'Half-day leave is not allowed'}, status=400)

                if policy.require_cover_person and float(days_count) >= policy.require_cover_after_days:
                    if not data.get('cover_person_id'):
                        return Response({'message': f'A cover person is required for leave of {policy.require_cover_after_days}+ days'}, status=400)

                if not policy.allow_negative_balance:
                    current_fy = FiscalYear.objects.filter(is_current=True).first()
                    if current_fy:
                        balance = LeaveBalance.objects.filter(
                            employee_id=employee.id,
                            leave_type_id=leave_type_id,
                            fiscal_year=current_fy,
                        ).first()
                        if balance and balance.remaining_days < days_count:
                            return Response({'message': f'Insufficient leave balance. Available: {float(balance.remaining_days)} days'}, status=400)

        lr = LeaveRequest.objects.create(
            employee_id=employee.id,
            leave_type_id=leave_type_id,
            start_date=start,
            end_date=end,
            days_count=days_count,
            is_half_day=is_half_day,
            half_day_period=data.get('half_day_period'),
            reason=data.get('reason'),
            document_url=data.get('document_url'),
            cover_person_id=data.get('cover_person_id') or None,
        )
        return Response(serialize_leave_request(lr), status=201)


class MyPayslipsView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        payslips = Payslip.objects.filter(employee_id=employee.id, status='paid')
        year = request.GET.get('year')
        if year:
            payslips = payslips.filter(year=year)
        return Response([serialize_payslip(ps) for ps in payslips])


class MyPayslipPDFView(APIView):
    @require_auth
    def get(self, request, ps_id):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)
        try:
            ps = Payslip.objects.get(id=ps_id, employee_id=employee.id, status='paid')
        except Payslip.DoesNotExist:
            return Response({'message': 'Payslip not found'}, status=404)
        pr = ps.payroll_run
        org = pr.organization if pr else None
        pdf_bytes = generate_payslip_pdf(ps, employee, org)
        from django.http import HttpResponse
        month_name = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][ps.month - 1]
        filename = f"Payslip_{employee.full_name.replace(' ', '_')}_{month_name}_{ps.year}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BonusListView(APIView):
    @require_permission('hrms.bonus.read')
    def get(self, request):
        qs = Bonus.objects.all()
        emp = request.GET.get('employee_id')
        if emp:
            qs = qs.filter(employee_id=emp)
        month = request.GET.get('month')
        year = request.GET.get('year')
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return Response([serialize_bonus(b) for b in qs])

    @require_permission('hrms.bonus.add')
    def post(self, request):
        data = request.data
        try:
            emp = Employee.objects.get(id=data.get('employee_id'))
        except Employee.DoesNotExist:
            return Response({'message': 'Employee not found'}, status=404)
        b = Bonus.objects.create(
            employee_id=emp.id,
            bonus_type=data.get('bonus_type', 'other'),
            amount=Decimal(str(data.get('amount', 0))),
            reason=data.get('reason'),
            month=int(data.get('month')),
            year=int(data.get('year')),
            is_taxable=data.get('is_taxable', True),
            status=data.get('status', 'pending'),
        )
        return Response(serialize_bonus(b), status=201)


class BonusDetailView(APIView):
    @require_permission('hrms.bonus.update')
    def patch(self, request, bonus_id):
        try:
            b = Bonus.objects.get(id=bonus_id)
        except Bonus.DoesNotExist:
            return Response({'message': 'Bonus not found'}, status=404)
        data = request.data
        for field in ['bonus_type', 'reason', 'is_taxable', 'status']:
            if field in data:
                setattr(b, field, data[field])
        if 'amount' in data:
            b.amount = Decimal(str(data['amount']))
        if 'month' in data:
            b.month = int(data['month'])
        if 'year' in data:
            b.year = int(data['year'])
        if data.get('status') == 'approved':
            b.approved_by = request.session.get('userId')
            b.approved_at = timezone.now()
        b.save()
        return Response(serialize_bonus(b))

    @require_permission('hrms.bonus.delete')
    def delete(self, request, bonus_id):
        try:
            b = Bonus.objects.get(id=bonus_id)
        except Bonus.DoesNotExist:
            return Response({'message': 'Bonus not found'}, status=404)
        b.delete()
        return Response({'message': 'Bonus deleted'})


class BonusBulkUploadView(APIView):
    @require_auth
    def get(self, request):
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'employee_email', 'bonus_type', 'amount', 'reason',
            'month', 'year', 'is_taxable', 'status',
        ])
        writer.writerow([
            'john@example.com', 'dashain', '10000', 'Dashain festival bonus',
            '10', '2026', 'yes', 'pending',
        ])
        writer.writerow([
            'jane@example.com', 'performance', '5000', 'Q3 performance',
            '10', '2026', 'yes', 'approved',
        ])
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bonus_bulk_template.csv"'
        return response

    @require_permission('hrms.bonus.add')
    def post(self, request):
        import csv
        import io

        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No file uploaded'}, status=400)

        fname = file.name.lower()
        rows = []

        if fname.endswith('.csv'):
            decoded = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                rows.append({k.strip().lower().replace(' ', '_'): (v.strip() if v else '') for k, v in row.items() if k})
        elif fname.endswith('.xlsx'):
            import openpyxl
            try:
                wb = openpyxl.load_workbook(file, read_only=True)
            except Exception:
                return Response({'message': 'Invalid Excel file.'}, status=400)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or '').strip().lower().replace(' ', '_') for c in row]
                    continue
                if not any(row):
                    continue
                rd = {}
                for j, val in enumerate(row):
                    if j < len(headers) and headers[j]:
                        rd[headers[j]] = str(val).strip() if val is not None else ''
                rows.append(rd)
        else:
            return Response({'message': 'Unsupported file format. Use CSV or Excel (.xlsx).'}, status=400)

        if not rows:
            return Response({'message': 'No data rows found in the file.'}, status=400)

        emp_map = {}
        for emp in Employee.objects.filter(status='active'):
            emp_map[emp.email.lower()] = emp

        valid_types = {c[0] for c in BONUS_TYPE_CHOICES}

        created = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            email = (row.get('employee_email') or row.get('email') or '').lower()
            if not email:
                errors.append(f'Row {i}: Missing employee_email')
                continue

            emp = emp_map.get(email)
            if not emp:
                errors.append(f'Row {i}: Employee not found for email {email}')
                continue

            amount_str = row.get('amount') or '0'
            try:
                amount = Decimal(str(amount_str))
            except Exception:
                errors.append(f'Row {i}: Invalid amount "{amount_str}"')
                continue

            if amount <= 0:
                errors.append(f'Row {i}: Amount must be positive')
                continue

            month_str = row.get('month') or ''
            year_str = row.get('year') or ''
            try:
                month = int(month_str)
                year = int(float(year_str))
            except (ValueError, TypeError):
                errors.append(f'Row {i}: Invalid month/year')
                continue

            if month < 1 or month > 12:
                errors.append(f'Row {i}: Month must be 1-12')
                continue

            bonus_type = (row.get('bonus_type') or 'other').lower()
            if bonus_type not in valid_types:
                bonus_type = 'other'

            is_taxable_str = (row.get('is_taxable') or 'yes').lower()
            is_taxable = is_taxable_str in ('yes', 'true', '1', 'y')

            status_val = (row.get('status') or 'pending').lower()
            if status_val not in ('pending', 'approved', 'paid', 'cancelled'):
                status_val = 'pending'

            reason = row.get('reason') or ''

            Bonus.objects.create(
                employee_id=emp.id,
                bonus_type=bonus_type,
                amount=amount,
                reason=reason if reason else None,
                month=month,
                year=year,
                is_taxable=is_taxable,
                status=status_val,
                approved_by=request.session.get('userId') if status_val == 'approved' else None,
                approved_at=timezone.now() if status_val == 'approved' else None,
            )
            created += 1

        return Response({
            'created': created,
            'errors': errors,
            'message': f'{created} bonus(es) created' + (f', {len(errors)} error(s)' if errors else ''),
        })


class TravelExpenseListView(APIView):
    @require_permission('hrms.expense.read')
    def get(self, request):
        qs = TravelExpense.objects.all()
        emp = request.GET.get('employee_id')
        if emp:
            qs = qs.filter(employee_id=emp)
        month = request.GET.get('month')
        year = request.GET.get('year')
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return Response([serialize_travel_expense(te) for te in qs])

    @require_permission('hrms.expense.add')
    def post(self, request):
        data = request.data
        try:
            emp = Employee.objects.get(id=data.get('employee_id'))
        except Employee.DoesNotExist:
            return Response({'message': 'Employee not found'}, status=404)
        exp_date = date.fromisoformat(data['expense_date']) if data.get('expense_date') else date.today()
        te = TravelExpense.objects.create(
            employee_id=emp.id,
            category=data.get('category', 'travel'),
            description=data.get('description', ''),
            amount=Decimal(str(data.get('amount', 0))),
            expense_date=exp_date,
            receipt_url=data.get('receipt_url'),
            month=int(data.get('month', exp_date.month)),
            year=int(data.get('year', exp_date.year)),
            include_in_salary=data.get('include_in_salary', True),
            status=data.get('status', 'pending'),
        )
        return Response(serialize_travel_expense(te), status=201)


class TravelExpenseDetailView(APIView):
    @require_permission('hrms.expense.update')
    def patch(self, request, expense_id):
        try:
            te = TravelExpense.objects.get(id=expense_id)
        except TravelExpense.DoesNotExist:
            return Response({'message': 'Travel expense not found'}, status=404)
        data = request.data
        for field in ['category', 'description', 'receipt_url', 'include_in_salary', 'status', 'rejection_reason']:
            if field in data:
                setattr(te, field, data[field])
        if 'amount' in data:
            te.amount = Decimal(str(data['amount']))
        if 'expense_date' in data:
            te.expense_date = date.fromisoformat(data['expense_date'])
        if 'month' in data:
            te.month = int(data['month'])
        if 'year' in data:
            te.year = int(data['year'])
        if data.get('status') == 'approved':
            te.approved_by = request.session.get('userId')
            te.approved_at = timezone.now()
        te.save()
        return Response(serialize_travel_expense(te))

    @require_permission('hrms.expense.delete')
    def delete(self, request, expense_id):
        try:
            te = TravelExpense.objects.get(id=expense_id)
        except TravelExpense.DoesNotExist:
            return Response({'message': 'Travel expense not found'}, status=404)
        te.delete()
        return Response({'message': 'Travel expense deleted'})


class AdvancePaymentListView(APIView):
    @require_permission('hrms.advance.read')
    def get(self, request):
        qs = AdvancePayment.objects.all()
        emp = request.GET.get('employee_id')
        if emp:
            qs = qs.filter(employee_id=emp)
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return Response([serialize_advance_payment(ap) for ap in qs])

    @require_permission('hrms.advance.add')
    def post(self, request):
        data = request.data
        try:
            emp = Employee.objects.get(id=data.get('employee_id'))
        except Employee.DoesNotExist:
            return Response({'message': 'Employee not found'}, status=404)
        amount = Decimal(str(data.get('amount', 0)))
        monthly_ded = Decimal(str(data.get('monthly_deduction', 0)))
        if monthly_ded <= 0:
            return Response({'message': 'Monthly deduction must be greater than 0'}, status=400)
        ap = AdvancePayment.objects.create(
            employee_id=emp.id,
            amount=amount,
            reason=data.get('reason'),
            request_date=date.fromisoformat(data['request_date']) if data.get('request_date') else date.today(),
            monthly_deduction=monthly_ded,
            deduction_start_month=int(data.get('deduction_start_month', date.today().month)),
            deduction_start_year=int(data.get('deduction_start_year', date.today().year)),
            remaining_balance=amount,
            status=data.get('status', 'pending'),
        )
        return Response(serialize_advance_payment(ap), status=201)


class AdvancePaymentDetailView(APIView):
    @require_permission('hrms.advance.update')
    def patch(self, request, advance_id):
        try:
            ap = AdvancePayment.objects.get(id=advance_id)
        except AdvancePayment.DoesNotExist:
            return Response({'message': 'Advance payment not found'}, status=404)
        data = request.data
        for field in ['reason', 'status']:
            if field in data:
                setattr(ap, field, data[field])
        if 'monthly_deduction' in data:
            ap.monthly_deduction = Decimal(str(data['monthly_deduction']))
        if data.get('status') == 'approved':
            ap.approved_by = request.session.get('userId')
            ap.approved_at = timezone.now()
            ap.status = 'approved'
        if data.get('status') == 'active':
            ap.status = 'active'
        ap.save()
        return Response(serialize_advance_payment(ap))

    @require_permission('hrms.advance.delete')
    def delete(self, request, advance_id):
        try:
            ap = AdvancePayment.objects.get(id=advance_id)
        except AdvancePayment.DoesNotExist:
            return Response({'message': 'Advance payment not found'}, status=404)
        ap.delete()
        return Response({'message': 'Advance payment deleted'})


class EmployeePortalAccessView(APIView):
    @require_permission('hrms.staff.write')
    def get(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        from accounts.models import User as AccountUser, Role, UserRole
        has_access = False
        user_info = None
        if emp.user_id:
            try:
                user = AccountUser.objects.get(id=emp.user_id)
                has_access = True
                user_roles = UserRole.objects.filter(user_id=user.id)
                roles = []
                for ur in user_roles:
                    try:
                        r = Role.objects.get(id=ur.role_id)
                        roles.append({'id': r.id, 'name': r.name})
                    except Role.DoesNotExist:
                        pass
                user_info = {
                    'user_id': user.id,
                    'email': user.email,
                    'portal_access': user.portal_access,
                    'is_active': user.is_active,
                    'force_password_change': user.force_password_change,
                    'last_login_at': user.last_login_at.isoformat() if user.last_login_at else None,
                    'roles': roles,
                }
            except AccountUser.DoesNotExist:
                emp.user_id = None
                emp.save(update_fields=['user_id'])

        all_roles = list(Role.objects.all().values('id', 'name'))

        return Response({
            'employee_id': str(emp.id),
            'employee_name': emp.full_name,
            'employee_email': emp.email,
            'has_access': has_access,
            'user_info': user_info,
            'available_roles': all_roles,
        })

    @require_permission('hrms.staff.write')
    def post(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        from accounts.models import User as AccountUser, UserRole
        import bcrypt
        import re

        if emp.user_id:
            try:
                existing = AccountUser.objects.get(id=emp.user_id)
                return Response({'error': f'Employee already has portal access (email: {existing.email})'}, status=400)
            except AccountUser.DoesNotExist:
                emp.user_id = None
                emp.save(update_fields=['user_id'])

        existing_user = AccountUser.objects.filter(email=emp.email).first()
        if existing_user:
            emp.user_id = existing_user.id
            emp.save(update_fields=['user_id'])
            portal_access = request.data.get('portal_access', 'employee')
            if portal_access in ('employee', 'both', 'admin'):
                existing_user.portal_access = portal_access
                existing_user.save(update_fields=['portal_access'])
            role_id = request.data.get('role_id')
            if role_id:
                if not UserRole.objects.filter(user_id=existing_user.id, role_id=role_id).exists():
                    UserRole.objects.create(user_id=existing_user.id, role_id=role_id)
            return Response({
                'message': f'Linked existing user account for {emp.email}',
                'user_id': existing_user.id,
            })

        password = request.data.get('password', '')
        if not password or len(password) < 8:
            return Response({'error': 'Password must be at least 8 characters'}, status=400)

        portal_access = request.data.get('portal_access', 'employee')
        if portal_access not in ('admin', 'employee', 'both'):
            portal_access = 'employee'

        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
        user = AccountUser.objects.create(
            email=emp.email,
            full_name=emp.full_name,
            password_hash=hashed,
            is_active=True,
            force_password_change=True,
            portal_access=portal_access,
        )

        emp.user_id = user.id
        emp.save(update_fields=['user_id', 'updated_at'])

        role_id = request.data.get('role_id')
        if role_id:
            UserRole.objects.create(user_id=user.id, role_id=role_id)

        try:
            from config.settings import DEFAULT_FROM_EMAIL
            from accounts.views import send_styled_email
            portal_url = 'https://people.studyinfocentre.com'
            send_styled_email(
                to_email=emp.email,
                subject='Your People Portal Access Has Been Created',
                body=f'''
                <p style="color:#475569;font-size:15px;line-height:1.6;">Hello <strong>{emp.full_name}</strong>,</p>
                <p style="color:#475569;font-size:15px;line-height:1.6;">Your portal access has been created. You can now log in using:</p>
                <table style="width:100%;border-collapse:collapse;margin:16px 0;">
                <tr><td style="padding:8px 12px;border:1px solid #e2e8f0;font-weight:600;width:140px;background:#f8fafc;">Email</td>
                <td style="padding:8px 12px;border:1px solid #e2e8f0;">{emp.email}</td></tr>
                <tr><td style="padding:8px 12px;border:1px solid #e2e8f0;font-weight:600;background:#f8fafc;">Temporary Password</td>
                <td style="padding:8px 12px;border:1px solid #e2e8f0;">{password}</td></tr>
                </table>
                <p style="color:#475569;font-size:15px;line-height:1.6;">Please change your password after first login.</p>
                <p style="text-align:center;margin:24px 0;">
                <a href="{portal_url}" style="display:inline-block;padding:12px 32px;background:#3b82f6;color:white;text-decoration:none;border-radius:8px;font-weight:600;">Login to Portal</a>
                </p>
                ''',
            )
        except Exception:
            pass

        return Response({
            'message': f'Portal access created for {emp.full_name}',
            'user_id': user.id,
        }, status=201)

    @require_permission('hrms.staff.write')
    def patch(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        from accounts.models import User as AccountUser, UserRole
        if not emp.user_id:
            return Response({'error': 'Employee does not have portal access'}, status=400)

        try:
            user = AccountUser.objects.get(id=emp.user_id)
        except AccountUser.DoesNotExist:
            return Response({'error': 'Linked user account not found'}, status=404)

        portal_access = request.data.get('portal_access')
        if portal_access and portal_access in ('admin', 'employee', 'both'):
            user.portal_access = portal_access
            user.save(update_fields=['portal_access'])

        is_active = request.data.get('is_active')
        if is_active is not None:
            user.is_active = bool(is_active)
            user.save(update_fields=['is_active'])

        role_id = request.data.get('role_id')
        if role_id is not None:
            UserRole.objects.filter(user_id=user.id).delete()
            if role_id:
                UserRole.objects.create(user_id=user.id, role_id=role_id)

        reset_password = request.data.get('reset_password')
        if reset_password:
            import bcrypt
            if len(reset_password) < 8:
                return Response({'error': 'Password must be at least 8 characters'}, status=400)
            hashed = bcrypt.hashpw(reset_password.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
            user.password_hash = hashed
            user.force_password_change = True
            user.save(update_fields=['password_hash', 'force_password_change'])

        return Response({'message': 'Portal access updated'})

    @require_permission('hrms.staff.write')
    def delete(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        from accounts.models import User as AccountUser
        if not emp.user_id:
            return Response({'error': 'Employee does not have portal access'}, status=400)

        try:
            user = AccountUser.objects.get(id=emp.user_id)
            user.is_active = False
            user.save(update_fields=['is_active'])
        except AccountUser.DoesNotExist:
            pass

        emp.user_id = None
        emp.save(update_fields=['user_id', 'updated_at'])

        return Response({'message': 'Portal access revoked'})


class EmployeeStatusChangeView(APIView):
    @require_permission('hrms.staff.write')
    def patch(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        new_status = request.data.get('status')
        valid_statuses = ['active', 'inactive', 'terminated', 'resigned', 'on_notice']
        if new_status not in valid_statuses:
            return Response({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}, status=400)

        emp.status = new_status
        emp.save(update_fields=['status', 'updated_at'])

        return Response({
            'id': str(emp.id),
            'full_name': emp.full_name,
            'status': emp.status,
            'message': f'Employee status changed to {new_status}',
        })


class StaffProfileListView(APIView):
    @require_permission('hrms.staff.read')
    def get(self, request):
        org = request.GET.get('organization_id')
        dept = request.GET.get('department_id')
        status_filter = request.GET.get('status', 'active')
        if status_filter == 'all':
            qs = Employee.objects.all()
        else:
            qs = Employee.objects.filter(status=status_filter)
        if org:
            qs = qs.filter(organization_id=org)
        if dept:
            qs = qs.filter(department_id=dept)
        result = []
        for emp in qs:
            sal = SalaryStructure.objects.filter(
                employee_id=emp.id, status='active'
            ).order_by('-effective_from').first()
            org_name = None
            dept_name = None
            if emp.organization_id:
                try:
                    org_obj = Organization.objects.get(id=emp.organization_id)
                    org_name = org_obj.name
                except Organization.DoesNotExist:
                    pass
            if emp.department_id:
                try:
                    dept_obj = Department.objects.get(id=emp.department_id)
                    dept_name = dept_obj.name
                except Department.DoesNotExist:
                    pass
            active_advances = AdvancePayment.objects.filter(
                employee_id=emp.id, status__in=['approved', 'active']
            ).aggregate(total=Sum('remaining_balance'))['total'] or 0

            result.append({
                'id': str(emp.id),
                'full_name': emp.full_name,
                'email': emp.email,
                'phone': emp.phone,
                'position': emp.position,
                'department': emp.department,
                'organization_id': str(emp.organization_id) if emp.organization_id else None,
                'organization_name': org_name,
                'department_id': str(emp.department_id) if emp.department_id else None,
                'department_name': dept_name,
                'gender': emp.gender,
                'country': emp.country,
                'marital_status': emp.marital_status,
                'date_of_birth': emp.date_of_birth.isoformat() if emp.date_of_birth else None,
                'join_date': emp.join_date.isoformat() if emp.join_date else None,
                'employment_type': emp.employment_type,
                'bank_name': emp.bank_name,
                'bank_account_number': emp.bank_account_number,
                'bank_branch': emp.bank_branch,
                'citizenship_no': emp.citizenship_no,
                'pan_no': emp.pan_no,
                'passport_number': emp.passport_number,
                'permanent_address': emp.permanent_address,
                'temporary_address': emp.temporary_address,
                'emergency_contact_name': emp.emergency_contact_name,
                'emergency_contact_phone': emp.emergency_contact_phone,
                'employee_id_number': emp.employee_id_number,
                'probation_end_date': emp.probation_end_date.isoformat() if emp.probation_end_date else None,
                'contract_end_date': emp.contract_end_date.isoformat() if emp.contract_end_date else None,
                'working_days_per_week': emp.working_days_per_week,
                'salary_amount': float(emp.salary_amount) if emp.salary_amount else None,
                'salary_currency': emp.salary_currency,
                'profile_photo_url': emp.profile_photo_url,
                'status': emp.status,
                'has_portal_access': bool(emp.user_id),
                'user_id': emp.user_id,
                'salary_structure': {
                    'id': str(sal.id),
                    'basic_salary': float(sal.basic_salary),
                    'allowances': sal.allowances or {},
                    'deductions': sal.deductions or {},
                    'cit_type': sal.cit_type,
                    'cit_value': float(sal.cit_value),
                    'ssf_applicable': sal.ssf_applicable,
                    'ssf_employee_percentage': float(sal.ssf_employee_percentage),
                    'ssf_employer_percentage': float(sal.ssf_employer_percentage),
                    'tax_applicable': sal.tax_applicable,
                    'effective_from': sal.effective_from.isoformat() if sal.effective_from else None,
                } if sal else None,
                'outstanding_advance': float(active_advances),
            })
        return Response(result)


class Employee360View(APIView):
    @require_permission('hrms.staff.read')
    def get(self, request, employee_id):
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        user_perms = set(request.session.get('userPermissions', []))
        can_salary = 'hrms.salary.read' in user_perms
        can_payroll = 'hrms.payroll.read' in user_perms or 'hrms.payslip.read' in user_perms
        can_bonus = 'hrms.bonus.read' in user_perms
        can_advance = 'hrms.advance.read' in user_perms
        can_expense = 'hrms.expense.read' in user_perms
        can_tax = 'hrms.tax.read' in user_perms

        confidential_unlocked = False
        unlock_ts = request.session.get('confidential_unlocked_at')
        if unlock_ts:
            try:
                unlock_time = datetime.fromisoformat(unlock_ts)
                if (timezone.now() - unlock_time).total_seconds() < 300:
                    confidential_unlocked = True
            except (ValueError, TypeError):
                pass

        org_name = None
        dept_name = None
        org_reg_label = 'Registration No.'
        org_pan_label = 'PAN No.'
        if emp.organization_id:
            try:
                org_obj = Organization.objects.get(id=emp.organization_id)
                org_name = org_obj.name
                org_reg_label = org_obj.registration_label or 'Registration No.'
                org_pan_label = org_obj.pan_label or 'PAN No.'
            except Organization.DoesNotExist:
                pass
        if emp.department_id:
            try:
                dept_name = Department.objects.get(id=emp.department_id).name
            except Department.DoesNotExist:
                pass

        salary_data = None
        if can_salary and confidential_unlocked:
            sal = SalaryStructure.objects.filter(
                employee_id=emp.id, status='active'
            ).order_by('-effective_from').first()
            if sal:
                salary_data = {
                    'id': str(sal.id),
                    'basic_salary': float(sal.basic_salary),
                    'allowances': sal.allowances or {},
                    'deductions': sal.deductions or {},
                    'cit_type': sal.cit_type,
                    'cit_value': float(sal.cit_value),
                    'ssf_applicable': sal.ssf_applicable,
                    'ssf_employee_percentage': float(sal.ssf_employee_percentage),
                    'ssf_employer_percentage': float(sal.ssf_employer_percentage),
                    'tax_applicable': sal.tax_applicable,
                    'effective_from': sal.effective_from.isoformat() if sal.effective_from else None,
                }

        now = timezone.now()
        current_month = now.month
        current_year = now.year
        att_records = AttendanceRecord.objects.filter(
            employee_id=emp.id, date__month=current_month, date__year=current_year
        )
        att_summary = {
            'month': current_month,
            'year': current_year,
            'present': att_records.filter(status='present').count(),
            'absent': att_records.filter(status='absent').count(),
            'late': att_records.filter(status='late').count(),
            'half_day': att_records.filter(status='half_day').count(),
            'on_leave': att_records.filter(status__in=['on_leave', 'paid_leave', 'unpaid_leave', 'sick_leave', 'casual_leave']).count(),
            'total_records': att_records.count(),
        }

        leave_balances = []
        for lb in LeaveBalance.objects.filter(employee_id=emp.id).select_related('leave_type'):
            leave_balances.append({
                'leave_type': lb.leave_type.name,
                'leave_type_code': lb.leave_type.code,
                'color': lb.leave_type.color,
                'allocated': float(lb.allocated_days),
                'used': float(lb.used_days),
                'carried_forward': float(lb.carried_forward_days),
                'remaining': float(lb.allocated_days + lb.carried_forward_days - lb.used_days),
            })

        recent_payslips = []
        if can_payroll:
            for ps in Payslip.objects.filter(employee_id=emp.id).order_by('-year', '-month')[:6]:
                recent_payslips.append({
                    'id': str(ps.id),
                    'month': ps.month,
                    'year': ps.year,
                    'basic_salary': float(ps.basic_salary),
                    'gross_salary': float(ps.gross_salary),
                    'cit_deduction': float(ps.cit_deduction),
                    'ssf_employee_deduction': float(ps.ssf_employee_deduction),
                    'ssf_employer_contribution': float(ps.ssf_employer_contribution),
                    'tax_deduction': float(ps.tax_deduction),
                    'bonus_amount': float(ps.bonus_amount),
                    'travel_reimbursement': float(ps.travel_reimbursement),
                    'advance_deduction': float(ps.advance_deduction),
                    'unpaid_leave_deduction': float(ps.unpaid_leave_deduction),
                    'total_deductions': float(ps.total_deductions),
                    'net_salary': float(ps.net_salary),
                    'working_days': ps.working_days,
                    'present_days': ps.present_days,
                    'status': ps.status,
                    'view_token': ps.view_token,
                })

        bonuses = []
        if can_bonus:
            for b in Bonus.objects.filter(employee_id=emp.id).order_by('-year', '-month')[:10]:
                bonuses.append({
                    'id': str(b.id),
                    'bonus_type': b.bonus_type,
                    'amount': float(b.amount),
                    'reason': b.reason,
                    'month': b.month,
                    'year': b.year,
                    'is_taxable': b.is_taxable,
                    'status': b.status,
                })

        advances = []
        active_advances_total = 0
        if can_advance and confidential_unlocked:
            for a in AdvancePayment.objects.filter(employee_id=emp.id).order_by('-request_date')[:10]:
                advances.append({
                    'id': str(a.id),
                    'amount': float(a.amount),
                    'reason': a.reason,
                    'request_date': a.request_date.isoformat() if a.request_date else None,
                    'monthly_deduction': float(a.monthly_deduction),
                    'total_deducted': float(a.total_deducted),
                    'remaining_balance': float(a.remaining_balance),
                    'status': a.status,
                })
            active_advances_total = AdvancePayment.objects.filter(
                employee_id=emp.id, status__in=['approved', 'active']
            ).aggregate(total=Sum('remaining_balance'))['total'] or 0

        expenses = []
        if can_expense:
            for e in TravelExpense.objects.filter(employee_id=emp.id).order_by('-expense_date')[:10]:
                expenses.append({
                    'id': str(e.id),
                    'category': e.category,
                    'description': e.description,
                    'amount': float(e.amount),
                    'expense_date': e.expense_date.isoformat() if e.expense_date else None,
                    'status': e.status,
                    'include_in_salary': e.include_in_salary,
                })

        tax_summary = {'year': current_year, 'total_tax': 0, 'total_cit': 0, 'total_ssf': 0}
        if can_tax:
            annual_tax_paid = Payslip.objects.filter(
                employee_id=emp.id, year=current_year
            ).aggregate(
                total_tax=Sum('tax_deduction'),
                total_cit=Sum('cit_deduction'),
                total_ssf=Sum('ssf_employee_deduction'),
            )
            tax_summary = {
                'year': current_year,
                'total_tax': float(annual_tax_paid['total_tax'] or 0),
                'total_cit': float(annual_tax_paid['total_cit'] or 0),
                'total_ssf': float(annual_tax_paid['total_ssf'] or 0),
            }

        emp_data = {
            'id': str(emp.id),
            'full_name': emp.full_name,
            'email': emp.email,
            'phone': emp.phone,
            'position': emp.position,
            'department': emp.department,
            'organization_name': org_name,
            'registration_label': org_reg_label,
            'pan_label': org_pan_label,
            'department_name': dept_name,
            'organization_id': str(emp.organization_id) if emp.organization_id else None,
            'department_id': str(emp.department_id) if emp.department_id else None,
            'gender': emp.gender,
            'country': emp.country,
            'marital_status': emp.marital_status,
            'date_of_birth': emp.date_of_birth.isoformat() if emp.date_of_birth else None,
            'join_date': emp.join_date.isoformat() if emp.join_date else None,
            'employment_type': emp.employment_type,
            'citizenship_no': emp.citizenship_no,
            'pan_no': emp.pan_no,
            'passport_number': emp.passport_number,
            'employee_id_number': emp.employee_id_number,
            'permanent_address': emp.permanent_address,
            'temporary_address': emp.temporary_address,
            'salary_currency': emp.salary_currency,
            'profile_photo_url': emp.profile_photo_url,
            'status': emp.status,
            'probation_end_date': emp.probation_end_date.isoformat() if emp.probation_end_date else None,
            'contract_end_date': emp.contract_end_date.isoformat() if emp.contract_end_date else None,
            'working_days_per_week': emp.working_days_per_week,
            'emergency_contact_name': emp.emergency_contact_name,
            'emergency_contact_phone': emp.emergency_contact_phone,
        }
        if can_salary:
            emp_data['bank_name'] = emp.bank_name
            emp_data['bank_account_number'] = emp.bank_account_number
            emp_data['bank_branch'] = emp.bank_branch

        return Response({
            'employee': emp_data,
            'salary_structure': salary_data,
            'attendance_summary': att_summary,
            'leave_balances': leave_balances,
            'recent_payslips': recent_payslips,
            'bonuses': bonuses,
            'advances': advances,
            'expenses': expenses,
            'outstanding_advance': float(active_advances_total),
            'tax_summary': tax_summary,
            'confidential_unlocked': confidential_unlocked,
        })


def serialize_tax_slab(s):
    return {
        'id': str(s.id),
        'organization_id': str(s.organization_id) if s.organization_id else None,
        'fiscal_year_id': str(s.fiscal_year_id) if s.fiscal_year_id else None,
        'country': s.country,
        'marital_status': s.marital_status,
        'slab_order': s.slab_order,
        'lower_limit': float(s.lower_limit),
        'upper_limit': float(s.upper_limit) if s.upper_limit is not None else None,
        'rate': float(s.rate),
        'is_active': s.is_active,
    }


class TaxSlabListView(APIView):
    @require_permission('hrms.tax.read')
    def get(self, request):
        qs = TaxSlab.objects.filter(is_active=True)
        org = request.GET.get('organization_id')
        if org:
            qs = qs.filter(Q(organization_id=org) | Q(organization__isnull=True))
        country = request.GET.get('country')
        if country:
            qs = qs.filter(Q(country=country) | Q(country__isnull=True))
        ms = request.GET.get('marital_status')
        if ms:
            qs = qs.filter(marital_status=ms)
        return Response([serialize_tax_slab(s) for s in qs])

    @require_permission('hrms.tax.update')
    def post(self, request):
        data = request.data
        s = TaxSlab.objects.create(
            organization_id=data.get('organization_id') or None,
            fiscal_year_id=data.get('fiscal_year_id') or None,
            country=data.get('country') or None,
            marital_status=data.get('marital_status', 'single'),
            slab_order=int(data.get('slab_order', 1)),
            lower_limit=Decimal(str(data.get('lower_limit', 0))),
            upper_limit=Decimal(str(data['upper_limit'])) if data.get('upper_limit') is not None else None,
            rate=Decimal(str(data.get('rate', 0))),
            is_active=data.get('is_active', True),
        )
        return Response(serialize_tax_slab(s), status=201)


class TaxSlabDetailView(APIView):
    @require_permission('hrms.tax.update')
    def patch(self, request, slab_id):
        try:
            s = TaxSlab.objects.get(id=slab_id)
        except TaxSlab.DoesNotExist:
            return Response({'message': 'Tax slab not found'}, status=404)
        data = request.data
        for field in ['marital_status', 'slab_order', 'lower_limit', 'upper_limit', 'rate', 'is_active', 'organization_id', 'fiscal_year_id', 'country']:
            if field in data:
                val = data[field]
                if field in ('lower_limit', 'rate') and val is not None:
                    val = Decimal(str(val))
                if field == 'upper_limit':
                    val = Decimal(str(val)) if val is not None else None
                if field == 'slab_order' and val is not None:
                    val = int(val)
                setattr(s, field, val)
        s.save()
        return Response(serialize_tax_slab(s))

    @require_permission('hrms.tax.update')
    def delete(self, request, slab_id):
        try:
            s = TaxSlab.objects.get(id=slab_id)
        except TaxSlab.DoesNotExist:
            return Response({'message': 'Tax slab not found'}, status=404)
        s.delete()
        return Response({'message': 'Tax slab deleted'})


class TaxSlabBulkSaveView(APIView):
    @require_permission('hrms.tax.update')
    def post(self, request):
        data = request.data
        marital_status = data.get('marital_status', 'single')
        slabs = data.get('slabs', [])
        org_id = data.get('organization_id') or None
        country = data.get('country') or None

        TaxSlab.objects.filter(
            marital_status=marital_status,
            organization_id=org_id,
            country=country,
        ).delete()

        created = []
        for i, slab in enumerate(slabs):
            s = TaxSlab.objects.create(
                organization_id=org_id,
                country=country,
                marital_status=marital_status,
                slab_order=i + 1,
                lower_limit=Decimal(str(slab.get('lower_limit', 0))),
                upper_limit=Decimal(str(slab['upper_limit'])) if slab.get('upper_limit') is not None else None,
                rate=Decimal(str(slab.get('rate', 0))),
                is_active=True,
            )
            created.append(serialize_tax_slab(s))
        return Response(created, status=201)


class GovernmentTaxRecordsView(APIView):
    @require_permission('hrms.payroll.read')
    def get(self, request):
        year = request.GET.get('year', datetime.now().year)
        org_id = request.GET.get('organization_id')

        payslips = Payslip.objects.filter(
            payroll_run__year=int(year),
            payroll_run__status__in=['processed', 'approved', 'paid', 'completed'],
        )
        if org_id:
            payslips = payslips.filter(payroll_run__organization_id=org_id)

        emp_cache = {}
        def get_emp_info(eid):
            if eid not in emp_cache:
                try:
                    emp = Employee.objects.get(id=eid)
                    emp_cache[eid] = {'name': emp.full_name, 'pan': emp.pan_no or ''}
                except Employee.DoesNotExist:
                    emp_cache[eid] = {'name': str(eid), 'pan': ''}
            return emp_cache[eid]

        monthly_records = []
        for month in range(1, 13):
            month_slips = payslips.filter(payroll_run__month=month)
            staff_details = []
            if month_slips.exists():
                for slip in month_slips.select_related('payroll_run'):
                    cit = float(slip.cit_deduction or 0)
                    ssf_emp = float(slip.ssf_employee_deduction or 0)
                    ssf_empr = float(slip.ssf_employer_contribution or 0)
                    tax = float(slip.tax_deduction or 0)
                    emp_info = get_emp_info(slip.employee_id)
                    staff_details.append({
                        'employee_id': str(slip.employee_id),
                        'employee_name': emp_info['name'],
                        'employee_pan': emp_info['pan'],
                        'gross_salary': float(slip.gross_salary or 0),
                        'cit': cit,
                        'ssf_employee': ssf_emp,
                        'ssf_employer': ssf_empr,
                        'tax': tax,
                        'total_govt': cit + ssf_emp + ssf_empr + tax,
                    })
            agg_gross = sum(s['gross_salary'] for s in staff_details)
            agg_cit = sum(s['cit'] for s in staff_details)
            agg_ssf_emp = sum(s['ssf_employee'] for s in staff_details)
            agg_ssf_empr = sum(s['ssf_employer'] for s in staff_details)
            agg_tax = sum(s['tax'] for s in staff_details)
            monthly_records.append({
                'month': month,
                'year': int(year),
                'employee_count': len(staff_details),
                'total_gross': agg_gross,
                'total_cit': agg_cit,
                'total_ssf_employee': agg_ssf_emp,
                'total_ssf_employer': agg_ssf_empr,
                'total_tax': agg_tax,
                'total_payable_to_govt': agg_cit + agg_ssf_emp + agg_ssf_empr + agg_tax,
                'staff': staff_details,
            })

        totals = {
            'total_cit': sum(r['total_cit'] for r in monthly_records),
            'total_ssf_employee': sum(r['total_ssf_employee'] for r in monthly_records),
            'total_ssf_employer': sum(r['total_ssf_employer'] for r in monthly_records),
            'total_tax': sum(r['total_tax'] for r in monthly_records),
            'total_payable_to_govt': sum(r['total_payable_to_govt'] for r in monthly_records),
        }

        return Response({
            'monthly': monthly_records,
            'annual_totals': totals,
            'year': int(year),
        })


class CountryTaxLabelListView(APIView):
    @require_auth
    def get(self, request):
        qs = CountryTaxLabel.objects.filter(is_active=True).order_by('country')
        return Response([{
            'id': str(c.id),
            'country': c.country,
            'tax_id_label': c.tax_id_label,
        } for c in qs])

    @require_permission('hrms.organization.add')
    def post(self, request):
        data = request.data
        country = (data.get('country') or '').strip()
        tax_id_label = (data.get('tax_id_label') or 'Tax ID No.').strip()
        if not country:
            return Response({'message': 'Country name is required'}, status=400)
        if CountryTaxLabel.objects.filter(country__iexact=country).exists():
            return Response({'message': f'Country "{country}" already exists'}, status=400)
        c = CountryTaxLabel.objects.create(country=country, tax_id_label=tax_id_label)
        return Response({'id': str(c.id), 'country': c.country, 'tax_id_label': c.tax_id_label}, status=201)


class CountryTaxLabelDetailView(APIView):
    @require_permission('hrms.organization.add')
    def patch(self, request, label_id):
        try:
            c = CountryTaxLabel.objects.get(id=label_id)
        except CountryTaxLabel.DoesNotExist:
            return Response({'message': 'Not found'}, status=404)
        data = request.data
        if 'country' in data:
            new_country = (data['country'] or '').strip()
            if new_country and new_country != c.country:
                if CountryTaxLabel.objects.filter(country__iexact=new_country).exclude(id=c.id).exists():
                    return Response({'message': f'Country "{new_country}" already exists'}, status=400)
                c.country = new_country
        if 'tax_id_label' in data:
            c.tax_id_label = (data['tax_id_label'] or 'Tax ID No.').strip()
        c.save()
        return Response({'id': str(c.id), 'country': c.country, 'tax_id_label': c.tax_id_label})

    @require_permission('hrms.organization.add')
    def delete(self, request, label_id):
        try:
            c = CountryTaxLabel.objects.get(id=label_id)
        except CountryTaxLabel.DoesNotExist:
            return Response({'message': 'Not found'}, status=404)
        c.delete()
        return Response({'message': 'Deleted'})


class PublicPayslipPDFView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, token):
        try:
            ps = Payslip.objects.get(view_token=token, status__in=['completed', 'paid', 'approved'])
        except Payslip.DoesNotExist:
            return Response({'message': 'Payslip not found or link expired'}, status=404)
        try:
            emp = Employee.objects.get(id=ps.employee_id)
        except Employee.DoesNotExist:
            return Response({'message': 'Employee not found'}, status=404)
        pr = ps.payroll_run
        org = pr.organization if pr else None
        pdf_bytes = generate_payslip_pdf(ps, emp, org)
        from django.http import HttpResponse
        month_name = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][ps.month - 1]
        filename = f"Payslip_{emp.full_name.replace(' ', '_')}_{month_name}_{ps.year}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response


class ConfidentialOTPSendView(APIView):
    @require_auth
    def post(self, request):
        from accounts.models import User, LoginVerificationCode
        from accounts.views import send_otp_email
        user_id = request.session.get('userId')
        if not user_id:
            return Response({'message': 'Not authenticated'}, status=401)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'message': 'User not found'}, status=404)

        otp_code = secrets.randbelow(900000) + 100000
        otp_str = str(otp_code)
        code_hash = hashlib.sha256(otp_str.encode()).hexdigest()
        expires_at = timezone.now() + timedelta(minutes=5)

        LoginVerificationCode.objects.filter(
            user_id=user_id, status='pending'
        ).update(status='expired')

        LoginVerificationCode.objects.create(
            user_id=user_id,
            code_hash=code_hash,
            expires_at=expires_at,
            status='pending',
        )

        try:
            send_otp_email(user.email, otp_str, portal='people')
        except Exception as e:
            print(f'[OTP-Confidential] Email send failed: {e}')
            return Response({'message': 'Failed to send OTP email'}, status=500)

        masked_email = user.email[:2] + '***@' + user.email.split('@')[1] if '@' in user.email else '***'
        return Response({
            'message': 'OTP sent successfully',
            'maskedEmail': masked_email,
        })


class ConfidentialOTPVerifyView(APIView):
    @require_auth
    def post(self, request):
        from accounts.models import LoginVerificationCode
        user_id = request.session.get('userId')
        if not user_id:
            return Response({'message': 'Not authenticated'}, status=401)

        code = (request.data.get('code') or '').strip()
        if not code or len(code) != 6:
            return Response({'message': 'Invalid OTP format'}, status=400)

        code_hash = hashlib.sha256(code.encode()).hexdigest()
        try:
            otp_record = LoginVerificationCode.objects.filter(
                user_id=user_id, status='pending',
                expires_at__gt=timezone.now(),
            ).latest('created_at')
        except LoginVerificationCode.DoesNotExist:
            return Response({'message': 'No valid OTP found. Please request a new one.'}, status=400)

        if otp_record.attempts >= 5:
            otp_record.status = 'expired'
            otp_record.save(update_fields=['status'])
            return Response({'message': 'Too many attempts. Please request a new OTP.'}, status=429)

        otp_record.attempts += 1
        otp_record.save(update_fields=['attempts'])

        if otp_record.code_hash != code_hash:
            remaining = 5 - otp_record.attempts
            return Response({'message': f'Invalid OTP. {remaining} attempts remaining.'}, status=400)

        otp_record.status = 'used'
        otp_record.used_at = timezone.now()
        otp_record.save(update_fields=['status', 'used_at'])

        request.session['confidential_unlocked_at'] = timezone.now().isoformat()
        return Response({'message': 'Verified successfully', 'unlocked': True})


class MyExpensesView(APIView):
    @require_auth
    def get(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        qs = TravelExpense.objects.filter(employee_id=employee.id).order_by('-expense_date')
        return Response([{
            'id': str(e.id),
            'category': e.category,
            'description': e.description,
            'amount': float(e.amount),
            'expense_date': e.expense_date.isoformat() if e.expense_date else None,
            'receipt_url': get_presigned_url(e.receipt_url) if e.receipt_url else None,
            'status': e.status,
            'rejection_reason': e.rejection_reason,
        } for e in qs[:20]])

    @require_auth
    def post(self, request):
        user_id = request.session.get('userId')
        employee = get_employee_for_user(user_id)
        if not employee:
            return Response({'message': 'No employee profile linked'}, status=404)

        data = request.data
        category = data.get('category', 'travel')
        description = (data.get('description') or '').strip()
        amount = data.get('amount')
        expense_date_str = data.get('expense_date')
        receipt_url = data.get('receipt_url')

        if not description:
            return Response({'message': 'Description is required'}, status=400)
        if not amount or float(amount) <= 0:
            return Response({'message': 'Amount must be greater than 0'}, status=400)

        exp_date = date.fromisoformat(expense_date_str) if expense_date_str else date.today()

        te = TravelExpense.objects.create(
            employee_id=employee.id,
            category=category,
            description=description,
            amount=float(amount),
            expense_date=exp_date,
            receipt_url=receipt_url or None,
            month=exp_date.month,
            year=exp_date.year,
            status='pending',
        )

        return Response({
            'id': str(te.id),
            'category': te.category,
            'description': te.description,
            'amount': float(te.amount),
            'expense_date': te.expense_date.isoformat() if te.expense_date else None,
            'receipt_url': te.receipt_url,
            'status': te.status,
        }, status=201)


class DepartmentLeaveAllocationView(APIView):
    @require_permission('hrms.leave_type.read')
    def get(self, request, leave_type_id):
        allocations = DepartmentLeaveAllocation.objects.filter(
            leave_type_id=leave_type_id
        ).select_related('department')
        return Response([{
            'id': str(a.id),
            'department_id': str(a.department_id),
            'department_name': a.department.name,
            'allocated_days': float(a.allocated_days),
        } for a in allocations])

    @require_permission('hrms.leave_type.add')
    def post(self, request, leave_type_id):
        data = request.data
        dept_id = data.get('department_id')
        allocated_days = data.get('allocated_days')
        if not dept_id or allocated_days is None:
            return Response({'message': 'department_id and allocated_days required'}, status=400)

        da, created = DepartmentLeaveAllocation.objects.update_or_create(
            department_id=dept_id,
            leave_type_id=leave_type_id,
            defaults={'allocated_days': float(allocated_days)},
        )
        return Response({
            'id': str(da.id),
            'department_id': str(da.department_id),
            'allocated_days': float(da.allocated_days),
        }, status=201 if created else 200)

    @require_permission('hrms.leave_type.delete')
    def delete(self, request, leave_type_id):
        alloc_id = request.data.get('allocation_id') or request.GET.get('allocation_id')
        if alloc_id:
            DepartmentLeaveAllocation.objects.filter(id=alloc_id, leave_type_id=leave_type_id).delete()
        return Response({'message': 'Deleted'})


class HolidayBulkUploadView(APIView):
    @require_permission('hrms.holiday.add')
    def get(self, request):
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['organization_short_code', 'name', 'date', 'is_optional'])
        writer.writerow(['SIC', 'New Year', '2026-01-01', 'no'])
        writer.writerow(['SIC', 'Republic Day', '2026-05-28', 'no'])
        writer.writerow(['SIC', 'Optional Holiday', '2026-06-15', 'yes'])
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="holidays_template.csv"'
        return response

    @require_permission('hrms.holiday.add')
    def post(self, request):
        import csv
        import io

        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No file uploaded'}, status=400)

        fname = file.name.lower()
        rows = []

        if fname.endswith('.csv'):
            decoded = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                rows.append({k.strip().lower().replace(' ', '_'): (v.strip() if v else '') for k, v in row.items() if k})
        elif fname.endswith('.xlsx'):
            import openpyxl
            try:
                wb = openpyxl.load_workbook(file, read_only=True)
            except Exception:
                return Response({'message': 'Invalid Excel file. Please upload a valid .xlsx file.'}, status=400)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or '').strip().lower().replace(' ', '_') for c in row]
                    continue
                if not any(row):
                    continue
                rd = {}
                for j, val in enumerate(row):
                    if j < len(headers) and headers[j]:
                        rd[headers[j]] = str(val).strip() if val is not None else ''
                rows.append(rd)
        else:
            return Response({'message': 'Unsupported file format. Use CSV or Excel (.xlsx).'}, status=400)

        if not rows:
            return Response({'message': 'No data rows found in the file.'}, status=400)

        org_map = {}
        for org in Organization.objects.all():
            org_map[org.short_code.lower()] = org
            org_map[org.name.lower()] = org

        created = 0
        skipped = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            org_code = row.get('organization_short_code') or row.get('organization') or row.get('org_code') or ''
            name = row.get('name') or row.get('holiday_name') or ''
            date_str = row.get('date') or row.get('holiday_date') or ''
            is_opt = row.get('is_optional') or row.get('optional') or 'no'

            if not name or not date_str:
                errors.append(f'Row {i}: Missing name or date')
                continue

            org = org_map.get(org_code.lower())
            if not org:
                errors.append(f'Row {i}: Organization "{org_code}" not found')
                continue

            try:
                from datetime import datetime as dt
                parsed_date = None
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                    try:
                        parsed_date = dt.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if not parsed_date:
                    raise ValueError(f'Invalid date format: {date_str}')
            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')
                continue

            optional = is_opt.lower() in ('yes', 'true', '1', 'optional')

            if Holiday.objects.filter(organization=org, date=parsed_date, name=name).exists():
                skipped += 1
                continue

            Holiday.objects.create(
                organization=org,
                name=name,
                date=parsed_date,
                is_optional=optional,
            )
            created += 1

        return Response({
            'message': f'{created} holidays created, {skipped} duplicates skipped.',
            'created': created,
            'skipped': skipped,
            'errors': errors,
        })


class StaffBulkUploadView(APIView):
    @require_auth
    def get(self, request):
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'full_name', 'email', 'phone', 'position', 'organization_short_code',
            'department_name', 'gender', 'country', 'marital_status',
            'date_of_birth', 'join_date', 'employment_type',
            'citizenship_no', 'pan_no', 'passport_number', 'employee_id_number',
            'bank_name', 'bank_account_number', 'bank_branch',
            'permanent_address', 'temporary_address',
            'emergency_contact_name', 'emergency_contact_phone',
            'probation_end_date', 'contract_end_date',
            'salary_amount', 'salary_currency',
            'basic_salary', 'cit_type', 'cit_value',
            'ssf_applicable', 'tax_applicable',
        ])
        writer.writerow([
            'John Doe', 'john@example.com', '9801234567', 'Software Engineer', 'SIC',
            'Engineering', 'Male', 'Nepal', 'Single',
            '1995-01-15', '2025-01-01', 'full_time',
            'CTZN-12345', 'PAN-67890', '', 'EMP-001',
            'Nepal Bank', '1234567890', 'Kathmandu',
            'Kathmandu, Nepal', 'Lalitpur, Nepal',
            'Jane Doe', '9809876543',
            '', '',
            '50000', 'NPR',
            '40000', 'none', '0',
            'no', 'yes',
        ])
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="staff_template.csv"'
        return response

    @require_permission('hrms.staff.write')
    def post(self, request):
        import csv
        import io

        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No file uploaded'}, status=400)

        fname = file.name.lower()
        rows = []

        if fname.endswith('.csv'):
            decoded = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                rows.append({k.strip().lower().replace(' ', '_'): (v.strip() if v else '') for k, v in row.items() if k})
        elif fname.endswith('.xlsx'):
            import openpyxl
            try:
                wb = openpyxl.load_workbook(file, read_only=True)
            except Exception:
                return Response({'message': 'Invalid Excel file. Please upload a valid .xlsx file.'}, status=400)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or '').strip().lower().replace(' ', '_') for c in row]
                    continue
                if not any(row):
                    continue
                rd = {}
                for j, val in enumerate(row):
                    if j < len(headers) and headers[j]:
                        rd[headers[j]] = str(val).strip() if val is not None else ''
                rows.append(rd)
        else:
            return Response({'message': 'Unsupported file format. Use CSV or Excel (.xlsx).'}, status=400)

        if not rows:
            return Response({'message': 'No data rows found in the file.'}, status=400)

        org_map = {}
        for org in Organization.objects.all():
            org_map[org.short_code.lower()] = org
            org_map[org.name.lower()] = org

        dept_map = {}
        for dept in Department.objects.all():
            key = f"{str(dept.organization_id).lower()}_{dept.name.lower()}"
            dept_map[key] = dept

        def parse_date(val):
            if not val or val.lower() == 'none':
                return None
            from datetime import datetime as dt
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                try:
                    return dt.strptime(val, fmt).date()
                except ValueError:
                    continue
            return None

        created = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            full_name = row.get('full_name') or row.get('name') or ''
            email = row.get('email') or ''

            if not full_name or not email:
                errors.append(f'Row {i}: Missing full_name or email')
                continue

            org_code = row.get('organization_short_code') or row.get('organization') or row.get('org_code') or ''
            org = org_map.get(org_code.lower()) if org_code else None

            dept_name = row.get('department_name') or row.get('department') or ''
            dept = None
            if org and dept_name:
                dept_key = f"{str(org.id).lower()}_{dept_name.lower()}"
                dept = dept_map.get(dept_key)

            emp_data = {
                'full_name': full_name,
                'phone': row.get('phone') or None,
                'position': row.get('position') or None,
                'department': dept_name or None,
                'organization_id': org.id if org else None,
                'department_id': dept.id if dept else None,
                'gender': row.get('gender') or None,
                'country': row.get('country') or None,
                'marital_status': row.get('marital_status') or None,
                'date_of_birth': parse_date(row.get('date_of_birth')),
                'join_date': parse_date(row.get('join_date')),
                'employment_type': row.get('employment_type') or 'full_time',
                'citizenship_no': row.get('citizenship_no') or None,
                'pan_no': row.get('pan_no') or None,
                'passport_number': row.get('passport_number') or None,
                'employee_id_number': row.get('employee_id_number') or None,
                'bank_name': row.get('bank_name') or None,
                'bank_account_number': row.get('bank_account_number') or None,
                'bank_branch': row.get('bank_branch') or None,
                'permanent_address': row.get('permanent_address') or None,
                'temporary_address': row.get('temporary_address') or None,
                'emergency_contact_name': row.get('emergency_contact_name') or None,
                'emergency_contact_phone': row.get('emergency_contact_phone') or None,
                'probation_end_date': parse_date(row.get('probation_end_date')),
                'contract_end_date': parse_date(row.get('contract_end_date')),
                'salary_currency': row.get('salary_currency') or 'NPR',
            }

            salary_str = row.get('salary_amount') or ''
            if salary_str:
                try:
                    emp_data['salary_amount'] = Decimal(salary_str)
                except Exception:
                    pass

            try:
                emp, is_new = Employee.objects.get_or_create(
                    email=email,
                    defaults=emp_data,
                )
                if not is_new:
                    for k, v in emp_data.items():
                        if v is not None:
                            setattr(emp, k, v)
                    emp.save()
                    updated += 1
                else:
                    created += 1

                basic_str = row.get('basic_salary') or ''
                if basic_str:
                    try:
                        basic = Decimal(basic_str)
                        cit_type = row.get('cit_type') or 'none'
                        cit_value = Decimal(row.get('cit_value') or '0')
                        ssf_val = (row.get('ssf_applicable') or 'no').lower()
                        ssf_applicable = ssf_val in ('yes', 'true', '1')
                        tax_val = (row.get('tax_applicable') or 'yes').lower()
                        tax_applicable = tax_val in ('yes', 'true', '1')

                        existing_ss = SalaryStructure.objects.filter(employee_id=emp.id, status='active').first()
                        if existing_ss:
                            existing_ss.basic_salary = basic
                            existing_ss.cit_type = cit_type
                            existing_ss.cit_value = cit_value
                            existing_ss.ssf_applicable = ssf_applicable
                            existing_ss.tax_applicable = tax_applicable
                            existing_ss.save()
                        else:
                            SalaryStructure.objects.create(
                                employee_id=emp.id,
                                basic_salary=basic,
                                allowances={},
                                deductions={},
                                cit_type=cit_type,
                                cit_value=cit_value,
                                ssf_applicable=ssf_applicable,
                                ssf_employee_percentage=Decimal('11'),
                                ssf_employer_percentage=Decimal('20'),
                                tax_applicable=tax_applicable,
                                status='active',
                            )
                    except Exception:
                        pass

            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')

        return Response({
            'message': f'{created} employees created, {updated} updated.',
            'created': created,
            'updated': updated,
            'errors': errors,
        })
